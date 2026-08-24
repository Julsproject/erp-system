"""Physical Stock Count (cycle count): scan what's actually on the shelf and
reconcile against what the system thinks is on hand, with live variance —
instead of the manual scan-to-Excel-then-import round trip.

Only one count session is open at a time, store-wide, to keep this simple —
nothing scoped to a category/location yet. A count is delta-based: each
line snapshots system_qty the moment it's first scanned, and Complete
applies (counted - system_qty) as a signed stock movement, so it's still
correct even if sales happen elsewhere on the system while counting.
"""
import io
import json
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import audit, models
from .database import get_db
from .deps import get_current_user, is_floor_staff
from .products import ADJUSTMENT_REASON_LABELS, ADJUSTMENT_REASONS, _parse_upload
from .search_utils import multi_word_ilike
from .templating import templates

router = APIRouter()

PAGE_SIZE = 15

# --- Bulk import (Excel / CSV) — fill in a count sheet instead of scanning
# every item by hand. Columns read by header name, like the Products import.
# Anything NOT listed here is treated as a per-unit column named after one of
# the product's own units (Sack, Elf Load, ...), so the reference-only columns
# below must be mapped even though nothing reads their value — otherwise a
# filled "Unit" cell would be misread as a unit quantity.
HEADER_MAP_COUNT = {
    "barcode": "barcode", "bar code": "barcode", "upc": "barcode", "ean": "barcode",
    "product name": "name", "name": "name", "product": "name",
    "counted qty": "counted", "counted": "counted", "count": "counted", "qty": "counted", "quantity": "counted",
    "unit": "unit_label", "base unit": "unit_label",
    "system qty (reference only)": "system_ref", "system qty": "system_ref", "system": "system_ref",
}
FIELDS_COUNT = ["barcode", "name", "counted", "unit_label", "system_ref"]
BASE_HEADERS_COUNT = ["Barcode", "Product Name", "Unit", "System Qty (reference only)", "Counted Qty"]


# Cells the count sheet writes as "not applicable" for a unit this product
# doesn't have. They read as filled-in text, so they'd otherwise come back on
# re-import looking like a quantity.
_BLANK_CELLS = {"", "—", "–", "-", "n/a", "na"}


def _is_blank_cell(value) -> bool:
    return str(value or "").strip().lower() in _BLANK_CELLS


def _dec(value, default="0") -> Decimal:
    try:
        return Decimal(str(value).strip().replace(",", "") or default)
    except (InvalidOperation, AttributeError, ValueError):
        return Decimal(default)


def _find_products(db: Session, q: str):
    """Same priority as POS: an exact barcode match wins outright (that's
    what a scan produces); otherwise fall back to a name search so counting
    still works for products with no barcode on file."""
    q = (q or "").strip()
    if not q:
        return []
    barcode_hit = (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True), models.Product.barcode == q)
        .first()
    )
    if barcode_hit:
        return [barcode_hit]
    return (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True), multi_word_ilike(models.Product.name, q))
        .order_by(models.Product.name)
        .limit(15)
        .all()
    )


def _line_dict(line: models.StockCountLine) -> dict:
    variance = Decimal(str(line.counted_qty or 0)) - Decimal(str(line.system_qty or 0))
    try:
        breakdown = json.loads(line.unit_breakdown) if line.unit_breakdown else {}
    except (ValueError, TypeError):
        breakdown = {}
    product = line.product
    units = []
    if product:
        base_name = product.unit_type.name if product.unit_type else "unit"
        units.append({"name": base_name, "factor": 1.0, "qty": breakdown.get(base_name, "")})
        for u in product.units:
            units.append({"name": u.name, "factor": float(u.factor_to_base), "qty": breakdown.get(u.name, "")})
    return {
        "id": line.id,
        "product_id": line.product_id,
        "name": line.product_name,
        "system_qty": float(line.system_qty or 0),
        "counted_qty": float(line.counted_qty or 0),
        "variance": float(variance),
        "units": units if len(units) > 1 else [],  # only worth showing when there's an actual ladder
        "shelf": (product.shelf.name if product and product.shelf else None),
    }


def _classify_count_import_rows(db: Session, count: models.StockCount, numbered_rows):
    """Read-only pass: match each row to a product (barcode wins, else exact
    name), skipping rows with nothing counted (a count sheet is filled in
    gradually, so most rows may still be blank).

    A row can express its count either as a plain base-unit number in the
    Counted Qty column, or spread across the per-unit columns (Sack, Elf
    Load, ...) — or both, which simply add up. Every unit is converted to
    base units via factor_to_base, exactly like the per-unit entry in the
    scan UI, because stock is stored as one base-unit number."""
    existing_line_by_product = {l.product_id: l for l in count.lines}
    classified = []
    for line_no, record in numbered_rows:
        barcode = (record["barcode"] or "").strip()
        name = (record["name"] or "").strip()
        counted_raw = (record["counted"] or "").strip()
        extra = {k: v for k, v in (record.get("_extra") or {}).items() if not _is_blank_cell(v)}
        if not barcode and not name:
            continue
        if not counted_raw and not extra:
            classified.append({"line_no": line_no, "name": name or barcode, "action": "blank"})
            continue

        product = None
        if barcode:
            product = (
                db.query(models.Product)
                .filter(models.Product.is_active.is_(True), models.Product.barcode == barcode)
                .first()
            )
        if not product and name:
            product = (
                db.query(models.Product)
                .filter(models.Product.is_active.is_(True), func.lower(models.Product.name) == name.lower())
                .first()
            )
        if not product:
            classified.append({
                "line_no": line_no, "name": name or barcode, "action": "error",
                "message": f"No active product matches {'barcode ' + barcode if barcode else 'name “' + name + '”'}.",
            })
            continue

        base_name = product.unit_type.name if product.unit_type else "unit"
        factor_by_key = {base_name.strip().lower(): (base_name, Decimal("1"))}
        for u in product.units:
            factor_by_key[u.name.strip().lower()] = (u.name, Decimal(str(u.factor_to_base or 0)))

        total = Decimal("0")
        breakdown = {}
        err = None
        for raw_label, raw_val in [(base_name, counted_raw)] + list(extra.items()):
            val = str(raw_val or "").strip()
            if not val:
                continue
            hit = factor_by_key.get(raw_label.strip().lower())
            if not hit:
                # A per-unit column that this particular product doesn't have
                # is normal — the sheet carries one column per unit found
                # across ALL products, so most rows leave most of them blank.
                # A *filled* one, though, is a real mistake worth reporting.
                err = f"“{raw_label}” isn't a unit of {product.name}."
                break
            unit_name, factor = hit
            try:
                qty = Decimal(val.replace(",", ""))
            except InvalidOperation:
                err = f"“{val}” isn't a valid quantity for {unit_name}."
                break
            if qty < 0:
                err = f"Quantity for {unit_name} can't be negative."
                break
            total += qty * factor
            breakdown[unit_name] = val
        if err:
            classified.append({"line_no": line_no, "name": product.name, "action": "error", "message": err})
            continue

        classified.append({
            "line_no": line_no, "name": product.name, "action": "set",
            "product": product, "counted": total,
            # Only worth recording a breakdown when more than the base unit was
            # used — a plain number stays a plain number, same as a typed edit.
            "breakdown": breakdown if len(breakdown) > 1 or (breakdown and base_name not in breakdown) else None,
            "line": existing_line_by_product.get(product.id),
        })
    return classified


def _run_count_import(db: Session, count: models.StockCount, classified, assign_shelf=None):
    """Apply each matched row's Counted Qty as a line's new count — same
    effect as typing it into the Counted field by hand. A later row for the
    same product in the same file simply overwrites an earlier one.

    assign_shelf, if given, is set on any counted product that doesn't
    already have a shelf — never overwrites an existing assignment. Meant
    for exactly this workflow: download the count sheet for one shelf,
    walk it, count everything including items that were never assigned a
    shelf before, and have counting them also be what puts them on the map."""
    applied = skipped = shelved = 0
    errors = []
    lines_by_product = {}
    for item in classified:
        line_no, name, action = item["line_no"], item["name"], item["action"]
        if action == "blank":
            skipped += 1
            continue
        if action == "error":
            errors.append({"row": line_no, "name": name, "message": item["message"]})
            continue
        product = item["product"]
        line = lines_by_product.get(product.id) or item["line"]
        if not line:
            line = models.StockCountLine(
                stock_count_id=count.id, product_id=product.id, product_name=product.name,
                system_qty=Decimal(str(product.total_qty or 0)), counted_qty=Decimal("0"),
            )
            db.add(line)
        line.counted_qty = item["counted"]
        line.unit_breakdown = json.dumps(item["breakdown"]) if item.get("breakdown") else None
        lines_by_product[product.id] = line
        if assign_shelf and product.shelf_id is None:
            product.shelf = assign_shelf
            shelved += 1
        applied += 1
    db.commit()
    return {
        "applied": applied, "skipped": skipped, "errors": errors, "shelved": shelved,
        "assign_shelf_name": assign_shelf.name if assign_shelf else None,
        "total": len(classified),
    }


def _uncounted_negatives(db: Session, count: models.StockCount):
    """Active products sitting at negative on-hand that aren't in this count.

    Negative stock is normal while a backlog of past sales is being encoded
    against an inventory that was never opened — but by the time a count is
    completed, every one of those should have been physically counted and
    corrected. Anything still negative and *not* in the count is a product
    nobody put eyes on, and completing without it leaves that negative on the
    books indefinitely. Surfaced as a warning (not a block) — a shop may
    legitimately be counting only one shelf.
    """
    counted_ids = {
        pid for (pid,) in db.query(models.StockCountLine.product_id)
        .filter(models.StockCountLine.stock_count_id == count.id).all()
    }
    on_hand = func.coalesce(models.Product.beginning_stock, 0) + func.coalesce(models.Product.stock_qty, 0)
    q = (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True), on_hand < 0)
        .order_by(on_hand, models.Product.name)
    )
    if counted_ids:
        q = q.filter(~models.Product.id.in_(counted_ids))
    return q.all()


@router.get("/stock-count", response_class=HTMLResponse)
def stock_count_list(request: Request, page: int = 1, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    page = max(page, 1)
    open_count = db.query(models.StockCount).filter(models.StockCount.status == "open").first()
    past_q = db.query(models.StockCount).filter(models.StockCount.status != "open").order_by(models.StockCount.id.desc())
    total = past_q.count()
    pages = max((total + PAGE_SIZE - 1) // PAGE_SIZE, 1)
    page = min(page, pages)
    past = past_q.offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    return templates.TemplateResponse(
        "stock_count/list.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "open_count": open_count, "past": past, "page": page, "pages": pages, "total": total},
    )


@router.post("/stock-count/start")
def stock_count_start(db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    existing = db.query(models.StockCount).filter(models.StockCount.status == "open").first()
    if existing:
        return RedirectResponse(f"/stock-count/{existing.id}", status_code=302)
    count = models.StockCount(status="open", created_by=user.id)
    db.add(count)
    db.flush()
    count.ref_no = f"SC-{count.id:06d}"
    db.commit()
    return RedirectResponse(f"/stock-count/{count.id}", status_code=302)


@router.get("/stock-count/{count_id:int}", response_class=HTMLResponse)
def stock_count_view(count_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    count = db.get(models.StockCount, count_id)
    if not count:
        return RedirectResponse("/stock-count", status_code=302)
    line_rows = (
        db.query(models.StockCountLine)
        .filter(models.StockCountLine.stock_count_id == count.id)
        .order_by(models.StockCountLine.product_name)
        .all()
    )
    lines = [_line_dict(l) for l in line_rows]
    variance_count = sum(1 for l in lines if l["counted_qty"] != l["system_qty"])
    shelves = db.query(models.Shelf).order_by(models.Shelf.name).all()
    missing_negatives = _uncounted_negatives(db, count) if count.status == "open" else []
    return templates.TemplateResponse(
        "stock_count/session.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "count": count, "lines": lines, "variance_count": variance_count,
         "adjustment_reasons": ADJUSTMENT_REASONS, "shelves": shelves,
         "missing_negatives": missing_negatives},
    )


@router.post("/stock-count/{count_id:int}/scan")
async def stock_count_scan(count_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user or not is_floor_staff(user):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return JSONResponse({"ok": False, "error": "This count isn't open anymore."}, status_code=400)

    data = await request.json()
    add_qty = _dec(data.get("qty"), "1")

    # A disambiguation pick (from a previous ambiguous name search) sends the
    # product id directly instead of searching again.
    product_id = data.get("product_id")
    if product_id:
        product = db.get(models.Product, int(product_id))
        if not product or not product.is_active:
            return JSONResponse({"ok": False, "error": "That product isn't available anymore."}, status_code=404)
    else:
        q = (data.get("q") or "").strip()
        if not q:
            return JSONResponse({"ok": False, "error": "Scan or type something to search."}, status_code=400)
        matches = _find_products(db, q)
        if not matches:
            return JSONResponse({"ok": False, "error": f"No product matches “{q}”."}, status_code=404)
        if len(matches) > 1:
            return JSONResponse({"ok": False, "choices": [{"id": p.id, "name": p.name} for p in matches]})
        product = matches[0]
    line = (
        db.query(models.StockCountLine)
        .filter(models.StockCountLine.stock_count_id == count.id, models.StockCountLine.product_id == product.id)
        .first()
    )
    if not line:
        line = models.StockCountLine(
            stock_count_id=count.id, product_id=product.id, product_name=product.name,
            system_qty=Decimal(str(product.total_qty or 0)), counted_qty=Decimal("0"),
        )
        db.add(line)
    line.counted_qty = Decimal(str(line.counted_qty or 0)) + add_qty
    db.commit()
    return {"ok": True, "line": _line_dict(line)}


@router.post("/stock-count/{count_id:int}/add-shelf")
async def stock_count_add_shelf(count_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Bulk-add every product on one shelf as a line (counted_qty starts at
    0, same as a freshly-scanned line) — lets a count be worked shelf by
    shelf instead of hunting the alphabetical product list. Shelves are
    assigned ahead of time (product form or bulk import), not from here."""
    if not user or not is_floor_staff(user):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return JSONResponse({"ok": False, "error": "This count isn't open anymore."}, status_code=400)

    data = await request.json()
    try:
        shelf_id = int(data.get("shelf_id") or 0)
    except (TypeError, ValueError):
        shelf_id = 0
    if not shelf_id:
        return JSONResponse({"ok": False, "error": "Pick a shelf."}, status_code=400)
    products = (
        db.query(models.Product)
        .filter(models.Product.shelf_id == shelf_id, models.Product.is_active.is_(True))
        .all()
    )
    if not products:
        return JSONResponse({"ok": False, "error": "No products are assigned to that shelf yet."}, status_code=404)

    existing_ids = {
        pid for (pid,) in db.query(models.StockCountLine.product_id)
        .filter(models.StockCountLine.stock_count_id == count.id).all()
    }
    added = 0
    for product in products:
        if product.id in existing_ids:
            continue
        db.add(models.StockCountLine(
            stock_count_id=count.id, product_id=product.id, product_name=product.name,
            system_qty=Decimal(str(product.total_qty or 0)), counted_qty=Decimal("0"),
        ))
        added += 1
    db.commit()

    line_rows = (
        db.query(models.StockCountLine)
        .filter(models.StockCountLine.stock_count_id == count.id)
        .order_by(models.StockCountLine.product_name)
        .all()
    )
    return {"ok": True, "added": added, "lines": [_line_dict(l) for l in line_rows]}


@router.post("/stock-count/{count_id:int}/assign-shelf")
async def stock_count_assign_shelf(count_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Put every counted product that has no shelf yet onto one shelf.

    The same idea the count-sheet import already offers, but available while
    counting on screen — you're standing at the shelf, you pull in what's
    already assigned to it, you scan the few strays that were never mapped,
    and this is what puts those strays on the map. Like the import, it only
    fills blanks: a product already assigned to another shelf is left alone,
    so this can never quietly move stock somewhere it isn't."""
    if not user or not is_floor_staff(user):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return JSONResponse({"ok": False, "error": "This count isn't open anymore."}, status_code=400)

    data = await request.json()
    try:
        shelf_id = int(data.get("shelf_id") or 0)
    except (TypeError, ValueError):
        shelf_id = 0
    if not shelf_id:
        return JSONResponse({"ok": False, "error": "Pick a shelf."}, status_code=400)
    shelf = db.get(models.Shelf, shelf_id)
    if not shelf:
        return JSONResponse({"ok": False, "error": "That shelf no longer exists."}, status_code=404)

    assigned = []
    for line in count.lines:
        product = line.product
        if product and product.shelf_id is None:
            product.shelf = shelf
            assigned.append(line)
    db.commit()

    line_rows = (
        db.query(models.StockCountLine)
        .filter(models.StockCountLine.stock_count_id == count.id)
        .order_by(models.StockCountLine.product_name)
        .all()
    )
    return {
        "ok": True, "assigned": len(assigned), "shelf": shelf.name,
        "lines": [_line_dict(l) for l in line_rows],
    }


@router.post("/stock-count/{count_id:int}/add-negatives")
def stock_count_add_negatives(count_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Pull every still-negative uncounted product into this count in one go,
    so the shop can go count exactly those instead of hunting them down from
    the warning list by hand."""
    if not user or not is_floor_staff(user):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return JSONResponse({"ok": False, "error": "This count isn't open anymore."}, status_code=400)

    for product in _uncounted_negatives(db, count):
        db.add(models.StockCountLine(
            stock_count_id=count.id, product_id=product.id, product_name=product.name,
            system_qty=Decimal(str(product.total_qty or 0)), counted_qty=Decimal("0"),
        ))
    db.commit()
    return RedirectResponse(f"/stock-count/{count.id}", status_code=302)


@router.get("/stock-count/{count_id:int}/import", response_class=HTMLResponse)
def stock_count_import_form(count_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return RedirectResponse("/stock-count", status_code=302)
    shelves = db.query(models.Shelf).order_by(models.Shelf.name).all()
    return templates.TemplateResponse(
        "stock_count/import.html",
        {"request": request, "app_name": request.app.title, "user": user, "count": count,
         "shelves": shelves, "result": None},
    )


@router.post("/stock-count/{count_id:int}/import", response_class=HTMLResponse)
async def stock_count_import_upload(
    count_id: int, request: Request, file: UploadFile = File(...), assign_shelf_id: int = Form(0),
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return RedirectResponse("/stock-count", status_code=302)

    contents = await file.read()
    rows, error = _parse_upload(
        file.filename, contents,
        header_map=HEADER_MAP_COUNT, fields=FIELDS_COUNT,
        required_field="name", required_label="Product Name",
        keep_extra_columns=True,  # the per-unit columns are named per product
    )
    if error:
        return templates.TemplateResponse(
            "stock_count/import.html",
            {"request": request, "app_name": request.app.title, "user": user, "count": count,
             "result": {"error": error}},
        )

    assign_shelf = db.get(models.Shelf, assign_shelf_id) if assign_shelf_id else None
    classified = _classify_count_import_rows(db, count, list(enumerate(rows, start=2)))
    result = _run_count_import(db, count, classified, assign_shelf=assign_shelf)
    result["filename"] = file.filename
    if result["applied"]:
        summary = f"Stock count {count.ref_no}: bulk import from “{file.filename}” set {result['applied']} line(s)"
        if result["shelved"]:
            summary += f", assigned {result['shelved']} product(s) to shelf “{result['assign_shelf_name']}”"
        audit.record(
            db, user=user, request=request, action="stock_count", entity_type="stock_count",
            entity_id=count.id, entity_label=count.ref_no, summary=summary,
        )
        db.commit()
    return templates.TemplateResponse(
        "stock_count/import.html",
        {"request": request, "app_name": request.app.title, "user": user, "count": count, "result": result},
    )


@router.get("/stock-count/{count_id:int}/import/template")
def stock_count_import_template(count_id: int, shelf_id: int = 0, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    count = db.get(models.StockCount, count_id)
    if not count:
        return RedirectResponse("/stock-count", status_code=302)

    existing_by_product = {l.product_id: l for l in count.lines}
    query = db.query(models.Product).filter(models.Product.is_active.is_(True))
    shelf = db.get(models.Shelf, shelf_id) if shelf_id else None
    if shelf:
        query = query.filter(models.Product.shelf_id == shelf.id)
    products = query.order_by(models.Product.name).all()

    # One extra column per distinct unit found across the products being
    # exported, so a product sold by the Sack can be counted as "2 Sack"
    # instead of forcing the counter to work out 2 x 25 = 50 kg by hand.
    # A shop with no unit ladders gets the plain 5-column sheet unchanged.
    reserved = set(HEADER_MAP_COUNT)
    unit_headers, seen = [], set()
    for p in products:
        for u in p.units:
            label = (u.name or "").strip()
            key = label.lower()
            if label and key not in seen and key not in reserved:
                seen.add(key)
                unit_headers.append(label)
    unit_headers.sort()
    headers = BASE_HEADERS_COUNT + unit_headers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = shelf.name[:31] if shelf else "Count Sheet"
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    unit_fill = PatternFill("solid", fgColor="0F766E")  # per-unit columns read as a distinct group
    for i, cell in enumerate(ws[1]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = unit_fill if i >= len(BASE_HEADERS_COUNT) else header_fill

    for p in products:
        line = existing_by_product.get(p.id)
        base_name = p.unit_type.name if p.unit_type else "unit"
        try:
            breakdown = json.loads(line.unit_breakdown) if line and line.unit_breakdown else {}
        except (ValueError, TypeError):
            breakdown = {}
        # Pre-fill from however the count was already entered: a per-unit
        # breakdown goes back into its own columns, a plain number into
        # Counted Qty — so re-downloading mid-count shows the real state.
        if breakdown:
            base_cell = breakdown.get(base_name, "")
        else:
            base_cell = float(line.counted_qty) if line else ""
        row = [p.barcode or "", p.name, base_name, float(p.total_qty or 0), base_cell]
        own_units = {(u.name or "").strip().lower() for u in p.units}
        for label in unit_headers:
            key = label.lower()
            row.append(breakdown.get(label, "") if key in own_units else "—")
        ws.append(row)

    widths = [18, 32, 12, 22, 14] + [12] * len(unit_headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = f"_{shelf.name}" if shelf else ""
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{count.ref_no}{suffix}_count_sheet.xlsx"'},
    )


@router.get("/stock-count/{count_id:int}/search")
def stock_count_search(count_id: int, q: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Live typeahead as the scan box is typed into by hand (a barcode
    scanner still just types-and-Enters, unaffected by this) — same
    matching rules as an actual scan, via _find_products."""
    if not user or not is_floor_staff(user):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    matches = _find_products(db, q)
    return {"products": [{"id": p.id, "name": p.name} for p in matches]}


@router.post("/stock-count/{count_id:int}/line/{line_id:int}/set")
async def stock_count_set_line(count_id: int, line_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user or not is_floor_staff(user):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return JSONResponse({"ok": False, "error": "This count isn't open anymore."}, status_code=400)
    line = db.get(models.StockCountLine, line_id)
    if not line or line.stock_count_id != count.id:
        return JSONResponse({"ok": False, "error": "Line not found."}, status_code=404)

    data = await request.json()
    try:
        new_qty = Decimal(str(data.get("counted_qty", "")).strip().replace(",", ""))
    except InvalidOperation:
        return JSONResponse({"ok": False, "error": "Enter a valid quantity."}, status_code=400)
    if new_qty < 0:
        return JSONResponse({"ok": False, "error": "Quantity can't be negative."}, status_code=400)
    line.counted_qty = new_qty
    line.unit_breakdown = None  # a plain-number edit overrides any per-unit breakdown
    db.commit()
    return {"ok": True, "line": _line_dict(line)}


@router.post("/stock-count/{count_id:int}/line/{line_id:int}/set-units")
async def stock_count_set_line_units(count_id: int, line_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Per-unit count entry — e.g. '2 FORWARD, 3 Elf, 1 Elf 1/2 physically on
    the shelf' — resolved into the one counted_qty (base units) everything
    else reads, so this is purely a friendlier way to arrive at that number."""
    if not user or not is_floor_staff(user):
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return JSONResponse({"ok": False, "error": "This count isn't open anymore."}, status_code=400)
    line = db.get(models.StockCountLine, line_id)
    if not line or line.stock_count_id != count.id:
        return JSONResponse({"ok": False, "error": "Line not found."}, status_code=404)
    product = line.product
    if not product:
        return JSONResponse({"ok": False, "error": "Product not found."}, status_code=404)

    data = await request.json()
    raw = data.get("units") or {}
    if not isinstance(raw, dict):
        return JSONResponse({"ok": False, "error": "Invalid data."}, status_code=400)

    base_name = product.unit_type.name if product.unit_type else "unit"
    factor_by_name = {base_name: Decimal("1")}
    for u in product.units:
        factor_by_name[u.name] = Decimal(str(u.factor_to_base or 0))

    total = Decimal("0")
    breakdown = {}
    for name, val in raw.items():
        if name not in factor_by_name:
            continue
        qty = _dec(val, "0")
        if qty < 0:
            return JSONResponse({"ok": False, "error": f"Quantity for {name} can't be negative."}, status_code=400)
        if qty == 0 and str(val).strip() == "":
            continue  # skip untouched fields entirely, don't record a stray "0"
        breakdown[name] = str(val).strip()
        total += qty * factor_by_name[name]

    line.counted_qty = total
    line.unit_breakdown = json.dumps(breakdown) if breakdown else None
    db.commit()
    return {"ok": True, "line": _line_dict(line)}


@router.post("/stock-count/{count_id:int}/complete")
def stock_count_complete(count_id: int, request: Request, reason: str = Form("count_correction"),
                          db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    count = db.get(models.StockCount, count_id)
    if not count or count.status != "open":
        return RedirectResponse("/stock-count", status_code=302)

    reason_label = ADJUSTMENT_REASON_LABELS.get(reason, reason)
    for line in count.lines:
        variance = Decimal(str(line.counted_qty or 0)) - Decimal(str(line.system_qty or 0))
        if variance == 0:
            continue
        product = db.get(models.Product, line.product_id, with_for_update=True)
        if not product:
            continue
        before_qty = Decimal(str(product.stock_qty or 0))
        product.stock_qty = before_qty + variance
        unit_cost = Decimal(str(product.cost_price or 0))
        db.add(models.StockMovement(
            product_id=product.id, qty_base=variance, reason="stock_count", ref=count.ref_no,
            unit_cost=unit_cost, value=variance * unit_cost, note=reason_label,
        ))
        audit.record(
            db, user=user, request=request, action="stock_count", entity_type="product",
            entity_id=product.id, entity_label=product.name,
            summary=f"Stock count {count.ref_no}: counted {line.counted_qty:g}, system said {line.system_qty:g}",
            changes={"stock_qty": [str(before_qty), str(product.stock_qty)]},
        )

    count.status = "completed"
    count.completed_by = user.id
    count.completed_at = func.now()
    db.commit()
    return RedirectResponse(f"/stock-count/{count.id}", status_code=302)


@router.post("/stock-count/{count_id:int}/cancel")
def stock_count_cancel(count_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_floor_staff(user):
        return RedirectResponse("/pos", status_code=302)
    count = db.get(models.StockCount, count_id)
    if count and count.status == "open":
        count.status = "cancelled"
        count.completed_by = user.id
        count.completed_at = func.now()
        db.commit()
    return RedirectResponse("/stock-count", status_code=302)
