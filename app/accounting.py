"""Accounting: Chart of Accounts, the posting engine, and the reports that
read off it (General Ledger, Trial Balance).

Phase 1 only auto-posts Sales (see pos.py's _finalize_sale / void_sale).
Purchases, Expenses, and Banking are deliberately not wired yet — see the
accounting module plan for why, and the order they're meant to land in.
"""
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import models
from .database import get_db
from .deps import get_current_user, is_staff
from .templating import templates

router = APIRouter()

MANILA = ZoneInfo("Asia/Manila")
CENTS = Decimal("0.01")
ZERO = Decimal("0")

ACCOUNT_TYPES = [
    ("asset", "Asset"), ("liability", "Liability"), ("equity", "Equity"),
    ("revenue", "Revenue"), ("cost_of_sales", "Cost of Sales"), ("expense", "Expense"),
]


def _dec(value, default="0") -> Decimal:
    try:
        return Decimal(str(value).strip().replace(",", "") or default)
    except (InvalidOperation, AttributeError, ValueError):
        return Decimal(default)


def _money(value) -> Decimal:
    return _dec(value).quantize(CENTS, rounding=ROUND_HALF_UP)


def _today() -> date:
    return datetime.now(MANILA).date()


def _local_date(col):
    return func.date(func.timezone("Asia/Manila", col))


def _parse_date(s: str):
    try:
        return date.fromisoformat(s) if s else None
    except ValueError:
        return None


def _resolve_period(days: int, date_from: str, date_to: str):
    """Same custom-range-overrides-preset logic used across every other
    report module — duplicated per this codebase's convention rather than
    imported, so each module can evolve its own filters independently."""
    today = _today()
    df, dt = _parse_date(date_from), _parse_date(date_to)
    custom = bool(df and dt)
    if custom:
        if dt > today:
            dt = today
        if df > dt:
            df, dt = dt, df
        if (dt - df).days > 365:
            df = dt - timedelta(days=365)
        return df, dt, custom
    if days not in (7, 30, 90):
        days = 30
    return today - timedelta(days=days - 1), today, custom


# --------------------------------------------------------------------------- #
# Posting engine
# --------------------------------------------------------------------------- #
class PostingError(Exception):
    pass


def _resolve_mapping(db: Session, function_key: str) -> models.Account:
    mapping = db.query(models.AccountMapping).filter(models.AccountMapping.function_key == function_key).first()
    if not mapping:
        raise PostingError(f"No account mapped for “{function_key}”. Set it up in Accounting Setup first.")
    if not mapping.account.is_active:
        raise PostingError(f"The account mapped to “{function_key}” ({mapping.account.name}) is inactive.")
    return mapping.account


def _next_journal_no(db: Session) -> str:
    # Row-lock the table via a real row (the max existing entry) so two
    # concurrent checkouts on different terminals can't both compute the
    # same next number — same discipline _finalize_sale already uses for
    # stock via with_for_update.
    last = (
        db.query(models.JournalEntry)
        .order_by(models.JournalEntry.id.desc())
        .with_for_update()
        .first()
    )
    next_n = (last.id + 1) if last else 1
    return f"JE-{next_n:06d}"


def post_journal(db: Session, *, txn_date, source_type: str, source_id, description: str,
                  lines: list, reference_no: str = None, entered_by_id: int = None, status: str = "posted") -> models.JournalEntry:
    """lines: [{"function_key": "SALE_CASH", "amount": Decimal, "side": "debit"|"credit", "memo": "..."}, ...]
    A line can give "_account_id" directly instead of "function_key" — for
    the handful of cases (like an Expense's per-category account) where the
    target account is a real FK on the source record, not a fixed function
    the whole shop shares one mapping for. The manual Journal Entry screen
    uses "_account_id" for every line and passes status="draft" so it lands
    in the books only once someone approves it (see post_draft_entry) — every
    balance-summing query in this module excludes status="draft" for exactly
    this reason.
    Does NOT commit — caller commits once, so the source transaction and its
    journal entry land or roll back together."""
    resolved = []
    total_debit = ZERO
    total_credit = ZERO
    for ln in lines:
        amount = _money(ln["amount"])
        if amount <= 0:
            continue
        if ln.get("_account_id"):
            account = db.get(models.Account, ln["_account_id"])
            if not account or not account.is_active:
                raise PostingError(f"Account #{ln['_account_id']} is missing or inactive.")
        else:
            account = _resolve_mapping(db, ln["function_key"])
        debit = amount if ln["side"] == "debit" else ZERO
        credit = amount if ln["side"] == "credit" else ZERO
        total_debit += debit
        total_credit += credit
        resolved.append({"account_id": account.id, "debit": debit, "credit": credit, "memo": ln.get("memo")})

    if not resolved:
        raise PostingError("Nothing to post — every line was zero.")
    if total_debit != total_credit:
        raise PostingError(f"Journal entry doesn't balance: debit {total_debit} vs credit {total_credit}.")

    entry = models.JournalEntry(
        journal_no=_next_journal_no(db), txn_date=txn_date, reference_no=reference_no,
        source_type=source_type, source_id=source_id, description=description,
        entered_by_id=entered_by_id, status=status,
    )
    db.add(entry)
    db.flush()  # need entry.id for the lines
    for ln in resolved:
        db.add(models.JournalLine(entry_id=entry.id, **ln))
    return entry


def post_draft_entry(db: Session, entry: models.JournalEntry) -> bool:
    """The 'approve' step of the manual Journal Entry workflow — flips a
    draft into the books. Returns False (no-op) if it wasn't a draft."""
    if entry.status != "draft":
        return False
    entry.status = "posted"
    return True


def delete_draft_entry(db: Session, entry: models.JournalEntry) -> bool:
    """Drafts never affected any balance, so unlike a posted entry they can
    just be deleted outright — no reversal needed. Returns False (no-op) if
    it wasn't a draft."""
    if entry.status != "draft":
        return False
    for line in list(entry.lines):
        db.delete(line)
    db.delete(entry)
    return True


def reverse_journal(db: Session, entry: models.JournalEntry, *, reason: str = None, entered_by_id: int = None) -> models.JournalEntry:
    """Insert a mirror-image entry (every debit becomes a credit and vice
    versa) and mark the original reversed. Never mutates or deletes the
    original entry's lines."""
    reversal = models.JournalEntry(
        journal_no=_next_journal_no(db), txn_date=_today(), reference_no=entry.reference_no,
        source_type="reversal", source_id=entry.id,
        description=f"Reversal of {entry.journal_no}" + (f": {reason}" if reason else ""),
        is_reversal_of_id=entry.id, entered_by_id=entered_by_id,
    )
    db.add(reversal)
    db.flush()
    for line in entry.lines:
        db.add(models.JournalLine(
            entry_id=reversal.id, account_id=line.account_id,
            debit=line.credit, credit=line.debit, memo=line.memo,
        ))
    entry.status = "reversed"
    return reversal


SALE_FUNCTION_KEYS = {
    "cash": "SALE_CASH", "gcash": "SALE_GCASH", "maya": "SALE_MAYA", "other_ewallet": "SALE_OTHER_EWALLET",
    "card": "SALE_CARD", "bank_transfer": "SALE_BANK_TRANSFER",
    "cheque": "SALE_BANK_TRANSFER",  # a cheque clears into the bank — reuse that mapping rather than add a new one
}


def post_sale(db: Session, sale: models.Sale, *, method_rows: list, receivable_amount: Decimal, entered_by_id: int = None):
    """Called from pos.py's _finalize_sale right after sale.id is flushed.
    Posts revenue + VAT + whatever the customer actually paid with — no
    COGS/Inventory leg yet (see the accounting plan: that waits on Purchases
    posting, which is what actually builds up the Inventory account)."""
    lines = []
    for method, amount in method_rows:
        if method in ("receivable", "cheque"):
            continue  # folded into the single AR line below instead, to avoid one JE line per split
        fkey = SALE_FUNCTION_KEYS.get(method)
        if not fkey:
            continue
        lines.append({"function_key": fkey, "amount": amount, "side": "debit", "memo": method})
    if receivable_amount and receivable_amount > 0:
        lines.append({"function_key": "AR", "amount": receivable_amount, "side": "debit", "memo": "credit/cheque"})

    net = Decimal(str(sale.net_amount or 0))
    vat = Decimal(str(sale.vat_amount or 0))
    if net > 0:
        lines.append({"function_key": "SALES_REVENUE", "amount": net, "side": "credit"})
    if vat > 0:
        lines.append({"function_key": "OUTPUT_VAT", "amount": vat, "side": "credit"})

    if not lines:
        return None
    txn_date = sale.created_at.date() if sale.created_at else _today()
    return post_journal(
        db, txn_date=txn_date, source_type="sale", source_id=sale.id,
        description=f"Sale {sale.invoice_no}", reference_no=sale.invoice_no,
        lines=lines, entered_by_id=entered_by_id,
    )


def reverse_sale_posting(db: Session, sale: models.Sale, *, reason: str, entered_by_id: int = None):
    """Called from pos.py's void_sale. No-op (not an error) if the sale
    predates Phase 1 and never got a journal entry in the first place."""
    entry = (
        db.query(models.JournalEntry)
        .filter(models.JournalEntry.source_type == "sale", models.JournalEntry.source_id == sale.id,
                models.JournalEntry.status == "posted")
        .first()
    )
    if not entry:
        return None
    return reverse_journal(db, entry, reason=reason, entered_by_id=entered_by_id)


def post_receivable_settlement(db: Session, sale: models.Sale, *, amount: Decimal, method: str, entered_by_id: int = None):
    """Called from sales.py's settle_pay and credits.py's pay_full_submit —
    a customer paying down what they owe on a credit sale. Dr the mapped
    money account, Cr Accounts Receivable. The cheque branch in both
    callers defers to a PDC and posts nothing yet, same as everywhere else
    a cheque appears in this app."""
    fkey = SALE_FUNCTION_KEYS.get(method)
    if not fkey:
        return None  # unrecognized method — nothing sane to post
    lines = [
        {"function_key": fkey, "amount": amount, "side": "debit", "memo": method},
        {"function_key": "AR", "amount": amount, "side": "credit"},
    ]
    return post_journal(
        db, txn_date=_today(), source_type="sale_settlement", source_id=sale.id,
        description=f"Payment on {sale.invoice_no}", reference_no=sale.invoice_no,
        lines=lines, entered_by_id=entered_by_id,
    )


PURCHASE_PAY_FUNCTION_KEYS = {
    "cash": "PURCHASE_PAY_CASH", "gcash": "PURCHASE_PAY_GCASH",
    "maya": "PURCHASE_PAY_MAYA", "other_ewallet": "PURCHASE_PAY_OTHER_EWALLET",
    "bank_transfer": "PURCHASE_PAY_BANK_TRANSFER", "cheque": "PURCHASE_PAY_CHEQUE", "other": "PURCHASE_PAY_OTHER",
}


def post_purchase_receive(db: Session, purchase: models.Purchase, *, is_payable: bool, payment_method: str = None, entered_by_id: int = None):
    """Called from purchases.py's create_purchase for a receive. Dr Inventory
    always; Cr Accounts Payable if left unpaid, otherwise Cr whichever money
    account the payment method maps to (a cheque here is treated the same
    way the operational Purchase record already treats it — settled at
    receive time, not deferred like a Sales cheque is via receivable)."""
    total = Decimal(str(purchase.total or 0))
    if total <= 0:
        return None
    credit_key = "AP" if is_payable else PURCHASE_PAY_FUNCTION_KEYS.get(payment_method, "PURCHASE_PAY_OTHER")
    lines = [
        {"function_key": "INVENTORY_MERCHANDISE", "amount": total, "side": "debit"},
        {"function_key": credit_key, "amount": total, "side": "credit", "memo": payment_method or "payable"},
    ]
    txn_date = purchase.created_at.date() if purchase.created_at else _today()
    return post_journal(
        db, txn_date=txn_date, source_type="purchase", source_id=purchase.id,
        description=f"Receive {purchase.ref_no}", reference_no=purchase.ref_no,
        lines=lines, entered_by_id=entered_by_id,
    )


def post_purchase_return(db: Session, purchase: models.Purchase, *, entered_by_id: int = None):
    """A return is posted as a straight vendor credit (reduces what's owed)
    regardless of how the original delivery was paid — the common
    small-business treatment. If the supplier actually refunds cash instead
    of issuing a credit, that needs a manual adjustment (Phase 2+'s manual
    Journal Entry UI, not yet built)."""
    total = Decimal(str(purchase.total or 0))
    if total <= 0:
        return None
    lines = [
        {"function_key": "AP", "amount": total, "side": "debit"},
        {"function_key": "INVENTORY_MERCHANDISE", "amount": total, "side": "credit"},
    ]
    txn_date = purchase.created_at.date() if purchase.created_at else _today()
    return post_journal(
        db, txn_date=txn_date, source_type="purchase", source_id=purchase.id,
        description=f"Return {purchase.ref_no}", reference_no=purchase.ref_no,
        lines=lines, entered_by_id=entered_by_id,
    )


def post_purchase_settlement(db: Session, purchase: models.Purchase, *, amount: Decimal, method: str, entered_by_id: int = None):
    """Called from purchases.py's settle_purchase_pay, only on the branch
    that actually records a PurchaseSettlement (the cheque branch defers to
    a PDC instead and posts nothing yet, same idea as Sales)."""
    credit_key = PURCHASE_PAY_FUNCTION_KEYS.get(method, "PURCHASE_PAY_OTHER")
    lines = [
        {"function_key": "AP", "amount": amount, "side": "debit"},
        {"function_key": credit_key, "amount": amount, "side": "credit", "memo": method},
    ]
    return post_journal(
        db, txn_date=_today(), source_type="purchase_settlement", source_id=purchase.id,
        description=f"Payment on {purchase.ref_no}", reference_no=purchase.ref_no,
        lines=lines, entered_by_id=entered_by_id,
    )


def reverse_purchase_posting(db: Session, purchase: models.Purchase, *, reason: str, entered_by_id: int = None):
    """Called from purchases.py's cancel_purchase. Reverses only the
    original receive/return entry (source_type="purchase") — NOT any
    PurchaseSettlement postings already made against it. Cancelling a
    purchase that's already been partially paid down is a pre-existing gap
    in the operational code too (cancel_purchase doesn't block on it), so
    this mirrors that rather than silently fixing scope beyond Phase 2."""
    entry = (
        db.query(models.JournalEntry)
        .filter(models.JournalEntry.source_type == "purchase", models.JournalEntry.source_id == purchase.id,
                models.JournalEntry.status == "posted")
        .first()
    )
    if not entry:
        return None
    return reverse_journal(db, entry, reason=reason, entered_by_id=entered_by_id)


EXPENSE_PAY_FUNCTION_KEYS = {
    "cash": "EXPENSE_PAY_CASH", "gcash": "EXPENSE_PAY_GCASH",
    "maya": "EXPENSE_PAY_MAYA", "other_ewallet": "EXPENSE_PAY_OTHER_EWALLET",
    "bank_transfer": "EXPENSE_PAY_BANK_TRANSFER", "cheque": "EXPENSE_PAY_CHEQUE",
}


def post_expense(db: Session, expense: models.Expense, *, entered_by_id: int = None):
    """Called from expenses.py's create_expense. Dr the expense category's
    own mapped account (falls back to EXPENSE_DEFAULT if the category has
    none set) plus an Input VAT line if vat_amount > 0; Cr whichever money
    account the payment method maps to — or, if the expense points at a
    specific paid_from_account (e.g. which Petty Cash box), credit that
    account's own gl_account_id directly instead of the generic mapping."""
    amount = Decimal(str(expense.amount or 0))
    if amount <= 0:
        return None
    vat = Decimal(str(expense.vat_amount or 0))
    net = amount - vat
    lines = []
    if expense.category and expense.category.account_id and expense.category.account.is_active:
        lines.append({"function_key": None, "amount": net, "side": "debit",
                       "_account_id": expense.category.account_id, "memo": expense.category.name})
    else:
        lines.append({"function_key": "EXPENSE_DEFAULT", "amount": net, "side": "debit"})
    if vat > 0:
        lines.append({"function_key": "INPUT_VAT", "amount": vat, "side": "debit"})
    paid_from = expense.paid_from_account
    if paid_from and paid_from.gl_account_id:
        lines.append({"function_key": None, "amount": amount, "side": "credit",
                       "_account_id": paid_from.gl_account_id, "memo": paid_from.name})
    else:
        pay_key = EXPENSE_PAY_FUNCTION_KEYS.get(expense.payment_method, "EXPENSE_PAY_CASH")
        lines.append({"function_key": pay_key, "amount": amount, "side": "credit", "memo": expense.payment_method})

    return post_journal(
        db, txn_date=expense.expense_date or _today(), source_type="expense", source_id=expense.id,
        description=f"Expense {expense.ref_no}" + (f" — {expense.payee}" if expense.payee else ""),
        reference_no=expense.reference_no, lines=lines, entered_by_id=entered_by_id,
    )


def reverse_expense_posting(db: Session, expense: models.Expense, *, reason: str, entered_by_id: int = None):
    """Called from expenses.py's void_expense. No-op if this expense predates
    Phase 3 and never got a journal entry."""
    entry = (
        db.query(models.JournalEntry)
        .filter(models.JournalEntry.source_type == "expense", models.JournalEntry.source_id == expense.id,
                models.JournalEntry.status == "posted")
        .first()
    )
    if not entry:
        return None
    return reverse_journal(db, entry, reason=reason, entered_by_id=entered_by_id)


def post_bank_transaction(db: Session, txn: models.BankTransaction, *, entered_by_id: int = None):
    """Called from banking.py's create_transaction. Only posts if BOTH the
    BankAccount has a gl_account_id AND the transaction was given a contra
    account — most deposits/withdrawals happen through Sales/Purchases/
    Expenses instead (which already post themselves via their own account);
    this is for the rest (capital injections, inter-account transfers,
    bank fees/interest not otherwise recorded)."""
    account = txn.account
    if not account or not account.gl_account_id or not txn.contra_account_id:
        return None
    amount = Decimal(str(txn.amount or 0))
    if amount <= 0:
        return None
    if txn.txn_type == "deposit":
        lines = [
            {"function_key": None, "_account_id": account.gl_account_id, "amount": amount, "side": "debit"},
            {"function_key": None, "_account_id": txn.contra_account_id, "amount": amount, "side": "credit"},
        ]
    else:
        lines = [
            {"function_key": None, "_account_id": txn.contra_account_id, "amount": amount, "side": "debit"},
            {"function_key": None, "_account_id": account.gl_account_id, "amount": amount, "side": "credit"},
        ]
    return post_journal(
        db, txn_date=txn.txn_date, source_type="bank_transaction", source_id=txn.id,
        description=(txn.description or f"{txn.txn_type.title()} — {account.name}"),
        reference_no=txn.reference_no, lines=lines, entered_by_id=entered_by_id,
    )


def reverse_bank_transaction_posting(db: Session, txn: models.BankTransaction, *, reason: str, entered_by_id: int = None):
    """Called from banking.py's void_transaction. No-op if this transaction
    was never posted (no gl_account_id/contra_account_id set at the time)."""
    entry = (
        db.query(models.JournalEntry)
        .filter(models.JournalEntry.source_type == "bank_transaction", models.JournalEntry.source_id == txn.id,
                models.JournalEntry.status == "posted")
        .first()
    )
    if not entry:
        return None
    return reverse_journal(db, entry, reason=reason, entered_by_id=entered_by_id)


# --------------------------------------------------------------------------- #
# Chart of Accounts
# --------------------------------------------------------------------------- #
@router.get("/accounting/coa", response_class=HTMLResponse)
def coa_list(request: Request, error: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    accounts = db.query(models.Account).order_by(models.Account.code).all()
    return templates.TemplateResponse(
        "accounting/coa.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "accounts": accounts, "account_types": ACCOUNT_TYPES, "error": error},
    )


@router.post("/accounting/coa")
def coa_create(
    code: str = Form(...), name: str = Form(...), account_type: str = Form(...),
    normal_balance: str = Form(...), parent_id: str = Form(""),
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    code = (code or "").strip()
    name = (name or "").strip()
    if not code or not name:
        return RedirectResponse("/accounting/coa?error=Code+and+name+are+required.", status_code=302)
    if db.query(models.Account).filter(models.Account.code == code).first():
        return RedirectResponse(f"/accounting/coa?error=Code+%E2%80%9C{code}%E2%80%9D+is+already+used.", status_code=302)
    account = models.Account(
        code=code, name=name, account_type=account_type, normal_balance=normal_balance,
        parent_id=int(parent_id) if parent_id else None, is_system=False,
    )
    db.add(account)
    db.commit()
    return RedirectResponse("/accounting/coa", status_code=302)


@router.post("/accounting/coa/{account_id:int}/rename")
def coa_rename(account_id: int, name: str = Form(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.Account, account_id)
    name = (name or "").strip()
    if account and name:
        account.name = name
        db.commit()
    return RedirectResponse("/accounting/coa", status_code=302)


@router.post("/accounting/coa/{account_id:int}/toggle")
def coa_toggle(account_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.Account, account_id)
    if not account:
        return RedirectResponse("/accounting/coa", status_code=302)
    if account.is_active:
        if account.is_system:
            return RedirectResponse(
                f"/accounting/coa?error=%E2%80%9C{account.name}%E2%80%9D+is+a+system+account+and+can%E2%80%99t+be+deactivated.",
                status_code=302,
            )
        still_mapped = db.query(models.AccountMapping.id).filter(models.AccountMapping.account_id == account.id).first()
        if still_mapped:
            return RedirectResponse(
                f"/accounting/coa?error=%E2%80%9C{account.name}%E2%80%9D+is+still+used+in+Accounting+Setup+%E2%80%94+remap+that+first.",
                status_code=302,
            )
    account.is_active = not account.is_active
    db.commit()
    return RedirectResponse("/accounting/coa", status_code=302)


# --------------------------------------------------------------------------- #
# Accounting Setup (function -> account mapping)
# --------------------------------------------------------------------------- #
@router.get("/accounting/mappings", response_class=HTMLResponse)
def mappings_list(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    mappings = db.query(models.AccountMapping).order_by(models.AccountMapping.function_key).all()
    accounts = db.query(models.Account).filter(models.Account.is_active.is_(True)).order_by(models.Account.code).all()
    expense_accounts = [a for a in accounts if a.account_type == "expense"]
    expense_categories = db.query(models.ExpenseCategory).order_by(models.ExpenseCategory.name).all()
    return templates.TemplateResponse(
        "accounting/mappings.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "mappings": mappings, "accounts": accounts,
         "expense_accounts": expense_accounts, "expense_categories": expense_categories},
    )


@router.post("/accounting/expense-category-account/{category_id:int}")
def expense_category_account_update(category_id: int, account_id: str = Form(""), db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    category = db.get(models.ExpenseCategory, category_id)
    if category:
        category.account_id = int(account_id) if account_id else None
        db.commit()
    return RedirectResponse("/accounting/mappings", status_code=302)


@router.post("/accounting/mappings/{mapping_id:int}")
def mappings_update(mapping_id: int, account_id: int = Form(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    mapping = db.get(models.AccountMapping, mapping_id)
    account = db.get(models.Account, account_id)
    if mapping and account and account.is_active:
        mapping.account_id = account.id
        db.commit()
    return RedirectResponse("/accounting/mappings", status_code=302)


# --------------------------------------------------------------------------- #
# General Ledger
# --------------------------------------------------------------------------- #
def _account_balance_before(db: Session, account: models.Account, before: date) -> Decimal:
    row = (
        db.query(func.coalesce(func.sum(models.JournalLine.debit), 0), func.coalesce(func.sum(models.JournalLine.credit), 0))
        .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
        .filter(models.JournalLine.account_id == account.id, models.JournalEntry.txn_date < before,
                models.JournalEntry.status != "draft")
        .one()
    )
    debit, credit = Decimal(str(row[0] or 0)), Decimal(str(row[1] or 0))
    return (debit - credit) if account.normal_balance == "debit" else (credit - debit)


@router.get("/accounting/general-ledger", response_class=HTMLResponse)
def general_ledger(
    request: Request, account_id: int = 0, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    accounts = db.query(models.Account).order_by(models.Account.code).all()
    account = db.get(models.Account, account_id) if account_id else (accounts[0] if accounts else None)

    rows = []
    opening = closing = ZERO
    if account:
        opening = _account_balance_before(db, account, period_start)
        entries = (
            db.query(models.JournalLine, models.JournalEntry)
            .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
            .filter(models.JournalLine.account_id == account.id,
                    models.JournalEntry.txn_date.between(period_start, period_end),
                    models.JournalEntry.status != "draft")
            .order_by(models.JournalEntry.txn_date, models.JournalEntry.id)
            .all()
        )
        running = opening
        for line, entry in entries:
            delta = (line.debit - line.credit) if account.normal_balance == "debit" else (line.credit - line.debit)
            running += delta
            rows.append({"entry": entry, "line": line, "balance": running})
        closing = running

    return templates.TemplateResponse(
        "accounting/general_ledger.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "accounts": accounts, "account": account, "rows": rows, "opening": opening, "closing": closing,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


# --------------------------------------------------------------------------- #
# Trial Balance
# --------------------------------------------------------------------------- #
def _trial_balance_rows(db: Session, as_of: date):
    accounts = db.query(models.Account).filter(models.Account.is_active.is_(True)).order_by(models.Account.code).all()
    rows = []
    total_debit = total_credit = ZERO
    for account in accounts:
        totals = (
            db.query(func.coalesce(func.sum(models.JournalLine.debit), 0), func.coalesce(func.sum(models.JournalLine.credit), 0))
            .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
            .filter(models.JournalLine.account_id == account.id, models.JournalEntry.txn_date <= as_of,
                    models.JournalEntry.status != "draft")
            .one()
        )
        debit, credit = Decimal(str(totals[0] or 0)), Decimal(str(totals[1] or 0))
        if debit == 0 and credit == 0:
            continue
        net = debit - credit
        row_debit = net if net > 0 else ZERO
        row_credit = -net if net < 0 else ZERO
        rows.append({"account": account, "debit": row_debit, "credit": row_credit})
        total_debit += row_debit
        total_credit += row_credit
    return rows, total_debit, total_credit


@router.get("/accounting/trial-balance", response_class=HTMLResponse)
def trial_balance(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    rows, total_debit, total_credit = _trial_balance_rows(db, period_end)
    return templates.TemplateResponse(
        "accounting/trial_balance.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "rows": rows, "total_debit": total_debit, "total_credit": total_credit,
         "balanced": total_debit == total_credit,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


@router.get("/accounting/trial-balance/export")
def export_trial_balance(days: int = 30, date_from: str = "", date_to: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, _ = _resolve_period(days, date_from, date_to)
    rows, total_debit, total_credit = _trial_balance_rows(db, period_end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.append(["Account Code", "Account Name", "Debit", "Credit"])
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in rows:
        ws.append([r["account"].code, r["account"].name, float(r["debit"]), float(r["credit"])])
    ws.append(["", "Total", float(total_debit), float(total_credit)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    widths = [14, 32, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"trial_balance_{period_end.isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --------------------------------------------------------------------------- #
# Cash Book / Bank Ledger — General Ledger, filtered to cash-like accounts
# --------------------------------------------------------------------------- #
def _cash_accounts(db: Session):
    linked_ids = [
        aid for (aid,) in db.query(models.BankAccount.gl_account_id)
        .filter(models.BankAccount.gl_account_id.isnot(None)).all()
    ]
    return (
        db.query(models.Account)
        .filter(
            models.Account.is_active.is_(True),
            (models.Account.system_key.in_(("CASH_ON_HAND", "GCASH", "MAYA", "OTHER_EWALLET", "BANK")) | models.Account.id.in_(linked_ids or [0])),
        )
        .order_by(models.Account.code)
        .all()
    )


@router.get("/accounting/cash-book", response_class=HTMLResponse)
def cash_book(
    request: Request, account_id: int = 0, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    accounts = _cash_accounts(db)

    # account_id == -1 means "All accounts" — the unified Cash Ledger
    # (Phase 5): every cash-type account's movement combined into one
    # chronological, running-balance table instead of one at a time.
    all_accounts = account_id == -1
    account = None
    if not all_accounts:
        account = db.get(models.Account, account_id) if account_id else (accounts[0] if accounts else None)
        if account and account not in accounts:
            account = accounts[0] if accounts else None

    rows = []
    opening = closing = ZERO
    if all_accounts and accounts:
        account_ids = [a.id for a in accounts]
        by_id = {a.id: a for a in accounts}
        for a in accounts:
            opening += _account_balance_before(db, a, period_start)
        entries = (
            db.query(models.JournalLine, models.JournalEntry)
            .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
            .filter(models.JournalLine.account_id.in_(account_ids),
                    models.JournalEntry.txn_date.between(period_start, period_end),
                    models.JournalEntry.status != "draft")
            .order_by(models.JournalEntry.txn_date, models.JournalEntry.id)
            .all()
        )
        running = opening
        for line, entry in entries:
            a = by_id[line.account_id]
            delta = (line.debit - line.credit) if a.normal_balance == "debit" else (line.credit - line.debit)
            running += delta
            rows.append({"entry": entry, "line": line, "account": a, "balance": running})
        closing = running
    elif account:
        opening = _account_balance_before(db, account, period_start)
        entries = (
            db.query(models.JournalLine, models.JournalEntry)
            .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
            .filter(models.JournalLine.account_id == account.id,
                    models.JournalEntry.txn_date.between(period_start, period_end),
                    models.JournalEntry.status != "draft")
            .order_by(models.JournalEntry.txn_date, models.JournalEntry.id)
            .all()
        )
        running = opening
        for line, entry in entries:
            delta = (line.debit - line.credit) if account.normal_balance == "debit" else (line.credit - line.debit)
            running += delta
            rows.append({"entry": entry, "line": line, "account": account, "balance": running})
        closing = running

    return templates.TemplateResponse(
        "accounting/cash_book.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "accounts": accounts, "account": account, "all_accounts": all_accounts,
         "rows": rows, "opening": opening, "closing": closing,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


# --------------------------------------------------------------------------- #
# Cash Flow Statement — cash-account movement grouped by what the OTHER
# side of each entry was, into Inflows vs Outflows.
# --------------------------------------------------------------------------- #
def _classify_counter(account: models.Account) -> str:
    """Labels the "why" of a cash movement from the account on the other
    side of the entry. system_key first (more precise than account_type
    for AR/AP, which are still technically 'asset'/'liability'), then
    account_type as a catch-all."""
    key = account.system_key or ""
    if key == "AR":
        return "Accounts Receivable Collections"
    if key == "AP":
        return "Supplier Payments"
    if key in ("SALES_REVENUE",):
        return "Sales"
    if key in ("OWNERS_EQUITY", "RETAINED_EARNINGS"):
        return "Owner's Equity"
    if account.account_type == "revenue":
        return "Other Income"
    if account.account_type == "expense":
        return "Operating Expenses"
    if account.account_type == "cost_of_sales":
        return "Cost of Sales"
    if account.account_type == "liability":
        return "Other Liabilities"
    if account.account_type == "asset":
        # Another cash-type account on the other side = an inter-account
        # transfer, not really an "inflow/outflow" from outside the business.
        return "Transfers Between Accounts" if account.system_key else "Other Assets"
    return "Other"


def _cash_flow_data(db: Session, period_start: date, period_end: date):
    """One entry's cash movement can have several non-cash lines (e.g. an
    expense split into net + Input VAT) — each is attributed a weighted
    share of the cash movement, proportional to its own debit+credit, so a
    ₱112 cash payment split Dr Expense 100 / Dr Input VAT 12 correctly
    shows as ₱100 Operating Expenses + ₱12 (whatever Input VAT classifies
    as) rather than all-or-nothing to one bucket."""
    cash_ids = {a.id for a in _cash_accounts(db)}
    entries = (
        db.query(models.JournalEntry)
        .join(models.JournalLine, models.JournalLine.entry_id == models.JournalEntry.id)
        .filter(models.JournalLine.account_id.in_(cash_ids or [0]),
                models.JournalEntry.txn_date.between(period_start, period_end),
                models.JournalEntry.status != "draft")
        .distinct()
        .all()
    )
    inflows, outflows = {}, {}
    for entry in entries:
        cash_lines = [ln for ln in entry.lines if ln.account_id in cash_ids]
        other_lines = [ln for ln in entry.lines if ln.account_id not in cash_ids]
        cash_net = sum((ln.debit - ln.credit) for ln in cash_lines)
        if cash_net == 0 or not other_lines:
            continue
        other_total = sum((ln.debit + ln.credit) for ln in other_lines) or Decimal("1")
        bucket = inflows if cash_net > 0 else outflows
        for ln in other_lines:
            weight = (ln.debit + ln.credit) / other_total
            label = _classify_counter(ln.account)
            bucket[label] = bucket.get(label, ZERO) + abs(cash_net) * weight
    return inflows, outflows


@router.get("/accounting/cash-flow-statement", response_class=HTMLResponse)
def cash_flow_statement(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    inflows, outflows = _cash_flow_data(db, period_start, period_end)

    inflow_rows = sorted(({"label": k, "amount": v} for k, v in inflows.items()), key=lambda r: -r["amount"])
    outflow_rows = sorted(({"label": k, "amount": v} for k, v in outflows.items()), key=lambda r: -r["amount"])
    total_in = sum((r["amount"] for r in inflow_rows), ZERO)
    total_out = sum((r["amount"] for r in outflow_rows), ZERO)

    return templates.TemplateResponse(
        "accounting/cash_flow_statement.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "inflow_rows": inflow_rows, "outflow_rows": outflow_rows,
         "total_in": total_in, "total_out": total_out, "net": total_in - total_out,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


@router.get("/accounting/cash-flow-statement/export")
def export_cash_flow_statement(days: int = 30, date_from: str = "", date_to: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, _ = _resolve_period(days, date_from, date_to)
    inflows, outflows = _cash_flow_data(db, period_start, period_end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"
    ws.append(["Type", "Category", "Amount"])
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for label, amount in sorted(inflows.items(), key=lambda kv: -kv[1]):
        ws.append(["Inflow", label, float(amount)])
    for label, amount in sorted(outflows.items(), key=lambda kv: -kv[1]):
        ws.append(["Outflow", label, float(amount)])
    total_in = sum(inflows.values(), ZERO)
    total_out = sum(outflows.values(), ZERO)
    ws.append(["", "Net Cash Flow", float(total_in - total_out)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    widths = [12, 34, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cash_flow_{period_end.isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --------------------------------------------------------------------------- #
# Balance Sheet + ledger-based P&L
# --------------------------------------------------------------------------- #
def _type_balance(db: Session, account_type: str, as_of: date) -> Decimal:
    """Cumulative balance of every active account of one account_type, signed
    by each account's own normal_balance, summed since inception through
    as_of. Cumulative (not period-only) because there's no formal
    period-close workflow — see plan Phase 2+ item 5."""
    accounts = db.query(models.Account).filter(
        models.Account.is_active.is_(True), models.Account.account_type == account_type,
    ).all()
    total = ZERO
    for account in accounts:
        total += _account_balance_before(db, account, as_of + timedelta(days=1))
    return total


def _pl_ledger_data(db: Session, period_start: date, period_end: date):
    """Ledger-based P&L: unlike Balance Sheet's cumulative-since-inception
    balances, this is scoped to the period — sum of JournalLine movement
    (not opening balance) for revenue/cost_of_sales/expense accounts."""
    def _period_total(account_type: str) -> Decimal:
        accounts = db.query(models.Account).filter(
            models.Account.is_active.is_(True), models.Account.account_type == account_type,
        ).all()
        total = ZERO
        for account in accounts:
            row = (
                db.query(func.coalesce(func.sum(models.JournalLine.debit), 0), func.coalesce(func.sum(models.JournalLine.credit), 0))
                .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
                .filter(models.JournalLine.account_id == account.id,
                        models.JournalEntry.txn_date.between(period_start, period_end),
                        models.JournalEntry.status != "draft")
                .one()
            )
            debit, credit = Decimal(str(row[0] or 0)), Decimal(str(row[1] or 0))
            total += (debit - credit) if account.normal_balance == "debit" else (credit - debit)
        return total

    revenue = _period_total("revenue")
    cost_of_sales = _period_total("cost_of_sales")
    expenses = _period_total("expense")
    return {
        "revenue": revenue, "cost_of_sales": cost_of_sales, "expenses": expenses,
        "gross_profit": revenue - cost_of_sales,
        "net_profit": revenue - cost_of_sales - expenses,
    }


@router.get("/accounting/balance-sheet", response_class=HTMLResponse)
def balance_sheet(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    _, period_end, custom = _resolve_period(days, date_from, date_to)

    def _accounts_of(account_type: str):
        accounts = db.query(models.Account).filter(
            models.Account.is_active.is_(True), models.Account.account_type == account_type,
        ).order_by(models.Account.code).all()
        rows = []
        total = ZERO
        for a in accounts:
            bal = _account_balance_before(db, a, period_end + timedelta(days=1))
            if bal == 0:
                continue
            rows.append({"account": a, "balance": bal})
            total += bal
        return rows, total

    asset_rows, total_assets = _accounts_of("asset")
    liability_rows, total_liabilities = _accounts_of("liability")
    equity_rows, equity_account_total = _accounts_of("equity")

    revenue = _type_balance(db, "revenue", period_end)
    cost_of_sales = _type_balance(db, "cost_of_sales", period_end)
    expenses = _type_balance(db, "expense", period_end)
    net_income = revenue - cost_of_sales - expenses
    total_equity = equity_account_total + net_income

    return templates.TemplateResponse(
        "accounting/balance_sheet.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "asset_rows": asset_rows, "liability_rows": liability_rows, "equity_rows": equity_rows,
         "total_assets": total_assets, "total_liabilities": total_liabilities,
         "equity_account_total": equity_account_total, "net_income": net_income, "total_equity": total_equity,
         "total_liab_and_equity": total_liabilities + total_equity,
         "balanced": total_assets == (total_liabilities + total_equity),
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_end": period_end, "custom": custom},
    )


@router.get("/accounting/profit-loss", response_class=HTMLResponse)
def ledger_profit_loss(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    data = _pl_ledger_data(db, period_start, period_end)

    # Reconcile against the existing operational P&L (app/reports.py) for
    # the same period, same idea as reconcile-sales — surface a mismatch
    # rather than silently trusting a second, independently-computed number.
    from . import reports as reports_module
    operational = reports_module._pl_data(db, period_start, period_end)

    return templates.TemplateResponse(
        "accounting/profit_loss.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "data": data, "operational": operational,
         "net_diff": data["net_profit"] - operational["net_profit"],
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


# --------------------------------------------------------------------------- #
# VAT Report — Output VAT (from Sales) vs Input VAT (from Expenses), netted
# --------------------------------------------------------------------------- #
def _vat_period_total(db: Session, system_key: str, period_start: date, period_end: date) -> Decimal:
    """Net period movement for the account currently tagged with this
    system_key, signed by its own normal_balance so it reads as a positive
    accrual regardless of which side it normally sits on."""
    account = db.query(models.Account).filter(models.Account.system_key == system_key).first()
    if not account:
        return ZERO
    row = (
        db.query(func.coalesce(func.sum(models.JournalLine.debit), 0), func.coalesce(func.sum(models.JournalLine.credit), 0))
        .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
        .filter(models.JournalLine.account_id == account.id,
                models.JournalEntry.txn_date.between(period_start, period_end),
                models.JournalEntry.status != "draft")
        .one()
    )
    debit, credit = Decimal(str(row[0] or 0)), Decimal(str(row[1] or 0))
    return (debit - credit) if account.normal_balance == "debit" else (credit - debit)


@router.get("/accounting/vat-report", response_class=HTMLResponse)
def vat_report(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)

    output_vat = _vat_period_total(db, "OUTPUT_VAT", period_start, period_end)
    input_vat = _vat_period_total(db, "INPUT_VAT", period_start, period_end)
    vat_payable = output_vat - input_vat

    # Reconciliation vs. the operational figures those postings came from.
    sale_vat_total = (
        db.query(func.coalesce(func.sum(models.Sale.vat_amount), 0))
        .filter(models.Sale.txn_type == "sale", models.Sale.is_voided.is_(False),
                _local_date(models.Sale.created_at).between(period_start, period_end))
        .scalar()
    )
    sale_vat_total = Decimal(str(sale_vat_total or 0))

    expense_vat_total = (
        db.query(func.coalesce(func.sum(models.Expense.vat_amount), 0))
        .filter(models.Expense.is_voided.is_(False), models.Expense.expense_date.between(period_start, period_end))
        .scalar()
    )
    expense_vat_total = Decimal(str(expense_vat_total or 0))

    return templates.TemplateResponse(
        "accounting/vat_report.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "output_vat": output_vat, "input_vat": input_vat, "vat_payable": vat_payable,
         "sale_vat_total": sale_vat_total, "expense_vat_total": expense_vat_total,
         "output_diff": output_vat - sale_vat_total, "input_diff": input_vat - expense_vat_total,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


# --------------------------------------------------------------------------- #
# AR / AP subledgers — per-customer / per-supplier drill-down of the AR/AP
# account's GL, joined via JournalEntry.source_id + source_type
# --------------------------------------------------------------------------- #
def _resolve_party(db: Session, entry: models.JournalEntry, kind: str):
    """Walks a reversal entry back to what it reversed so a voided sale's
    settlement still attributes to the right customer. Returns (id, name)."""
    source_type, source_id = entry.source_type, entry.source_id
    if source_type == "reversal" and entry.is_reversal_of_id:
        original = db.get(models.JournalEntry, entry.is_reversal_of_id)
        if original:
            source_type, source_id = original.source_type, original.source_id
    if kind == "ar" and source_type in ("sale", "sale_settlement") and source_id:
        sale = db.get(models.Sale, source_id)
        if sale and sale.customer:
            return sale.customer.id, sale.customer.name
    if kind == "ap" and source_type in ("purchase", "purchase_settlement") and source_id:
        purchase = db.get(models.Purchase, source_id)
        if purchase and purchase.supplier:
            return purchase.supplier.id, purchase.supplier.name
    return None, "Walk-in / Unspecified"


def _subledger(request: Request, db: Session, user, *, kind: str, system_key: str, party_id: int, days: int, date_from: str, date_to: str):
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    account = db.query(models.Account).filter(models.Account.system_key == system_key).first()

    summary = {}
    detail_rows = []
    opening = closing = ZERO
    if account:
        all_rows = (
            db.query(models.JournalLine, models.JournalEntry)
            .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
            .filter(models.JournalLine.account_id == account.id, models.JournalEntry.txn_date <= period_end,
                    models.JournalEntry.status != "draft")
            .order_by(models.JournalEntry.txn_date, models.JournalEntry.id)
            .all()
        )
        sign = 1 if account.normal_balance == "debit" else -1
        for line, entry in all_rows:
            pid, pname = _resolve_party(db, entry, kind)
            delta = sign * (line.debit - line.credit)
            key = pid or 0
            if key not in summary:
                summary[key] = {"id": pid, "name": pname, "balance": ZERO}
            summary[key]["balance"] += delta

            if entry.txn_date < period_start and (pid or 0) == party_id:
                opening += delta
            if party_id and (pid or 0) == party_id and period_start <= entry.txn_date <= period_end:
                detail_rows.append({"entry": entry, "line": line, "delta": delta})

        running = opening
        for r in detail_rows:
            running += r["delta"]
            r["balance"] = running
        closing = running

    summary_rows = sorted(
        [v for v in summary.values() if v["balance"] != 0],
        key=lambda r: r["balance"], reverse=True,
    )
    selected_party = next((v for v in summary.values() if (v["id"] or 0) == party_id), None) if party_id else None

    return templates.TemplateResponse(
        "accounting/ar_ap_subledger.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "kind": kind, "title": "Accounts Receivable" if kind == "ar" else "Accounts Payable",
         "party_label": "Customer" if kind == "ar" else "Supplier",
         "route": "/accounting/ar-subledger" if kind == "ar" else "/accounting/ap-subledger",
         "summary_rows": summary_rows, "party_id": party_id, "selected_party": selected_party,
         "detail_rows": detail_rows, "opening": opening, "closing": closing,
         "total_outstanding": sum((r["balance"] for r in summary_rows), ZERO),
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


@router.get("/accounting/ar-subledger", response_class=HTMLResponse)
def ar_subledger(
    request: Request, party_id: int = 0, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    return _subledger(request, db, user, kind="ar", system_key="AR", party_id=party_id, days=days, date_from=date_from, date_to=date_to)


@router.get("/accounting/ap-subledger", response_class=HTMLResponse)
def ap_subledger(
    request: Request, party_id: int = 0, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    return _subledger(request, db, user, kind="ap", system_key="AP", party_id=party_id, days=days, date_from=date_from, date_to=date_to)


# --------------------------------------------------------------------------- #
# Manual Journal Entry — draft/post/reverse workflow. Built last on purpose
# (see the accounting plan): a "fix it by hand" UI is exactly where a real
# posting bug could get hidden instead of caught, so every automatic
# posting path had to be trustworthy first. Two-step (draft, then a
# separate explicit "Post" click) so a mistake doesn't hit the books
# before someone's looked at it a second time — the same approve-gate idea
# as _can_void_sale in pos.py, just without a separate role for it since
# every other accounting screen in this app is already is_staff-gated.
# --------------------------------------------------------------------------- #
@router.get("/accounting/journal-entries", response_class=HTMLResponse)
def journal_entries_list(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    entries = (
        db.query(models.JournalEntry)
        .filter(models.JournalEntry.source_type == "manual")
        .order_by(models.JournalEntry.id.desc())
        .limit(100)
        .all()
    )
    return templates.TemplateResponse(
        "accounting/journal_entries.html",
        {"request": request, "app_name": request.app.title, "user": user, "entries": entries},
    )


@router.get("/accounting/journal-entries/new", response_class=HTMLResponse)
def journal_entry_new(request: Request, error: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    accounts = db.query(models.Account).filter(models.Account.is_active.is_(True)).order_by(models.Account.code).all()
    return templates.TemplateResponse(
        "accounting/journal_entry_form.html",
        {"request": request, "app_name": request.app.title, "user": user, "accounts": accounts,
         "today": _today().isoformat(), "error": error},
    )


@router.post("/accounting/journal-entries")
async def journal_entry_create(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    form = await request.form()
    txn_date = _parse_date(form.get("txn_date", "")) or _today()
    description = (form.get("description") or "").strip()
    reference_no = (form.get("reference_no") or "").strip() or None

    account_ids = form.getlist("account_id")
    sides = form.getlist("side")
    amounts = form.getlist("amount")
    memos = form.getlist("memo")

    lines = []
    for account_id, side, amount, memo in zip(account_ids, sides, amounts, memos):
        if not account_id or not amount:
            continue
        amt = _money(amount)
        if amt <= 0:
            continue
        lines.append({"_account_id": int(account_id), "amount": amt, "side": side, "memo": (memo or "").strip() or None})

    if not description:
        return RedirectResponse("/accounting/journal-entries/new?error=Description+is+required.", status_code=302)

    try:
        entry = post_journal(
            db, txn_date=txn_date, source_type="manual", source_id=None, description=description,
            lines=lines, reference_no=reference_no, entered_by_id=user.id, status="draft",
        )
    except PostingError as e:
        return RedirectResponse(f"/accounting/journal-entries/new?error={str(e)}", status_code=302)
    db.commit()
    return RedirectResponse(f"/accounting/journal-entries?created={entry.journal_no}", status_code=302)


@router.post("/accounting/journal-entries/{entry_id:int}/post")
def journal_entry_post(entry_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    entry = db.get(models.JournalEntry, entry_id)
    if entry and entry.source_type == "manual":
        post_draft_entry(db, entry)
        db.commit()
    return RedirectResponse("/accounting/journal-entries", status_code=302)


@router.post("/accounting/journal-entries/{entry_id:int}/delete")
def journal_entry_delete(entry_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    entry = db.get(models.JournalEntry, entry_id)
    if entry and entry.source_type == "manual":
        delete_draft_entry(db, entry)
        db.commit()
    return RedirectResponse("/accounting/journal-entries", status_code=302)


@router.post("/accounting/journal-entries/{entry_id:int}/reverse")
def journal_entry_reverse(entry_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    entry = db.get(models.JournalEntry, entry_id)
    if entry and entry.source_type == "manual" and entry.status == "posted":
        reverse_journal(db, entry, reason="Manual reversal", entered_by_id=user.id)
        db.commit()
    return RedirectResponse("/accounting/journal-entries", status_code=302)


# --------------------------------------------------------------------------- #
# Financial Dashboard — rollup of every other report. Built last on purpose:
# a rollup built before the reports underneath it were trustworthy would
# just have been showing confidently wrong numbers.
# --------------------------------------------------------------------------- #
@router.get("/accounting/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)

    _, total_debit, total_credit = _trial_balance_rows(db, period_end)
    balanced = total_debit == total_credit

    total_assets = _type_balance(db, "asset", period_end)
    total_liabilities = _type_balance(db, "liability", period_end)
    equity_accounts = _type_balance(db, "equity", period_end)
    revenue_to_date = _type_balance(db, "revenue", period_end)
    cost_of_sales_to_date = _type_balance(db, "cost_of_sales", period_end)
    expenses_to_date = _type_balance(db, "expense", period_end)
    net_income_to_date = revenue_to_date - cost_of_sales_to_date - expenses_to_date
    total_equity = equity_accounts + net_income_to_date

    pl_period = _pl_ledger_data(db, period_start, period_end)

    cash_total = ZERO
    for account in _cash_accounts(db):
        cash_total += _account_balance_before(db, account, period_end + timedelta(days=1))

    ar_account = db.query(models.Account).filter(models.Account.system_key == "AR").first()
    ap_account = db.query(models.Account).filter(models.Account.system_key == "AP").first()
    ar_total = _account_balance_before(db, ar_account, period_end + timedelta(days=1)) if ar_account else ZERO
    ap_total = _account_balance_before(db, ap_account, period_end + timedelta(days=1)) if ap_account else ZERO

    output_vat = _vat_period_total(db, "OUTPUT_VAT", period_start, period_end)
    input_vat = _vat_period_total(db, "INPUT_VAT", period_start, period_end)
    vat_payable = output_vat - input_vat

    return templates.TemplateResponse(
        "accounting/dashboard.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "balanced": balanced, "total_debit": total_debit, "total_credit": total_credit,
         "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,
         "net_income_to_date": net_income_to_date, "pl_period": pl_period,
         "cash_total": cash_total, "ar_total": ar_total, "ap_total": ap_total, "vat_payable": vat_payable,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )


# --------------------------------------------------------------------------- #
# Reconciliation: ledger Sales total vs the existing operational Sales report
# --------------------------------------------------------------------------- #
@router.get("/accounting/reconcile-sales", response_class=HTMLResponse)
def reconcile_sales(
    request: Request, days: int = 30, date_from: str = "", date_to: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    period_start, period_end, custom = _resolve_period(days, date_from, date_to)

    ledger_total = (
        db.query(func.coalesce(func.sum(models.JournalLine.credit), 0))
        .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
        .join(models.Account, models.JournalLine.account_id == models.Account.id)
        .filter(
            models.Account.system_key.in_(("SALES_REVENUE", "OUTPUT_VAT")),
            models.JournalEntry.txn_date.between(period_start, period_end),
            models.JournalEntry.status == "posted",
        )
        .scalar()
    )
    ledger_total = Decimal(str(ledger_total or 0))

    # Scoped to plain "sale" transactions only — Phase 1 doesn't post
    # refunds/exchanges, so comparing against the broader P&L revenue figure
    # (which nets those in) would show a mismatch that isn't actually a bug.
    report_total = (
        db.query(func.coalesce(func.sum(models.Sale.total), 0))
        .filter(
            models.Sale.txn_type == "sale", models.Sale.is_voided.is_(False),
            _local_date(models.Sale.created_at).between(period_start, period_end),
        )
        .scalar()
    )
    report_total = Decimal(str(report_total or 0))

    unposted = []
    if ledger_total != report_total:
        sales = (
            db.query(models.Sale)
            .filter(models.Sale.txn_type == "sale", models.Sale.is_voided.is_(False),
                    _local_date(models.Sale.created_at).between(period_start, period_end))
            .all()
        )
        posted_ids = {
            sid for (sid,) in db.query(models.JournalEntry.source_id)
            .filter(models.JournalEntry.source_type == "sale", models.JournalEntry.status == "posted",
                    models.JournalEntry.source_id.in_([s.id for s in sales]))
            .all()
        }
        unposted = [s for s in sales if s.id not in posted_ids]

    return templates.TemplateResponse(
        "accounting/reconcile_sales.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "ledger_total": ledger_total, "report_total": report_total,
         "diff": ledger_total - report_total, "unposted": unposted,
         "days": days, "date_from": date_from, "date_to": date_to,
         "period_start": period_start, "period_end": period_end, "custom": custom},
    )
