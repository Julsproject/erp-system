"""Cashier activity: a read-only history of what each person processed —
one day, or any date range.

Nothing here blocks anyone from working and nothing needs to be manually
counted or declared; it's computed straight from Sales/Payments/Settlements
already on file. Voided sales are excluded throughout, so this always agrees
with Sales, the Dashboard and the P&L for the same dates.

The physical-cash counterpart is Cash Drawer (shifts.py): this says what the
system thinks was collected, that says what was actually in the till.
"""
import io as _io
from datetime import date, datetime, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import models
from .database import get_db
from .deps import get_current_user, is_staff
from .templating import templates

router = APIRouter()

MANILA = ZoneInfo("Asia/Manila")
METHOD_LABELS = {
    "cash": "Cash", "gcash": "GCash", "maya": "Maya", "other_ewallet": "Other E-Wallet",
    "card": "Card", "bank_transfer": "Bank Transfer", "cheque": "Cheque", "receivable": "Receivable (Credit)",
}


def _today() -> date:
    return datetime.now(MANILA).date()


def _parse_day(s: str):
    try:
        return date.fromisoformat(s) if s else None
    except ValueError:
        return None


def _local_date(col):
    return func.date(func.timezone("Asia/Manila", col))


def _resolve_range(day: str, date_from: str, date_to: str):
    """Return (start, end, is_range). A single ?day= still works (that's what
    every existing link uses); a from/to pair widens it. Ends are swapped if
    entered backwards and a future end is pulled back to today, same
    forgiving behaviour as the other reports."""
    today = _today()
    df, dt = _parse_day(date_from), _parse_day(date_to)
    if df and dt:
        if dt > today:
            dt = today
        if df > dt:
            df, dt = dt, df
        return df, dt, (df != dt)
    d = _parse_day(day) or today
    return d, d, False


def _day_summary(db: Session, cashier_id: int, day: date, end: date = None) -> dict:
    """Everything this cashier processed between `day` and `end` inclusive.
    `end` defaults to `day`, so the original single-day callers are
    unchanged."""
    end = end or day
    # Voided sales are excluded everywhere else (Sales, Dashboard, P&L, the
    # accounting ledger), and they must be excluded here too — a void means
    # the sale is treated as never having happened, so leaving it in would
    # credit a cashier with money they didn't take and make this page
    # disagree with every other report for the same day.
    sales = (
        db.query(models.Sale)
        .filter(
            models.Sale.cashier_id == cashier_id,
            models.Sale.is_voided.is_(False),
            _local_date(models.Sale.created_at).between(day, end),
        )
        .order_by(models.Sale.id)
        .all()
    )
    pay_rows = (
        db.query(models.Payment.method, func.coalesce(func.sum(models.Payment.amount), 0))
        .join(models.Sale, models.Payment.sale_id == models.Sale.id)
        .filter(
            models.Sale.cashier_id == cashier_id,
            models.Sale.is_voided.is_(False),
            _local_date(models.Sale.created_at).between(day, end),
        )
        .group_by(models.Payment.method)
        .all()
    )
    collections = (
        db.query(func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
        .filter(
            models.ReceivableSettlement.cashier_id == cashier_id,
            _local_date(models.ReceivableSettlement.created_at).between(day, end),
        )
        .scalar()
    )
    by_method = {
        m: Decimal(str(a or 0)) for m, a in pay_rows
    }
    # Change is always handed back in physical cash (see pos.py: a refund's or
    # an overpaid exchange's/sale's change is cash out regardless of which
    # method funded the original payment) — net it out of the cash bucket, or
    # this would overstate cash on hand by whatever change was given that day.
    change_given = sum((s.change_amount or Decimal("0") for s in sales), Decimal("0"))
    if "cash" in by_method or change_given > 0:
        by_method["cash"] = by_method.get("cash", Decimal("0")) - change_given
    by_method_list = [
        {"label": METHOD_LABELS.get(m, (m or "").title()), "amount": a}
        for m, a in sorted(by_method.items(), key=lambda r: float(r[1] or 0), reverse=True)
    ]
    net_total = sum((s.total or Decimal("0") for s in sales), Decimal("0"))
    return {
        "sales": sales,
        "count": len(sales),
        "net_total": net_total,
        "by_method": by_method_list,
        "collections": Decimal(str(collections or 0)),
    }


def _all_rows(db: Session, start: date, end: date):
    """One row per person who actually processed something in the window.
    Anyone who sold nothing is left out rather than shown as a zero row —
    on a range that would otherwise list every staff member who has ever
    existed. Shared by the page and the Excel export."""
    people = db.query(models.User).filter(models.User.is_active.is_(True)).order_by(models.User.username).all()
    rows = []
    for p in people:
        summary = _day_summary(db, p.id, start, end)
        if summary["count"] > 0 or summary["collections"]:
            rows.append({"user": p, "summary": summary})
    return rows


@router.get("/activity", response_class=HTMLResponse)
def my_activity(request: Request, day: str = "", date_from: str = "", date_to: str = "",
                db: Session = Depends(get_db), user=Depends(get_current_user)):
    """A cashier's own activity for a day, or any range."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if is_staff(user):
        return RedirectResponse("/activity/all", status_code=302)
    start, end, is_range = _resolve_range(day, date_from, date_to)
    summary = _day_summary(db, user.id, start, end)
    return templates.TemplateResponse(
        "activity/mine.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "day": start, "end": end, "is_range": is_range, "summary": summary, "viewing": None},
    )


@router.get("/activity/all", response_class=HTMLResponse)
def all_activity(request: Request, day: str = "", date_from: str = "", date_to: str = "",
                 db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Every cashier's activity for a day or range, one row each."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/activity", status_code=302)
    start, end, is_range = _resolve_range(day, date_from, date_to)
    rows = _all_rows(db, start, end)
    totals = {
        "count": sum(r["summary"]["count"] for r in rows),
        "net_total": sum((r["summary"]["net_total"] for r in rows), Decimal("0")),
        "collections": sum((r["summary"]["collections"] for r in rows), Decimal("0")),
    }
    return templates.TemplateResponse(
        "activity/all.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "day": start, "end": end, "is_range": is_range, "rows": rows, "totals": totals},
    )


@router.get("/activity/user/{user_id:int}", response_class=HTMLResponse)
def user_activity(
    user_id: int,
    request: Request,
    day: str = "",
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Drilling into one person's day/range (or a cashier's own — same page)."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user) and user.id != user_id:
        return RedirectResponse("/activity", status_code=302)
    target = db.get(models.User, user_id)
    if not target:
        return RedirectResponse("/activity/all" if is_staff(user) else "/activity", status_code=302)
    start, end, is_range = _resolve_range(day, date_from, date_to)
    summary = _day_summary(db, user_id, start, end)
    return templates.TemplateResponse(
        "activity/mine.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "day": start, "end": end, "is_range": is_range, "summary": summary,
         "viewing": target if is_staff(user) else None},
    )


@router.get("/activity/export.xlsx")
def export_activity(
    day: str = "",
    date_from: str = "",
    date_to: str = "",
    user_id: int = 0,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Excel of whatever is on screen: the per-person summary for a day or
    range, and — when one person is selected — every transaction behind it."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    # A cashier may export only their own; staff may export anyone.
    if not is_staff(user):
        if user_id and user_id != user.id:
            return RedirectResponse("/activity", status_code=302)
        user_id = user.id

    start, end, _ = _resolve_range(day, date_from, date_to)
    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="1F6FEB")

    def _style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
        ws.freeze_panes = "A2"

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Cashier", "Transactions", "Total Sales", "Credit Collected"])
    _style_header(ws)
    rows = _all_rows(db, start, end)
    if user_id:
        rows = [r for r in rows if r["user"].id == user_id]
    for r in rows:
        ws.append([
            r["user"].full_name or r["user"].username,
            r["summary"]["count"],
            float(r["summary"]["net_total"] or 0),
            float(r["summary"]["collections"] or 0),
        ])
    if rows:
        ws.append([])
        ws.append([
            "TOTAL",
            sum(r["summary"]["count"] for r in rows),
            float(sum((r["summary"]["net_total"] for r in rows), Decimal("0"))),
            float(sum((r["summary"]["collections"] for r in rows), Decimal("0"))),
        ])
    for i, w in enumerate([26, 14, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Second sheet: the transactions themselves, but only when the export is
    # scoped to one person — across everybody it would be a duplicate of the
    # Sales export rather than an activity report.
    if user_id:
        ws2 = wb.create_sheet("Transactions")
        ws2.append(["Invoice", "Type", "Date", "Time", "Customer", "Payment", "Total"])
        _style_header(ws2)
        for sale in _day_summary(db, user_id, start, end)["sales"]:
            local = sale.created_at.astimezone(MANILA) if sale.created_at else None
            ws2.append([
                sale.invoice_no or "",
                (sale.txn_type or "").title(),
                local.strftime("%Y-%m-%d") if local else "",
                local.strftime("%I:%M %p") if local else "",
                sale.customer_name or "Walk-in",
                sale.payment_method or "",
                float(sale.total or 0),
            ])
        for i, w in enumerate([16, 12, 13, 11, 24, 18, 14], start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    span = start.isoformat() if start == end else f"{start.isoformat()}_{end.isoformat()}"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cashier_activity_{span}.xlsx"'},
    )
