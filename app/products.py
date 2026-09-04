"""Inventory / Products module: encode products with beginning stock.

Columns the client asked for: Product Name, Category, Unit Type, Cost of Sales,
Selling Price, Actual Beginning Stocks, Stocks Qty, Total Qty.
"""
import base64
import csv
import io
import json
import re
from urllib.parse import quote
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from difflib import SequenceMatcher
from zoneinfo import ZoneInfo

import barcode as barcode_lib
from barcode.writer import SVGWriter
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload
from starlette.concurrency import run_in_threadpool

from . import audit, models, pricing, settings_store
from .database import SessionLocal, get_db
from .deps import get_current_user, is_staff, safe_back_url
from .double_deductions import find_all_double_deduction_candidates, find_double_deduction_candidates
from .search_utils import multi_word_ilike
from .templating import templates

router = APIRouter()

PAGE_SIZE = 15
MANILA = ZoneInfo("Asia/Manila")
MONTH_END_SETTING_KEY = "last_month_end_period"
CENTS = Decimal("0.01")

BULK_MODES = ("pct", "amount", "cost_pct", "markup", "margin")

# Fields whose before/after we log on a product edit. Stock and price changes
# are the theft/accountability-sensitive ones the owner most wants visible.
AUDIT_FIELDS = ["name", "barcode", "cost_price", "selling_price", "beginning_stock", "stock_qty",
                "reorder_level", "is_vat", "markup_pct", "markup_price", "margin_pct", "margin_price"]

# Why a manual stock quantity changed — required whenever an edit moves
# Beginning Stock or Stock Qty, so the eventual peso value (qty * cost) has
# an accountable reason attached, not just a silent number change.
ADJUSTMENT_REASONS = [
    ("count_correction", "Count correction"),
    ("damage", "Damage / breakage"),
    ("theft", "Theft / loss"),
    ("expired", "Expired / spoiled"),
    ("initial_balance", "Initial balance correction"),
    ("other", "Other"),
]
ADJUSTMENT_REASON_LABELS = dict(ADJUSTMENT_REASONS)

# Throttles the month-end rollover check to at most once per real day per
# process — see maybe_run_month_end_rollover(). Not persisted; a restart just
# means the next request re-checks, which is harmless (the actual guard
# against double-rolling the same month lives in app_settings, not here).
_last_rollover_check_day = None


def _run_month_end_rollover(db: Session, period: str) -> int:
    """Fold every active product's Stocks Qty into Actual Beginning and reset
    Stocks Qty to 0 — the shop's own month-end close: whatever's left on the
    shelf becomes next month's opening balance, and Stocks Qty starts fresh
    for the new month's purchases. total_qty (and therefore stock valuation)
    is unchanged, so this is a reclassification, not a StockMovement/adjustment.

    Only a POSITIVE Stocks Qty gets folded in. A negative Stocks Qty (an
    oversold backlog that hasn't been reconciled by a Stock Count yet) is
    left exactly where it is instead of being carried into Actual Beginning
    — Beginning is meant to be a clean opening balance, and once a negative
    amount lands there it used to be permanent (a Stock Count only ever
    corrected Stocks Qty, never Beginning — see _apply_stock_count_correction
    for the fix on that side). Leaving it in Stocks Qty keeps it exactly as
    reachable for reconciliation as it always was.

    One MonthEndRolloverLine per product is recorded so the exact amount
    that moved is traceable later (see /reports/month-end-rollover), plus
    one audit summary row for the Activity Log."""
    products = db.query(models.Product).filter(models.Product.is_active.is_(True)).all()
    rolled = 0
    for p in products:
        old_beginning = Decimal(str(p.beginning_stock or 0))
        old_stock = Decimal(str(p.stock_qty or 0))
        qty_moved = max(old_stock, Decimal("0"))
        new_beginning = old_beginning + qty_moved
        new_stock = old_stock - qty_moved
        if new_beginning != old_beginning or new_stock != old_stock:
            p.beginning_stock = new_beginning
            p.stock_qty = new_stock
            rolled += 1
            db.add(models.MonthEndRolloverLine(
                product_id=p.id, product_name=p.name, period=period,
                qty_moved=qty_moved, old_beginning=old_beginning, new_beginning=new_beginning,
            ))
    if rolled:
        audit.record(
            db, action="update", entity_type="product", entity_label="Month-end rollover",
            summary=f"Month-end rollover into {period}: {rolled} product(s) — Stocks Qty folded into Actual Beginning",
        )
    return rolled


def maybe_run_month_end_rollover() -> None:
    """Auto-fires the month-end rollover on the first page load of a new
    calendar month (Asia/Manila). This app has no scheduler process, so — like
    notifications._maybe_sweep — it piggybacks on normal page loads instead:
    cheap in-memory guard skips the DB check after the first hit each day,
    and the real once-per-month guard lives in app_settings so it can never
    double-roll even across restarts or multiple app instances.
    """
    global _last_rollover_check_day
    today = datetime.now(MANILA).date()
    if _last_rollover_check_day == today:
        return
    _last_rollover_check_day = today
    current_period = today.strftime("%Y-%m")
    db = SessionLocal()
    try:
        last_period = settings_store.get_setting(db, MONTH_END_SETTING_KEY, "")
        if last_period == current_period:
            return
        if last_period:  # not the very first run ever — actually roll
            _run_month_end_rollover(db, current_period)
        # First run ever: just establish a baseline, don't roll over
        # retroactively for however many months of history already exist.
        settings_store.set_setting(db, MONTH_END_SETTING_KEY, current_period)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


def _product_snapshot(p: models.Product) -> dict:
    d = audit.snapshot(p, AUDIT_FIELDS)
    d["category"] = p.category.name if p.category else None
    d["subcategory"] = p.subcategory.name if p.subcategory else None
    d["unit_type"] = p.unit_type.name if p.unit_type else None
    return d

# Columns used for the import template and the upload parser. Order here is
# just what the downloaded template looks like — the parser reads columns by
# header name (see HEADER_MAP), so this can be reordered freely without
# touching how uploads are read, and doesn't affect the Inventory page's own
# column order either.
TEMPLATE_HEADERS = [
    "Barcode",
    "Product Name",
    "Category",
    "Sub Category",
    "Unit Type",
    "Shelf",
    "Cost of Sales",
    "Selling Price",
    "Actual Beginning Stocks",
    "Stocks Qty",
]

# Maps (normalized) spreadsheet headers -> internal field names, so columns can
# be in any order and tolerate small naming differences.
HEADER_MAP = {
    "product name": "name", "name": "name", "product": "name",
    "barcode": "barcode", "bar code": "barcode", "upc": "barcode", "ean": "barcode",
    "category": "category",
    "sub category": "subcategory", "subcategory": "subcategory", "sub-category": "subcategory",
    "unit type": "unit_type", "unit": "unit_type", "unit of measure": "unit_type", "uom": "unit_type",
    "shelf": "shelf", "shelf location": "shelf", "shelf #": "shelf", "shelf no": "shelf",
    "cost of sales": "cost", "cost": "cost", "cost price": "cost",
    "selling price": "selling", "price": "selling", "srp": "selling",
    "actual beginning stocks": "beginning", "beginning stock": "beginning",
    "beginning stocks": "beginning", "beginning": "beginning",
    "stocks qty": "stocks", "stock qty": "stocks", "stocks": "stocks", "stock": "stocks",
    "vat": "vat", "vat-able": "vat", "vatable": "vat",
}
FIELDS = ["name", "barcode", "category", "subcategory", "unit_type", "shelf", "cost", "selling", "beginning", "stocks", "vat"]

# --- Units & Conversions import (separate flow for products that sell in
# more than one unit, e.g. Sack/Elf Load/Kg) — see routes near the bottom of
# this file. Each row adds one unit to a product's ladder; several rows with
# the same Product Name build up that product's full ladder together.
TEMPLATE_HEADERS_UNITS = [
    "Product Name",
    "Unit Name",
    "Equivalent Qty",
    "Of Unit (blank = base unit)",
    "Fixed Price",
    "Markup %",
    "Margin %",
]
HEADER_MAP_UNITS = {
    "product name": "product", "product": "product",
    "unit name": "unit_name", "unit": "unit_name", "sell as": "unit_name",
    "equivalent qty": "factor", "equivalent": "factor", "factor": "factor",
    "qty": "factor", "1 unit =": "factor",
    "of unit (blank = base unit)": "relative_to", "of unit": "relative_to",
    "of": "relative_to", "relative to": "relative_to",
    "fixed price": "price", "price": "price",
    "markup %": "markup_pct", "markup": "markup_pct",
    "margin %": "margin_pct", "margin": "margin_pct",
}
FIELDS_UNITS = ["product", "unit_name", "factor", "relative_to", "price", "markup_pct", "margin_pct"]


def _to_decimal(value: str, default: str = "0") -> Decimal:
    value = (value or "").strip().replace(",", "")
    if value == "":
        value = default
    try:
        return Decimal(value)
    except InvalidOperation:
        return Decimal(default)


def _get_or_create_category(db: Session, name: str):
    name = (name or "").strip()
    if not name:
        return None
    existing = db.query(models.Category).filter(func.lower(models.Category.name) == name.lower()).first()
    if existing:
        return existing
    cat = models.Category(name=name)
    db.add(cat)
    db.flush()
    return cat


def _get_or_create_subcategory(db: Session, name: str, category: models.Category = None):
    name = (name or "").strip()
    if not name:
        return None
    existing = db.query(models.SubCategory).filter(func.lower(models.SubCategory.name) == name.lower()).first()
    if existing:
        # Backfill the parent link if this sub category was created before a
        # Category was ever paired with it.
        if existing.category_id is None and category is not None:
            existing.category_id = category.id
        return existing
    sub = models.SubCategory(name=name, category_id=category.id if category else None)
    db.add(sub)
    db.flush()
    return sub


def _get_or_create_unit_type(db: Session, name: str):
    name = (name or "").strip()
    if not name:
        return None
    existing = db.query(models.UnitType).filter(func.lower(models.UnitType.name) == name.lower()).first()
    if existing:
        return existing
    unit = models.UnitType(name=name)
    db.add(unit)
    db.flush()
    return unit


def _get_or_create_shelf(db: Session, name: str):
    name = (name or "").strip()
    if not name:
        return None
    existing = db.query(models.Shelf).filter(func.lower(models.Shelf.name) == name.lower()).first()
    if existing:
        return existing
    shelf = models.Shelf(name=name)
    db.add(shelf)
    db.flush()
    return shelf


_needs_review_expr = pricing.needs_review_expr


def low_stock_expr(default_pct):
    """SQL expression: true when a product is at or under its low-stock
    threshold. Uses the product's own "Low Stock Alert At" (reorder_level)
    if it's set (>0); otherwise, if the shop has a default_pct configured
    (Settings), falls back to that % of the product's Actual Beginning
    Stocks. No fallback fires for a product whose beginning_stock is also
    0/blank — there's nothing to take a percentage of. The single source of
    truth for this, imported wherever a low-stock condition is checked, so
    the nav badge, Dashboard and Notifications can never drift apart."""
    from sqlalchemy import and_, or_

    qty = models.Product.beginning_stock + models.Product.stock_qty
    own_threshold = and_(
        models.Product.reorder_level > 0,
        qty > 0,
        qty <= models.Product.reorder_level,
    )
    if default_pct is None:
        return own_threshold
    fallback_threshold = and_(
        models.Product.reorder_level <= 0,
        models.Product.beginning_stock > 0,
        qty > 0,
        qty <= models.Product.beginning_stock * (Decimal(str(default_pct)) / Decimal("100")),
    )
    return or_(own_threshold, fallback_threshold)


SORTABLE_COLUMNS = {
    "beginning_stock": models.Product.beginning_stock,
    "stock_qty": models.Product.stock_qty,
    "total_qty": models.Product.beginning_stock + models.Product.stock_qty,
}


def _find_conversion_issue_ids(db: Session) -> set:
    """Product ids whose unit-breakdown display (Sealed packs / Open
    container — see Product.unit_breakdown) shows a negative quantity on
    any tier. That's physically impossible (you can't have -2 packs on a
    shelf), so it's a reliable sign the item's total is oversold and needs
    a physical recount — not a guess, just reading the same numbers already
    shown on the row, gathered up instead of left to be spotted by eye.
    Computed in Python since unit_breakdown is derived from total_qty + the
    unit ladder, not a queryable column, so this only ever looks at active
    products that actually have a units ladder to begin with."""
    candidates = (
        db.query(models.Product)
        .options(selectinload(models.Product.units))
        .filter(models.Product.is_active.is_(True), models.Product.units.any())
        .all()
    )
    ids = set()
    for p in candidates:
        breakdown = p.unit_breakdown
        if breakdown and any(tier["qty"] < 0 for tier in breakdown):
            ids.add(p.id)
    return ids


MANUAL_ADJUSTMENT_WINDOW_DAYS = 7


def _find_recent_manual_adjustment_ids(db: Session) -> set:
    """Product ids with a manual stock adjustment (Product Edit's Beginning/
    Stocks Qty change, reason="adjustment" — either direction, dagdag or
    bawas) in the last MANUAL_ADJUSTMENT_WINDOW_DAYS days. A plain rolling
    review window, not a correctness check — an item ages off this list on
    its own once its most recent adjustment is more than a week old, no
    manual cleanup needed."""
    cutoff = datetime.now(MANILA) - timedelta(days=MANUAL_ADJUSTMENT_WINDOW_DAYS)
    rows = (
        db.query(models.StockMovement.product_id.distinct())
        .filter(models.StockMovement.reason == "adjustment", models.StockMovement.created_at >= cutoff)
        .all()
    )
    return {row[0] for row in rows}


@router.get("/products", response_class=HTMLResponse)
def list_products(
    request: Request,
    q: str = "",
    page: int = 1,
    alert: int = 0,
    category_id: int = 0,
    subcategory_id: int = 0,
    shelf_id: int = 0,
    sort: str = "",
    sort_dir: str = "asc",
    bulk_msg: str = "",
    deleted: int = 0,
    delete_blocked: int = 0,
    delete_blocked_name: str = "",
    merged: str = "",
    merge_error: str = "",
    flag: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = (q or "").strip()
    page = max(page, 1)
    sort_dir = "desc" if sort_dir == "desc" else "asc"

    # Every product with an open double-deduction candidate, store-wide —
    # computed once per request and reused for the Void tab's filter, its
    # badge count, and the existing per-row ⚠ flags below, instead of
    # scanning for candidates three separate times.
    void_map = find_all_double_deduction_candidates(db) if is_staff(user) else {}
    void_ids = set(void_map.keys())

    # Every product whose Sealed/Open Container breakdown shows a negative
    # number (see _find_conversion_issue_ids) — same "compute once, reuse
    # for the tab's filter + badge" idea as void_ids above.
    conversion_ids = _find_conversion_issue_ids(db) if is_staff(user) else set()

    # Every product with a manual stock adjustment in the last week — see
    # _find_recent_manual_adjustment_ids.
    adjustment_ids = _find_recent_manual_adjustment_ids(db) if is_staff(user) else set()

    query = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if q:
        query = query.filter(
            multi_word_ilike(models.Product.name, q) | (models.Product.barcode == q)
        )
    if alert:
        # Only products whose selling price no longer covers cost.
        query = query.filter(
            models.Product.cost_price > 0,
            models.Product.selling_price <= models.Product.cost_price,
        )
    if category_id == -1:  # sentinel for "No category" — a real Category.id is never negative
        query = query.filter(models.Product.category_id.is_(None))
    elif category_id:
        query = query.filter(models.Product.category_id == category_id)
    if subcategory_id:
        query = query.filter(models.Product.subcategory_id == subcategory_id)
    if shelf_id == -1:  # sentinel for "No shelf" — a real Shelf.id is never negative
        query = query.filter(models.Product.shelf_id.is_(None))
    elif shelf_id:
        query = query.filter(models.Product.shelf_id == shelf_id)
    if flag == "void":
        query = query.filter(models.Product.id.in_(void_ids or {-1}))
    elif flag == "conversion":
        query = query.filter(models.Product.id.in_(conversion_ids or {-1}))
    elif flag == "adjustments":
        query = query.filter(models.Product.id.in_(adjustment_ids or {-1}))

    total = query.count()
    pages = max((total + PAGE_SIZE - 1) // PAGE_SIZE, 1)
    page = min(page, pages)

    qty_expr = models.Product.beginning_stock + models.Product.stock_qty
    total_cost_value, total_retail_value = query.with_entities(
        func.coalesce(func.sum(qty_expr * models.Product.cost_price), 0),
        func.coalesce(func.sum(qty_expr * models.Product.selling_price), 0),
    ).one()
    total_cost_value = Decimal(str(total_cost_value or 0))
    total_retail_value = Decimal(str(total_retail_value or 0))

    sort_col = SORTABLE_COLUMNS.get(sort)
    if sort_col is not None:
        # Name as a tiebreaker keeps ties (e.g. a bunch of zero-stock
        # products) in a stable, predictable order instead of shuffling on
        # every page load.
        order = (sort_col.desc() if sort_dir == "desc" else sort_col.asc(), models.Product.name)
    else:
        order = (models.Product.name,)
    products = (
        query.order_by(*order)
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )

    # Which of these products (only this page's worth) have a past sale that
    # likely double-deducted stock already reflected in a completed Stock
    # Count — reused from void_map above rather than recomputed.
    page_ids = {p.id for p in products}
    dd_flags = {
        pid: {
            "count": len(rows),
            "min_date": min(r["sale_date"] for r in rows).date().isoformat(),
            "max_date": max(r["sale_date"] for r in rows).date().isoformat(),
        }
        for pid, rows in void_map.items() if pid in page_ids
    }

    # Quick-filter pills: only categories actually in use, with a live count
    # each, computed against the current search/alert filter (not the
    # category filter itself) so switching pills reflects what's really there.
    base_for_counts = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if q:
        base_for_counts = base_for_counts.filter(
            multi_word_ilike(models.Product.name, q) | (models.Product.barcode == q)
        )
    if alert:
        base_for_counts = base_for_counts.filter(
            models.Product.cost_price > 0,
            models.Product.selling_price <= models.Product.cost_price,
        )
    if flag == "void":
        base_for_counts = base_for_counts.filter(models.Product.id.in_(void_ids or {-1}))
    elif flag == "conversion":
        base_for_counts = base_for_counts.filter(models.Product.id.in_(conversion_ids or {-1}))
    elif flag == "adjustments":
        base_for_counts = base_for_counts.filter(models.Product.id.in_(adjustment_ids or {-1}))
    cat_counts = dict(
        base_for_counts.filter(models.Product.category_id.isnot(None))
        .with_entities(models.Product.category_id, func.count(models.Product.id))
        .group_by(models.Product.category_id)
        .all()
    )
    categories = (
        db.query(models.Category)
        .filter(models.Category.id.in_(cat_counts.keys()))
        .order_by(models.Category.name)
        .all()
    ) if cat_counts else []
    no_category_count = base_for_counts.filter(models.Product.category_id.is_(None)).count()

    # Sub Category pills: scoped to the selected Category (if any) — same
    # live-count-against-everything-but-itself idea as the Category pills.
    base_for_subcounts = base_for_counts
    if category_id:
        base_for_subcounts = base_for_subcounts.filter(models.Product.category_id == category_id)
    subcat_counts = dict(
        base_for_subcounts.filter(models.Product.subcategory_id.isnot(None))
        .with_entities(models.Product.subcategory_id, func.count(models.Product.id))
        .group_by(models.Product.subcategory_id)
        .all()
    )
    subcategories = (
        db.query(models.SubCategory)
        .filter(models.SubCategory.id.in_(subcat_counts.keys()))
        .order_by(models.SubCategory.name)
        .all()
    ) if subcat_counts else []

    # Shelf pills: same live-count idea, independent of Category/Sub Category
    # (a shelf can hold products from any category).
    shelf_counts = dict(
        base_for_counts.filter(models.Product.shelf_id.isnot(None))
        .with_entities(models.Product.shelf_id, func.count(models.Product.id))
        .group_by(models.Product.shelf_id)
        .all()
    )
    shelves = (
        db.query(models.Shelf)
        .filter(models.Shelf.id.in_(shelf_counts.keys()))
        .order_by(models.Shelf.name)
        .all()
    ) if shelf_counts else []
    no_shelf_count = base_for_counts.filter(models.Product.shelf_id.is_(None)).count()

    # Full, unfiltered lists (not just the pills above, which only show
    # categories currently in use) — for the inline "set category" editor's
    # autocomplete, same "type or pick" datalist the Add/Edit form uses.
    all_categories = db.query(models.Category).order_by(models.Category.name).all()
    all_subcategories = db.query(models.SubCategory).order_by(models.SubCategory.name).all()

    return templates.TemplateResponse(
        "products/list.html",
        {
            "request": request,
            "app_name": request.app.title,
            "user": user,
            "products": products,
            "q": q,
            "page": page,
            "pages": pages,
            "total": total,
            "total_cost_value": total_cost_value,
            "total_retail_value": total_retail_value,
            "alert": alert,
            "bulk_msg": bulk_msg,
            "deleted": deleted,
            "delete_blocked": delete_blocked,
            "delete_blocked_name": delete_blocked_name,
            "merged": merged,
            "merge_error": merge_error,
            "category_id": category_id,
            "categories": categories,
            "cat_counts": cat_counts,
            "no_category_count": no_category_count,
            "all_categories": all_categories,
            "all_subcategories": all_subcategories,
            "subcategory_id": subcategory_id,
            "subcategories": subcategories,
            "subcat_counts": subcat_counts,
            "shelf_id": shelf_id,
            "shelves": shelves,
            "shelf_counts": shelf_counts,
            "no_shelf_count": no_shelf_count,
            "sort": sort,
            "sort_dir": sort_dir,
            "dd_flags": dd_flags,
            "flag": flag,
            "void_count": len(void_ids),
            "conversion_count": len(conversion_ids),
            "adjustment_count": len(adjustment_ids),
            "last_rollover_period": settings_store.get_setting(db, MONTH_END_SETTING_KEY, ""),
        },
    )


@router.get("/products/export.xlsx")
def export_products_excel(
    q: str = "",
    alert: int = 0,
    category_id: int = 0,
    subcategory_id: int = 0,
    shelf_id: int = 0,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Excel export of Inventory — respects whatever filter/search/category
    pill is active on screen, same as the list page itself."""
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = (q or "").strip()
    query = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if q:
        query = query.filter(multi_word_ilike(models.Product.name, q) | (models.Product.barcode == q))
    if alert:
        query = query.filter(models.Product.cost_price > 0, models.Product.selling_price <= models.Product.cost_price)
    if category_id == -1:  # sentinel for "No category" — a real Category.id is never negative
        query = query.filter(models.Product.category_id.is_(None))
    elif category_id:
        query = query.filter(models.Product.category_id == category_id)
    if subcategory_id:
        query = query.filter(models.Product.subcategory_id == subcategory_id)
    if shelf_id == -1:  # sentinel for "No shelf" — a real Shelf.id is never negative
        query = query.filter(models.Product.shelf_id.is_(None))
    elif shelf_id:
        query = query.filter(models.Product.shelf_id == shelf_id)
    products = query.order_by(models.Product.name).all()

    is_admin_user = is_staff(user)
    headers = ["Product Name", "Barcode", "Category", "Sub Category", "Unit Type", "Shelf", "Cost of Sales"]
    if is_admin_user:
        headers.append("Selling Price")
    headers += ["Actual Beginning", "Stocks Qty", "Total Qty"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for p in products:
        row = [
            p.name, p.barcode or "", p.category.name if p.category else "",
            p.subcategory.name if p.subcategory else "", p.unit_type.name if p.unit_type else "",
            p.shelf.name if p.shelf else "",
            float(p.cost_price or 0),
        ]
        if is_admin_user:
            row.append(float(p.selling_price or 0))
        row += [float(p.beginning_stock or 0), float(p.stock_qty or 0), float(p.total_qty or 0)]
        ws.append(row)

    widths = [28, 16, 16, 16, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="inventory.xlsx"'},
    )


def _render_form(request, db, user, product=None, error=None, back=""):
    categories = db.query(models.Category).order_by(models.Category.name).all()
    subcategories = db.query(models.SubCategory).order_by(models.SubCategory.name).all()
    unit_types = db.query(models.UnitType).order_by(models.UnitType.name).all()
    shelves = db.query(models.Shelf).order_by(models.Shelf.name).all()
    return templates.TemplateResponse(
        "products/form.html",
        {
            "request": request,
            "app_name": request.app.title,
            "user": user,
            "product": product,
            "categories": categories,
            "subcategories": subcategories,
            "unit_types": unit_types,
            "shelves": shelves,
            "adjustment_reasons": ADJUSTMENT_REASONS,
            "error": error,
            # Where Back/Cancel/after-save should land — the filtered list the
            # user came from, so a search isn't lost just because they edited
            # one row. Falls back to a plain /products when absent.
            "back": safe_back_url(back, "/products"),
        },
    )


@router.get("/products/new", response_class=HTMLResponse)
def new_product(request: Request, back: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    return _render_form(request, db, product=None, user=user, back=back)


@router.get("/products/{product_id:int}/edit", response_class=HTMLResponse)
def edit_product(product_id: int, request: Request, back: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    product = db.get(models.Product, product_id)
    if not product:
        return RedirectResponse("/products", status_code=302)
    return _render_form(request, db, user, product=product, back=back)


@router.post("/products/{product_id:int}/rename")
def rename_product(product_id: int, data: dict, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Quick rename straight from the Inventory list/search — a typo or a
    naming-convention fix (e.g. "PILL BOX 10 X 10" spacing) shouldn't need
    the full find -> Edit page -> Save -> search-again round trip."""
    if not user:
        return JSONResponse({"ok": False, "error": "Please sign in again."}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"ok": False, "error": "You don't have permission to rename products."}, status_code=403)
    product = db.get(models.Product, product_id)
    if not product:
        return JSONResponse({"ok": False, "error": "Product not found."}, status_code=404)

    new_name = (data.get("name") or "").strip()
    if not new_name:
        return JSONResponse({"ok": False, "error": "Name can't be empty."}, status_code=400)

    old_name = product.name
    if new_name != old_name:
        product.name = new_name
        audit.record(
            db, user=user, request=request, action="update", entity_type="product",
            entity_id=product.id, entity_label=product.name,
            summary=f"Renamed “{old_name}” → “{new_name}”",
            changes={"name": [old_name, new_name]},
        )
        db.commit()
    return {"ok": True, "name": product.name}


@router.post("/products/{product_id:int}/set-category")
def set_product_category(product_id: int, data: dict, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Set a product's Category/Sub Category straight from the Inventory
    list — same "type or pick" fields as the Edit form, just without the
    full find -> Edit page -> Save round trip. Meant for the common case of
    quickly filling in a product that has no category at all yet."""
    if not user:
        return JSONResponse({"ok": False, "error": "Please sign in again."}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"ok": False, "error": "You don't have permission to do this."}, status_code=403)
    product = db.get(models.Product, product_id)
    if not product:
        return JSONResponse({"ok": False, "error": "Product not found."}, status_code=404)

    old_category = product.category.name if product.category else None
    old_subcategory = product.subcategory.name if product.subcategory else None
    product.category = _get_or_create_category(db, data.get("category"))
    product.subcategory = _get_or_create_subcategory(db, data.get("subcategory"), product.category)
    new_category = product.category.name if product.category else None
    new_subcategory = product.subcategory.name if product.subcategory else None
    if new_category != old_category or new_subcategory != old_subcategory:
        audit.record(
            db, user=user, request=request, action="update", entity_type="product",
            entity_id=product.id, entity_label=product.name,
            summary=f"Set category for “{product.name}”: {old_category or '—'} / {old_subcategory or '—'} → {new_category or '—'} / {new_subcategory or '—'}",
            changes={"category": [old_category, new_category], "subcategory": [old_subcategory, new_subcategory]},
        )
        db.commit()
    return JSONResponse({"ok": True, "category": new_category or "", "subcategory": new_subcategory or ""})


@router.get("/products/{product_id:int}/double-deductions")
def product_double_deductions(product_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Candidate list for the ⚠ flag on the Inventory row — the actual sales
    behind that product's flag, for the review modal to show and pre-fill a
    From/To range from."""
    if not user or not is_staff(user):
        return JSONResponse({"candidates": []}, status_code=403)
    rows = find_double_deduction_candidates(db, [product_id]).get(product_id, [])
    return JSONResponse({
        "candidates": [
            {
                "movement_id": r["movement_id"],
                "qty_base": r["qty_base"],
                "invoice_no": r["invoice_no"],
                "sale_date": r["sale_date"].date().isoformat(),
                "count_ref": r["count_ref"],
                "count_date": r["count_date"].isoformat(),
            }
            for r in rows
        ],
    })


@router.post("/products/{product_id:int}/void-deductions")
def void_double_deductions(product_id: int, data: dict, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Reverses the stock effect of past sale(s) that likely double-deducted
    this product's counted stock (see app/double_deductions.py) — adds the
    exact quantity back, healing a negative Actual Beginning first the same
    way a Stock Count correction does. Recomputes the candidate list itself
    rather than trusting whatever the client sends, filtered to the chosen
    date range — so this can only ever touch real, still-open evidence, never
    something already corrected or no longer valid by the time it's clicked."""
    if not user:
        return JSONResponse({"ok": False, "error": "Please sign in again."}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"ok": False, "error": "You don't have permission to do this."}, status_code=403)
    product = db.get(models.Product, product_id)
    if not product:
        return JSONResponse({"ok": False, "error": "Product not found."}, status_code=404)

    try:
        from_date = date.fromisoformat((data.get("from_date") or "").strip())
        to_date = date.fromisoformat((data.get("to_date") or "").strip())
    except ValueError:
        return JSONResponse({"ok": False, "error": "Pick a valid date range."}, status_code=400)
    if from_date > to_date:
        return JSONResponse({"ok": False, "error": "From date must be on or before To date."}, status_code=400)

    candidates = [
        r for r in find_double_deduction_candidates(db, [product_id]).get(product_id, [])
        if from_date <= r["sale_date"].date() <= to_date
    ]
    if not candidates:
        return JSONResponse({"ok": False, "error": "No matching sale(s) found in that date range."}, status_code=400)

    from .pos import _apply_stock_count_correction
    total_qty = Decimal("0")
    for c in candidates:
        qty = Decimal(str(abs(c["qty_base"])))
        _apply_stock_count_correction(product, qty)
        db.add(models.StockMovement(
            product_id=product.id,
            qty_base=qty,
            reason="correction",
            ref=c["invoice_no"],
            note=f"Double-deduction void — {c['invoice_no']} ({c['sale_date'].date().isoformat()}), already reflected in {c['count_ref']}",
            corrects_movement_id=c["movement_id"],
        ))
        total_qty += qty

    audit.record(
        db, user=user, request=request, action="void_double_deduction", entity_type="product",
        entity_id=product.id, entity_label=product.name,
        summary=f"Voided {len(candidates)} double-deducted sale(s) dated {from_date}–{to_date}, restored {total_qty} unit(s)",
        changes={"corrected": [
            {"movement_id": c["movement_id"], "invoice_no": c["invoice_no"], "qty": c["qty_base"], "count_ref": c["count_ref"]}
            for c in candidates
        ]},
    )
    db.commit()
    return JSONResponse({"ok": True, "corrected": len(candidates), "qty_restored": float(total_qty)})


@router.get("/products/void-invoice")
def void_invoice_search(invoice: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Every still-open double-deduction candidate line whose invoice
    matches (case-insensitive) — for the Void tab's invoice-search tool, so
    one invoice can be found and corrected across every product on it at
    once, instead of only from a single product's own Inventory row."""
    if not user or not is_staff(user):
        return JSONResponse({"items": []}, status_code=403)
    needle = (invoice or "").strip().lower()
    if not needle:
        return JSONResponse({"items": []})

    void_map = find_all_double_deduction_candidates(db)
    matches = [
        (pid, r) for pid, rows in void_map.items() for r in rows
        if needle in r["invoice_no"].lower()
    ]
    product_ids = {pid for pid, _ in matches}
    names = dict(
        db.query(models.Product.id, models.Product.name)
        .filter(models.Product.id.in_(product_ids))
        .all()
    ) if product_ids else {}
    return JSONResponse({
        "items": [
            {
                "product_id": pid,
                "product_name": names.get(pid, ""),
                "movement_id": r["movement_id"],
                "qty_base": r["qty_base"],
                "sale_date": r["sale_date"].date().isoformat(),
                "count_ref": r["count_ref"],
                "count_date": r["count_date"].isoformat(),
            }
            for pid, r in matches
        ],
    })


@router.post("/products/void-by-movements")
def void_by_movements(data: dict, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Voids a specific set of double-deduction candidate movements, picked
    from the Void tab's invoice search (which can span several products at
    once) — same rule as /products/{id}/void-deductions, just addressed by
    movement id instead of a per-product date range. Recomputes candidates
    itself and only ever touches whichever of the requested ids are still
    real, open evidence at the moment this runs — never trusts the client's
    list beyond which ids it's asking to act on."""
    if not user:
        return JSONResponse({"ok": False, "error": "Please sign in again."}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"ok": False, "error": "You don't have permission to do this."}, status_code=403)

    requested_ids = set()
    for mid in (data.get("movement_ids") or []):
        try:
            requested_ids.add(int(mid))
        except (TypeError, ValueError):
            continue
    if not requested_ids:
        return JSONResponse({"ok": False, "error": "Nothing selected."}, status_code=400)

    void_map = find_all_double_deduction_candidates(db)
    candidates = [
        (pid, r) for pid, rows in void_map.items() for r in rows
        if r["movement_id"] in requested_ids
    ]
    if not candidates:
        return JSONResponse({"ok": False, "error": "Nothing matched — it may already be corrected."}, status_code=400)

    from .pos import _apply_stock_count_correction
    product_ids = {pid for pid, _ in candidates}
    products = {
        p.id: p for p in db.query(models.Product).filter(models.Product.id.in_(product_ids)).all()
    }
    total_qty = Decimal("0")
    invoices = set()
    changes = []
    for pid, c in candidates:
        product = products.get(pid)
        if not product:
            continue
        qty = Decimal(str(abs(c["qty_base"])))
        _apply_stock_count_correction(product, qty)
        db.add(models.StockMovement(
            product_id=product.id,
            qty_base=qty,
            reason="correction",
            ref=c["invoice_no"],
            note=f"Double-deduction void — {c['invoice_no']} ({c['sale_date'].date().isoformat()}), already reflected in {c['count_ref']}",
            corrects_movement_id=c["movement_id"],
        ))
        total_qty += qty
        invoices.add(c["invoice_no"])
        changes.append({
            "product_id": pid, "movement_id": c["movement_id"], "invoice_no": c["invoice_no"],
            "qty": c["qty_base"], "count_ref": c["count_ref"],
        })

    audit.record(
        db, user=user, request=request, action="void_double_deduction", entity_type="product",
        entity_id=None, entity_label=", ".join(sorted(invoices)),
        summary=f"Voided {len(changes)} double-deducted line(s) across invoice(s) {', '.join(sorted(invoices))}, restored {total_qty} unit(s)",
        changes={"corrected": changes},
    )
    db.commit()
    return JSONResponse({"ok": True, "corrected": len(changes), "qty_restored": float(total_qty)})


def _save_from_form(product: models.Product, db: Session, form):
    product.name = (form.get("name") or "").strip()
    product.barcode = (form.get("barcode") or "").strip() or None
    product.category = _get_or_create_category(db, form.get("category"))
    product.subcategory = _get_or_create_subcategory(db, form.get("subcategory"), product.category)
    product.unit_type = _get_or_create_unit_type(db, form.get("unit_type"))
    product.shelf = _get_or_create_shelf(db, form.get("shelf"))
    product.cost_price = _to_decimal(form.get("cost_price"))
    # Selling prices are set on the Add-product form and from then on only
    # through the dedicated Selling Price tab (see /products/pricing) — the
    # edit form here doesn't carry these fields, so leave the fixed price and
    # both percentages untouched. Markup/margin prices still get refreshed
    # against whatever the cost is now, so they don't go stale if it changed.
    markup_pct = form.get("markup_pct") if "markup_pct" in form else product.markup_pct
    margin_pct = form.get("margin_pct") if "margin_pct" in form else product.margin_pct
    if "selling_price" in form:
        product.selling_price = _to_decimal(form.get("selling_price"))   # the fixed price
    pricing.apply_to(product, product.cost_price, markup_pct, margin_pct)
    product.beginning_stock = _to_decimal(form.get("beginning_stock"))
    product.stock_qty = _to_decimal(form.get("stock_qty"))
    product.reorder_level = _to_decimal(form.get("reorder_level"))
    product.is_vat = bool(form.get("is_vat"))

    # Units ladder (extra sellable units). Parallel arrays from the form.
    # Each row's factor can be typed relative to another unit already on the
    # ladder (e.g. "1 Elf Load = 6 Sack") instead of only relative to base —
    # see _resolve_unit_chain for how that gets flattened into factor_to_base.
    names = form.getlist("unit_name")
    factors = form.getlist("unit_factor")
    prices = form.getlist("unit_price")
    relative_tos = form.getlist("unit_relative_to")
    unit_markup_pcts = form.getlist("unit_markup_pct")
    unit_margin_pcts = form.getlist("unit_margin_pct")

    rows = []
    for i, nm in enumerate(names):
        nm = (nm or "").strip()
        if not nm:
            continue
        fac = _to_decimal(factors[i] if i < len(factors) else "1", "1")
        if fac <= 0:
            fac = Decimal("1")
        pr = _to_decimal(prices[i] if i < len(prices) else "0")
        rel_to = (relative_tos[i] if i < len(relative_tos) else "").strip() or None
        u_markup_pct = _to_decimal(unit_markup_pcts[i] if i < len(unit_markup_pcts) else "0")
        u_margin_pct = _to_decimal(unit_margin_pcts[i] if i < len(unit_margin_pcts) else "0")
        rows.append({
            "name": nm, "relative_factor": fac, "price": pr, "relative_to": rel_to,
            "markup_pct": u_markup_pct, "margin_pct": u_margin_pct,
        })

    _apply_unit_rows(product, db, rows)


def _apply_unit_rows(product: models.Product, db: Session, rows: list):
    """Replace product's entire unit ladder with the given rows
    ({name, relative_factor, price, relative_to, markup_pct, margin_pct}),
    resolving any "relative to another unit" chain into factor_to_base.
    Shared by the product edit form and the Units & Conversions bulk import."""
    # Break any self-referential link among the OLD units (e.g. Sack ->
    # Elf Load) before they get deleted below — with the link still in
    # place, SQLAlchemy can't decide which ProductUnit row to delete first
    # and raises CircularDependencyError. Nulling it out first removes the
    # cycle so the deletes can proceed in any order.
    for existing in product.units:
        existing.relative_to_unit_id = None
    db.flush()
    product.units.clear()

    base_name = product.unit_type.name if product.unit_type else None
    filtered = []
    for r in rows:
        # A ladder row named the same as the base unit is meaningless (the
        # base already exists as its own row-less entry) and produces a
        # confusing duplicate in every unit picker — drop it rather than
        # save it.
        if base_name and r["name"].lower() == base_name.lower():
            continue
        rel_to = r["relative_to"]
        if rel_to and base_name and rel_to.lower() == base_name.lower():
            rel_to = None  # typing the base unit's own name just means "relative to base"
        filtered.append({**r, "relative_to": rel_to})

    resolved = _resolve_unit_chain(filtered)

    order = 0
    name_to_unit = {}
    for r in filtered:
        unit_factor = resolved[r["name"]]
        unit_cost = (product.cost_price or Decimal("0")) * unit_factor
        pu = models.ProductUnit(
            name=r["name"], price=r["price"], sort_order=order,
            factor_to_base=unit_factor, relative_factor=r["relative_factor"],
            markup_pct=r["markup_pct"], markup_price=pricing.markup_price(unit_cost, r["markup_pct"]),
            margin_pct=r["margin_pct"], margin_price=pricing.margin_price(unit_cost, r["margin_pct"]),
        )
        product.units.append(pu)
        name_to_unit[r["name"]] = pu
        order += 1
    db.flush()  # assign ids so relative_to_unit_id below can point at them
    for r in filtered:
        if r["relative_to"] and r["relative_to"] in name_to_unit:
            name_to_unit[r["name"]].relative_to_unit_id = name_to_unit[r["relative_to"]].id


def _resolve_unit_chain(rows: list) -> dict:
    """rows: [{name, relative_factor, relative_to}]. Returns {name: factor_to_base},
    flattening any chain (A relative to B relative to base) into a single
    base-relative number via memoized recursion. Falls back to treating a row
    as base-relative if its parent is missing or a cycle is detected, so a
    typo in "relative to" never breaks saving."""
    by_name = {r["name"]: r for r in rows}
    resolved = {}

    def resolve(name, visiting):
        if name in resolved:
            return resolved[name]
        row = by_name.get(name)
        if not row:
            return Decimal("1")
        rel_to = row["relative_to"]
        if not rel_to or rel_to == name or rel_to in visiting:
            value = row["relative_factor"]
        else:
            value = row["relative_factor"] * resolve(rel_to, visiting | {name})
        resolved[name] = value
        return value

    for r in rows:
        resolve(r["name"], set())
    return resolved


@router.post("/products")
async def create_product(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    form = await request.form()
    back = (form.get("back") or "").strip()

    # The DB work below is synchronous SQLAlchemy — running it directly in
    # this async function would block the whole app for every other user
    # for as long as it takes. run_in_threadpool moves it off the event
    # loop, same as any plain `def` route gets for free; only the form
    # parsing above (which genuinely needs await) stays on the loop.
    def _do():
        if not (form.get("name") or "").strip():
            return _render_form(request, db, user, product=None, error="Product name is required.", back=back)
        barcode = (form.get("barcode") or "").strip()
        if barcode and db.query(models.Product).filter(models.Product.barcode == barcode).first():
            return _render_form(request, db, user, product=None, error=f"Barcode “{barcode}” is already assigned to another product.", back=back)
        product = models.Product()
        db.add(product)  # add before _save_from_form so its internal flush (for the units-ladder chain) can assign ids
        _save_from_form(product, db, form)
        db.flush()  # assign product.id so the audit row can reference it
        audit.record(
            db, user=user, request=request, action="create", entity_type="product",
            entity_id=product.id, entity_label=product.name,
            summary=f"Created product “{product.name}”",
        )
        db.commit()
        return RedirectResponse(safe_back_url(back, "/products"), status_code=status.HTTP_302_FOUND)

    return await run_in_threadpool(_do)


@router.post("/products/{product_id:int}")
async def update_product(product_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    form = await request.form()

    # See create_product just above — same reason for the threadpool hop.
    def _do():
        product = db.get(models.Product, product_id)
        if not product:
            return RedirectResponse("/products", status_code=302)
        back = form.get("back") or ""
        if not (form.get("name") or "").strip():
            return _render_form(request, db, user, product=product, error="Product name is required.", back=back)
        barcode = (form.get("barcode") or "").strip()
        if barcode and db.query(models.Product).filter(models.Product.barcode == barcode, models.Product.id != product.id).first():
            return _render_form(request, db, user, product=product, error=f"Barcode “{barcode}” is already assigned to another product.", back=back)
        old_total = Decimal(str(product.total_qty or 0))
        new_total_preview = _to_decimal(form.get("beginning_stock")) + _to_decimal(form.get("stock_qty"))
        adjustment_reason = (form.get("adjustment_reason") or "").strip()
        if new_total_preview != old_total and not adjustment_reason:
            return _render_form(request, db, user, product=product, back=back,
                                 error="Select a reason for the stock quantity change before saving.")

        before = _product_snapshot(product)
        _save_from_form(product, db, form)
        db.flush()
        after = _product_snapshot(product)
        new_total = Decimal(str(product.total_qty or 0))
        changes = audit.diff(before, after)
        if changes:
            # Flag a stock correction distinctly — it's the theft-sensitive edit.
            stock_touched = "stock_qty" in changes or "beginning_stock" in changes
            audit.record(
                db, user=user, request=request,
                action="adjust_stock" if stock_touched else "update",
                entity_type="product", entity_id=product.id, entity_label=product.name,
                summary=(f"Adjusted stock for “{product.name}”" if stock_touched else f"Edited “{product.name}”"),
                changes=changes,
            )
            # Record the net stock change as a movement too, so the Stock Card
            # ledger reconciles — manual edits used to leave no trace here. Value
            # it at current cost so it also shows up as shrinkage/gain in P&L.
            delta = new_total - old_total
            if delta != 0:
                unit_cost = Decimal(str(product.cost_price or 0))
                reason_label = ADJUSTMENT_REASON_LABELS.get(adjustment_reason, adjustment_reason or "manual edit")
                note_text = (form.get("adjustment_note") or "").strip()
                db.add(models.StockMovement(
                    product_id=product.id, qty_base=delta, reason="adjustment", ref="manual edit",
                    unit_cost=unit_cost, value=delta * unit_cost,
                    note=f"{reason_label}: {note_text}" if note_text else reason_label,
                ))
        db.commit()
        return RedirectResponse(safe_back_url(back, "/products"), status_code=status.HTTP_302_FOUND)

    return await run_in_threadpool(_do)


@router.get("/products/pricing", response_class=HTMLResponse)
def pricing_tool(request: Request, review: int = 0, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """The single unified costing + pricing workflow: search or browse on
    the left (with a 'Price Review Needed' filter folded in as a toggle,
    not a separate tab), edit cost + the 3-row pricing matrix on the right."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    min_margin = settings_store.min_margin_pct()
    review_count = db.query(func.count(models.Product.id)).filter(
        models.Product.is_active.is_(True), _needs_review_expr(min_margin)
    ).scalar() or 0
    return templates.TemplateResponse(
        "products/pricing.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "review_count": review_count,
            "initial_review": bool(review),
        },
    )


@router.get("/products/price-search")
def price_search(
    q: str = "",
    review: int = 0,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"error": "forbidden"}, status_code=403)
    q = (q or "").strip()
    min_margin = settings_store.min_margin_pct()

    review_count = db.query(func.count(models.Product.id)).filter(
        models.Product.is_active.is_(True), _needs_review_expr(min_margin)
    ).scalar() or 0

    query = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if review:
        query = query.filter(_needs_review_expr(min_margin))
    if q:
        barcode_hit = query.filter(models.Product.barcode == q).first()
        products = [barcode_hit] if barcode_hit else (
            query.filter(multi_word_ilike(models.Product.name, q)).order_by(models.Product.name).limit(100).all()
        )
    else:
        products = query.order_by(models.Product.name).limit(200 if review else 60).all()

    rows = []
    for p in products:
        flagged, reason = pricing.needs_review(p.selling_price, p.cost_price, min_margin)
        rows.append({
            "id": p.id, "name": p.name, "barcode": p.barcode,
            "cost_price": float(p.cost_price or 0),
            "selling_price": float(p.selling_price or 0),
            "markup_pct": float(p.markup_pct or 0),
            "margin_pct": float(p.margin_pct or 0),
            "needs_review": flagged, "review_reason": reason,
        })
    return {"products": rows, "review_count": review_count}


@router.post("/products/{product_id:int}/pricing")
def update_pricing(product_id: int, data: dict, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)
    product = db.get(models.Product, product_id)
    if not product:
        return JSONResponse({"ok": False, "error": "Product not found."}, status_code=404)
    before = _product_snapshot(product)
    if "cost_price" in data:
        product.cost_price = _to_decimal(data.get("cost_price"))
    product.selling_price = _to_decimal(data.get("selling_price"))
    pricing.apply_to(product, product.cost_price, data.get("markup_pct"), data.get("margin_pct"))
    db.flush()
    after = _product_snapshot(product)
    changes = audit.diff(before, after)
    if changes:
        audit.record(
            db, user=user, request=request, action="update",
            entity_type="product", entity_id=product.id, entity_label=product.name,
            summary=f"Updated selling price for “{product.name}”",
            changes=changes,
        )
    db.commit()
    min_margin = settings_store.min_margin_pct()
    flagged, reason = pricing.needs_review(product.selling_price, product.cost_price, min_margin)
    return {
        "ok": True,
        "cost_price": float(product.cost_price or 0),
        "selling_price": float(product.selling_price or 0),
        "markup_pct": float(product.markup_pct or 0),
        "margin_pct": float(product.margin_pct or 0),
        "needs_review": flagged, "review_reason": reason,
    }


@router.post("/products/{product_id:int}/archive")
def archive_product(product_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    product = db.get(models.Product, product_id)
    if product:
        product.is_active = False
        audit.record(
            db, user=user, request=request, action="archive", entity_type="product",
            entity_id=product.id, entity_label=product.name,
            summary=f"Archived “{product.name}”",
        )
        db.commit()
    return RedirectResponse("/products", status_code=status.HTTP_302_FOUND)


def _product_has_history(db: Session, product_id: int) -> bool:
    """True if this product is referenced anywhere real activity would be
    lost from — a sale, purchase, quotation, stock movement, stock count, or
    month-end rollover. Those foreign keys aren't ON DELETE CASCADE (on
    purpose — the whole point is that a sale/adjustment from three months ago
    must never silently disappear), so a hard DELETE on a product with any of
    these would either fail outright or, worse, quietly erase history. Only a
    product nobody has ever transacted against is safe to truly delete."""
    checks = [
        (models.SaleLine, models.SaleLine.product_id),
        (models.QuotationLine, models.QuotationLine.product_id),
        (models.PurchaseLine, models.PurchaseLine.product_id),
        (models.StockMovement, models.StockMovement.product_id),
        (models.StockCountLine, models.StockCountLine.product_id),
        (models.MonthEndRolloverLine, models.MonthEndRolloverLine.product_id),
    ]
    for model, col in checks:
        if db.query(model.id).filter(col == product_id).first():
            return True
    return False


@router.post("/products/{product_id:int}/delete")
def delete_product(product_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Permanently remove a product — only allowed when it has never been
    transacted against (see _product_has_history). Anything with real
    activity has to go through Archive instead, which is reversible and
    doesn't touch its history."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    product = db.get(models.Product, product_id)
    if not product:
        return RedirectResponse("/products", status_code=status.HTTP_302_FOUND)

    if _product_has_history(db, product.id):
        return RedirectResponse(
            f"/products?delete_blocked={product.id}&delete_blocked_name={product.name}",
            status_code=status.HTTP_302_FOUND,
        )

    name = product.name
    audit.record(
        db, user=user, request=request, action="delete", entity_type="product",
        entity_id=product.id, entity_label=name,
        summary=f"Deleted “{name}” (never had any activity)",
    )
    db.delete(product)
    db.commit()
    return RedirectResponse("/products?deleted=1", status_code=status.HTTP_302_FOUND)


_DUP_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_DUP_WS_RE = re.compile(r"\s+")
DUPLICATE_SIMILARITY_THRESHOLD = 0.92


def _normalize_dup_name(name: str) -> str:
    """Lowercased, punctuation stripped to spaces, whitespace collapsed —
    "Coke 1.5L" and "coke 1.5 l" normalize to the same string, so a straight
    equality check catches spacing/punctuation/case variants of the same
    typed name without any similarity math needed."""
    s = (name or "").lower()
    s = _DUP_PUNCT_RE.sub(" ", s)
    s = _DUP_WS_RE.sub(" ", s).strip()
    return s


def _find_duplicate_groups(products: list):
    """Groups active products that look like the same item entered more than
    once. Two passes, cheapest first:

    1. Exact match on the normalized name — spacing/punctuation/case is the
       only difference, so these are near-certain duplicates.
    2. Everything left over is bucketed by its first normalized word (a
       cheap, near-free grouping step), and only products sharing that first
       word are ever compared to each other with difflib — comparing every
       product against every other product would be ~2,000,000 comparisons
       at 1,000 products; bucketing first keeps each comparison set small.

    Returns a list of {"kind": "exact"|"near", "score": float, "products": [...]}
    dicts, one per group of 2+ products — singletons (nothing looks like a
    duplicate) are never included."""
    by_norm: dict[str, list] = {}
    for p in products:
        by_norm.setdefault(_normalize_dup_name(p.name), []).append(p)

    groups = []
    used_ids = set()
    for norm, members in by_norm.items():
        if len(members) > 1 and norm:
            groups.append({"kind": "exact", "score": 1.0, "products": members})
            used_ids.update(p.id for p in members)

    buckets: dict[str, list] = {}
    for p in products:
        if p.id in used_ids:
            continue
        norm = _normalize_dup_name(p.name)
        if not norm:
            continue
        first_word = norm.split(" ", 1)[0]
        buckets.setdefault(first_word, []).append((p, norm))

    for bucket in buckets.values():
        if len(bucket) < 2:
            continue
        n = len(bucket)
        parent = list(range(n))

        def find(i):
            while parent[i] != i:
                parent[i] = parent[parent[i]]
                i = parent[i]
            return i

        def union(i, j):
            ri, rj = find(i), find(j)
            if ri != rj:
                parent[ri] = rj

        pair_scores = {}
        for i in range(n):
            for j in range(i + 1, n):
                ratio = SequenceMatcher(None, bucket[i][1], bucket[j][1]).ratio()
                if ratio >= DUPLICATE_SIMILARITY_THRESHOLD:
                    union(i, j)
                    pair_scores[(i, j)] = ratio

        clusters: dict[int, list] = {}
        for i in range(n):
            clusters.setdefault(find(i), []).append(i)

        for idxs in clusters.values():
            if len(idxs) < 2:
                continue
            idx_set = set(idxs)
            scores = [s for (i, j), s in pair_scores.items() if i in idx_set and j in idx_set]
            groups.append({
                "kind": "near",
                "score": min(scores) if scores else DUPLICATE_SIMILARITY_THRESHOLD,
                "products": [bucket[i][0] for i in idxs],
            })

    groups.sort(key=lambda g: (-len(g["products"]), -g["score"]))
    return groups


@router.get("/products/duplicates", response_class=HTMLResponse)
def find_duplicates(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Scans every active product for likely duplicates (same item typed in
    more than once under a slightly different name/spacing) so reviewing a
    ~1,000-product catalogue for dupes is a short candidate list instead of a
    name-by-name manual search. Purely a report — nothing merges until the
    admin picks a keeper and confirms, via the same /products/merge endpoint
    Merge Duplicates already uses."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)

    products = (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True))
        .order_by(models.Product.name)
        .all()
    )
    groups = [
        {
            "kind": g["kind"],
            "score": g["score"],
            "products": [
                {
                    "id": p.id, "name": p.name, "barcode": p.barcode or "",
                    "category": p.category.name if p.category else "Uncategorized",
                    "qty": float(p.total_qty or 0),
                    "cost_price": float(p.cost_price or 0),
                }
                for p in g["products"]
            ],
        }
        for g in _find_duplicate_groups(products)
    ]
    return templates.TemplateResponse(
        "products/duplicates.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "groups": groups, "scanned": len(products),
        },
    )


@router.get("/products/merge/search")
def merge_search(q: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Product picker for the Merge Duplicates modal — same fuzzy name match
    as the Inventory list search, just a small JSON payload for a dropdown."""
    if not user or not is_staff(user):
        return JSONResponse({"products": []}, status_code=401)
    q = (q or "").strip()
    if not q:
        return {"products": []}
    products = (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True), multi_word_ilike(models.Product.name, q))
        .order_by(models.Product.name)
        .limit(20)
        .all()
    )
    return {
        "products": [
            {
                "id": p.id, "name": p.name,
                "category": p.category.name if p.category else "Uncategorized",
                "unit_name": p.unit_type.name if p.unit_type else "unit",
                "cost_price": float(p.cost_price or 0),
                "total_qty": float(p.total_qty or 0),
                "barcode": p.barcode or "",
            }
            for p in products
        ]
    }


# Every historical table a merge needs to re-point from the duplicate onto
# the surviving product, so past sales/purchases/adjustments keep reading
# correctly under the product that's left. Same set _product_has_history
# checks, since those are exactly the tables where losing the link would
# erase real activity.
_MERGE_HISTORY_TABLES = [
    models.SaleLine, models.QuotationLine, models.PurchaseLine,
    models.StockMovement, models.StockCountLine, models.MonthEndRolloverLine,
]


@router.post("/products/merge")
def merge_products(
    request: Request, keep_id: int = Form(...), merge_id: int = Form(...),
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    if keep_id == merge_id:
        return RedirectResponse("/products?merge_error=Pick+two+different+products", status_code=status.HTTP_302_FOUND)

    keep = db.get(models.Product, keep_id)
    dup = db.get(models.Product, merge_id)
    if not keep or not dup or not keep.is_active or not dup.is_active:
        return RedirectResponse("/products?merge_error=Product+not+found", status_code=status.HTTP_302_FOUND)

    keep.beginning_stock = (keep.beginning_stock or 0) + (dup.beginning_stock or 0)
    keep.stock_qty = (keep.stock_qty or 0) + (dup.stock_qty or 0)

    if not keep.barcode and dup.barcode:
        keep.barcode = dup.barcode
        dup.barcode = None

    for model in _MERGE_HISTORY_TABLES:
        db.query(model).filter(model.product_id == merge_id).update(
            {model.product_id: keep_id}, synchronize_session=False
        )
    db.query(models.ProductUnit).filter(models.ProductUnit.product_id == merge_id).delete(synchronize_session=False)

    dup.is_active = False
    audit.record(
        db, user=user, request=request, action="merge", entity_type="product",
        entity_id=keep.id, entity_label=keep.name,
        summary=f"Merged “{dup.name}” (#{dup.id}) into “{keep.name}” (#{keep.id}) — stock, sales, purchase and stock-count history moved over; “{dup.name}” archived",
    )
    db.commit()
    return RedirectResponse(f"/products?merged={keep.name}", status_code=status.HTTP_302_FOUND)


@router.get("/products/archived", response_class=HTMLResponse)
def list_archived_products(
    request: Request, q: str = "", db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    q = (q or "").strip()
    query = db.query(models.Product).filter(models.Product.is_active.is_(False))
    if q:
        query = query.filter(
            multi_word_ilike(models.Product.name, q) | (models.Product.barcode == q)
        )
    products = query.order_by(models.Product.name).all()
    return templates.TemplateResponse(
        "products/archived.html",
        {"request": request, "app_name": request.app.title, "user": user, "products": products, "q": q},
    )


@router.post("/products/{product_id:int}/restore")
def restore_product(product_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    product = db.get(models.Product, product_id)
    if product:
        product.is_active = True
        audit.record(
            db, user=user, request=request, action="restore", entity_type="product",
            entity_id=product.id, entity_label=product.name,
            summary=f"Restored “{product.name}” from archive",
        )
        db.commit()
    return RedirectResponse("/products/archived", status_code=status.HTTP_302_FOUND)


def _bulk_mode_label(mode: str, value: Decimal) -> str:
    sign = "+" if value >= 0 else ""
    if mode == "pct":
        return f"Bulk price update: {sign}{value:g}% on Fixed Price"
    if mode == "amount":
        return f"Bulk price update: {sign}{value:g} on Fixed Price"
    if mode == "cost_pct":
        return f"Bulk price update: Fixed Price set to Cost {sign}{value:g}%"
    if mode == "markup":
        return f"Bulk price update: Markup set to {value:g}%"
    return f"Bulk price update: Margin set to {value:g}%"


@router.post("/products/bulk-price", response_class=HTMLResponse)
def bulk_price_start(request: Request, ids: list[str] = Form([]), db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Step 1: show the selected products with a live client-side preview.
    Nothing is saved here — Apply (below) recomputes and saves for real."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    ids = sorted({int(i) for i in ids if i.isdigit()})
    if not ids:
        return RedirectResponse("/products", status_code=302)
    products = (
        db.query(models.Product)
        .filter(models.Product.id.in_(ids), models.Product.is_active.is_(True))
        .order_by(models.Product.name)
        .all()
    )
    return templates.TemplateResponse(
        "products/bulk_price.html",
        {"request": request, "app_name": request.app.title, "user": user, "products": products},
    )


@router.post("/products/bulk-price/apply")
async def bulk_price_apply(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Step 2: the server recomputes every price from scratch (never trusts
    numbers the browser sent) and saves, logging one audit row per product
    that actually changed — so this shows up on that product's normal price
    history exactly like a one-off edit would."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    form = await request.form()

    # Field names are dynamic (new_price_<id> per row), so this can't be
    # declared as fixed Form(...) params — instead the DB work (the actual
    # blocking part) runs off the event loop via run_in_threadpool, same
    # effect as converting to a plain `def` route.
    def _do():
        ids = sorted({int(i) for i in form.getlist("ids") if i.isdigit()})
        mode = (form.get("mode") or "").strip()
        if mode not in BULK_MODES:
            mode = "pct"
        try:
            value = Decimal(str(form.get("value") or "0").strip().replace(",", ""))
        except InvalidOperation:
            value = Decimal("0")

        if not ids:
            return RedirectResponse("/products", status_code=302)

        products = (
            db.query(models.Product)
            .filter(models.Product.id.in_(ids), models.Product.is_active.is_(True))
            .all()
        )
        label = _bulk_mode_label(mode, value)
        updated = 0
        skipped = 0
        for p in products:
            if mode in ("markup", "margin", "cost_pct") and (not p.cost_price or p.cost_price <= 0):
                skipped += 1
                continue
            before = _product_snapshot(p)

            # For the three Fixed-Price modes, the row's own editable field wins
            # when present — same trust level as a normal single-product price
            # edit (an admin explicitly typed this number), so it overrides the
            # uniform mode/value calculation for that one product only.
            override = None
            if mode in ("pct", "amount", "cost_pct"):
                raw_override = (form.get(f"new_price_{p.id}") or "").strip()
                if raw_override:
                    try:
                        override = max(Decimal(raw_override), Decimal("0")).quantize(CENTS, rounding=ROUND_HALF_UP)
                    except InvalidOperation:
                        override = None

            if mode == "pct":
                new_price = override if override is not None else (
                    Decimal(str(p.selling_price or 0)) * (1 + value / 100)
                ).quantize(CENTS, rounding=ROUND_HALF_UP)
                p.selling_price = max(new_price, Decimal("0"))
            elif mode == "amount":
                new_price = override if override is not None else (
                    Decimal(str(p.selling_price or 0)) + value
                ).quantize(CENTS, rounding=ROUND_HALF_UP)
                p.selling_price = max(new_price, Decimal("0"))
            elif mode == "cost_pct":
                # Same math as Markup, but writes straight into Fixed Price —
                # for the common case of seeding a never-set Fixed Price in bulk.
                new_price = override if override is not None else pricing.markup_price(p.cost_price, value)
                p.selling_price = max(new_price, Decimal("0"))
            elif mode == "markup":
                p.markup_pct = value
                p.markup_price = pricing.markup_price(p.cost_price, value)
            else:  # margin
                p.margin_pct = value
                p.margin_price = pricing.margin_price(p.cost_price, value)
            after = _product_snapshot(p)
            changes = audit.diff(before, after)
            if changes:
                audit.record(
                    db, user=user, request=request, action="update", entity_type="product",
                    entity_id=p.id, entity_label=p.name, summary=label, changes=changes,
                )
                updated += 1
        db.commit()
        msg = f"Updated {updated} product{'s' if updated != 1 else ''}"
        if skipped:
            msg += f", skipped {skipped} with no cost on file"
        return RedirectResponse(f"/products?bulk_msg={msg}", status_code=status.HTTP_302_FOUND)

    return await run_in_threadpool(_do)


def _generate_internal_barcode(product_id: int) -> str:
    """For products with no manufacturer barcode (locally-sourced, repackaged
    goods) — a locally-unique code we invent and print ourselves. No need to
    mimic a real EAN-13/UPC checksum scheme; Code128 (used to render it)
    reads any digit/letter string fine."""
    return f"LC{product_id:08d}"


def _barcode_data_uri(code: str) -> str:
    """Render `code` as a Code128 barcode (works for any barcode we already
    have on file — manufacturer EAN-13 digits included, Code128 just encodes
    them as text) and return it as an inline data: URI, ready for an <img>."""
    buf = io.BytesIO()
    bc = barcode_lib.get("code128", code, writer=SVGWriter())
    bc.write(buf, options={"write_text": False, "quiet_zone": 2})
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/svg+xml;base64,{b64}"


@router.post("/products/labels", response_class=HTMLResponse)
def print_labels(request: Request, ids: list[str] = Form([]), db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    ids = sorted({int(i) for i in ids if i.isdigit()})
    if not ids:
        return RedirectResponse("/products", status_code=302)
    products = (
        db.query(models.Product)
        .filter(models.Product.id.in_(ids), models.Product.is_active.is_(True))
        .order_by(models.Product.name)
        .all()
    )
    assigned = False
    for p in products:
        if not p.barcode:
            p.barcode = _generate_internal_barcode(p.id)
            assigned = True
    if assigned:
        db.commit()

    labels = [{"product": p, "barcode_uri": _barcode_data_uri(p.barcode)} for p in products]
    return templates.TemplateResponse(
        "products/labels.html",
        {"request": request, "app_name": request.app.title, "user": user, "labels": labels},
    )


@router.post("/products/bulk-shelf", response_class=HTMLResponse)
def bulk_shelf_start(request: Request, ids: list[str] = Form([]), db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Step 1: confirm which shelf, for the products selected in Inventory —
    same two-step shape as Bulk Price Update. Same idea Stock Count already
    has (assign a shelf to whatever's being counted), just reachable from
    Inventory directly instead of only while a count is open."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    ids = sorted({int(i) for i in ids if i.isdigit()})
    if not ids:
        return RedirectResponse("/products", status_code=302)
    products = (
        db.query(models.Product)
        .filter(models.Product.id.in_(ids), models.Product.is_active.is_(True))
        .order_by(models.Product.name)
        .all()
    )
    shelves = db.query(models.Shelf).order_by(models.Shelf.name).all()
    return templates.TemplateResponse(
        "products/bulk_shelf.html",
        {"request": request, "app_name": request.app.title, "user": user, "products": products, "shelves": shelves},
    )


@router.post("/products/bulk-shelf/apply")
def bulk_shelf_apply(
    request: Request, ids: list[str] = Form([]), shelf_id: str = Form("0"),
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    """Step 2: move every selected product onto the picked shelf — this
    REPLACES whatever shelf a product already had (unlike Stock Count's own
    bulk-assign, which only fills products that have none yet; here the
    point is a deliberate re-shelve of a chosen batch, so overwriting is the
    correct default)."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    ids = sorted({int(i) for i in ids if i.isdigit()})
    if not ids:
        return RedirectResponse("/products", status_code=302)
    try:
        shelf_id = int(shelf_id or 0)
    except ValueError:
        shelf_id = 0
    shelf = db.get(models.Shelf, shelf_id) if shelf_id else None
    if not shelf:
        return RedirectResponse("/products?bulk_msg=Pick+a+shelf+first.", status_code=status.HTTP_302_FOUND)

    products = (
        db.query(models.Product)
        .filter(models.Product.id.in_(ids), models.Product.is_active.is_(True))
        .all()
    )
    moved = 0
    for p in products:
        if p.shelf_id == shelf.id:
            continue
        old_shelf_name = p.shelf.name if p.shelf else "no shelf"
        p.shelf = shelf
        audit.record(
            db, user=user, request=request, action="update", entity_type="product",
            entity_id=p.id, entity_label=p.name,
            summary=f"Bulk-assigned “{p.name}” to shelf {shelf.name}",
            changes={"shelf": [old_shelf_name, shelf.name]},
        )
        moved += 1
    db.commit()
    msg = f"Moved {moved} product{'s' if moved != 1 else ''} to {shelf.name}"
    if moved < len(products):
        msg += f" ({len(products) - moved} already there)"
    return RedirectResponse(f"/products?bulk_msg={msg}", status_code=status.HTTP_302_FOUND)


# Human labels + in/out direction for the stock-movement reasons written across
# POS (sale/refund/exchange), Purchasing (purchase/return) and manual edits.
MOVEMENT_LABELS = {
    "sale": "Sale", "refund": "Refund (returned)",
    "exchange-return": "Exchange — returned in", "exchange-sale": "Exchange — sold out",
    "purchase": "Purchase received", "purchase-return": "Purchase return",
    "adjustment": "Manual adjustment", "void": "Sale voided",
    "sale-edit-reverse": "Item correction (reversed)",
    "correction": "Double-deduction void",
    "unvoid": "Sale restored",
}


@router.get("/products/{product_id:int}/stock-card", response_class=HTMLResponse)
def stock_card(product_id: int, request: Request, back: str = "", unit: int = 0, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Per-product stock ledger: every in/out movement with a running balance.

    The balance is anchored to the product's *current* on-hand total and worked
    backwards through the movements, so the newest row always shows the true
    current stock even if older movements pre-date movement tracking (e.g. a
    bulk import). The implied opening balance is shown for reconciliation.

    `unit`, when given, is one of this product's ProductUnit ids — everything
    stock-related (In/Out/Balance, the stat tiles) is shown converted into
    that unit instead of base (dividing by its factor_to_base, same math as
    ProductUnit.stock_as_this_unit), for a product sold in a bigger unit than
    it's tracked in. This never changes what's actually stored — base units
    remain the only thing written anywhere — purely a different lens on the
    same movements.
    """
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    product = db.get(models.Product, product_id)
    if not product:
        return RedirectResponse("/products", status_code=302)

    view_unit = next((u for u in product.units if u.id == unit), None) if unit else None
    view_factor = Decimal(str(view_unit.factor_to_base)) if view_unit and view_unit.factor_to_base else Decimal("1")
    view_unit_name = view_unit.name if view_unit else (product.unit_type.name if product.unit_type else "base unit")

    # This exact Stock Card view — unit + wherever it was itself opened
    # from — so a sale opened from here can send someone back to it
    # specifically, instead of dumping them on a blank new sale.
    self_url = f"/products/{product_id}/stock-card"
    self_qs = []
    if unit:
        self_qs.append(f"unit={unit}")
    if back:
        self_qs.append(f"back={quote(back)}")
    if self_qs:
        self_url += "?" + "&".join(self_qs)

    movements = (
        db.query(models.StockMovement)
        .filter(models.StockMovement.product_id == product_id)
        .order_by(models.StockMovement.created_at.asc(), models.StockMovement.id.asc())
        .all()
    )

    current_total = Decimal(str(product.total_qty or 0))
    total_delta = sum((Decimal(str(m.qty_base or 0)) for m in movements), Decimal("0"))
    opening = current_total - total_delta

    # What each sale-driven movement actually sold for. A StockMovement only
    # ever carries cost, never price — the price lives on the sale's own line,
    # so it's looked up here by the invoice # the movement already carries in
    # `ref`. Read from the sale rather than the product's current
    # selling_price on purpose: a row then shows what was really charged that
    # day, even after the item is re-priced later. Movements written before
    # `ref` was populated have nothing to match on and simply show no price.
    priced_by_ref = {}
    sale_id_by_ref = {}
    for line, sale in (
        db.query(models.SaleLine, models.Sale)
        .join(models.Sale, models.SaleLine.sale_id == models.Sale.id)
        .filter(models.SaleLine.product_id == product_id)
        .all()
    ):
        key = f"{sale.receipt_type}{sale.invoice_no}" if sale.receipt_type else sale.invoice_no
        priced_by_ref.setdefault(key, []).append(line)
        sale_id_by_ref[key] = sale.id

    # Invoice/PO # is clickable when it resolves to a real record — a sale
    # (any sale-driven reason shares the same ref/display-invoice, already
    # collected above), a purchase, or a stock count. Scoped to just the
    # refs these movements actually carry, not a store-wide scan.
    purchase_reasons = {"purchase", "purchase-return", "purchase-cancelled", "purchase-edit-reverse"}
    purchase_refs = {m.ref for m in movements if m.ref and m.reason in purchase_reasons}
    purchase_id_by_ref = dict(
        db.query(models.Purchase.ref_no, models.Purchase.id)
        .filter(models.Purchase.ref_no.in_(purchase_refs))
        .all()
    ) if purchase_refs else {}
    count_refs = {m.ref for m in movements if m.ref and m.reason == "stock_count"}
    count_id_by_ref = dict(
        db.query(models.StockCount.ref_no, models.StockCount.id)
        .filter(models.StockCount.ref_no.in_(count_refs))
        .all()
    ) if count_refs else {}

    def _sold_line(movement, delta):
        """The sale line behind this movement, or None. An exchange can put
        the sold and the returned side of the same product on one invoice, so
        where there's more than one candidate, pick the side whose direction
        matches this movement (out = the sold line, in = the returned one)."""
        lines = priced_by_ref.get(movement.ref or "")
        if not lines:
            return None
        if len(lines) > 1:
            want_sold = delta < 0
            match = next(
                (ln for ln in lines if (Decimal(str(ln.qty or 0)) > 0) == want_sold), None
            )
            if match:
                return match
        return lines[0]

    running = opening
    total_in = total_out = Decimal("0")
    rows = []
    for m in movements:
        delta = Decimal(str(m.qty_base or 0))
        running += delta
        if delta >= 0:
            total_in += delta
        else:
            total_out += -delta
        sold = _sold_line(m, delta)
        factor = Decimal(str(sold.unit_factor or 1)) if sold else Decimal("1")

        # Cost and price have to be read in the SAME unit or the row lies.
        # A movement's own unit_cost is always per base unit (e.g. per Kg),
        # so on a sale rung up per bag it would sit beside a per-bag selling
        # price and imply a huge margin when the real per-bag cost is 20x
        # what's shown. Scale the cost by the sold unit's factor so both
        # columns read per bag. SaleLine.unit_cost is the cost frozen at the
        # moment of sale — the same figure COGS uses — so it wins over the
        # movement's when present.
        movement_cost = Decimal(str(m.unit_cost)) if m.unit_cost is not None else None
        if sold:
            base_cost = Decimal(str(sold.unit_cost or 0))
            if base_cost <= 0 and movement_cost is not None:
                base_cost = movement_cost
            unit_cost_shown = base_cost * factor if base_cost > 0 else None
        else:
            unit_cost_shown = movement_cost

        if m.reason in purchase_reasons:
            ref_link = f"/purchases/{purchase_id_by_ref[m.ref]}" if m.ref in purchase_id_by_ref else None
        elif m.reason == "stock_count":
            ref_link = f"/stock-count/{count_id_by_ref[m.ref]}" if m.ref in count_id_by_ref else None
        elif m.ref in sale_id_by_ref:
            ref_link = f"/pos/receipt/{sale_id_by_ref[m.ref]}?from=stock_card&back={quote(self_url)}"
        else:
            ref_link = None

        rows.append({
            "movement": m,
            "label": MOVEMENT_LABELS.get(m.reason, (m.reason or "").replace("-", " ").title()),
            "in_qty": (delta / view_factor) if delta > 0 else None,
            "out_qty": (-delta / view_factor) if delta < 0 else None,
            "balance": running / view_factor,
            "unit_cost": unit_cost_shown,
            "sold_price": Decimal(str(sold.unit_price or 0)) if sold else None,
            # Labels both money columns. Only worth showing when it isn't the
            # base unit the qty columns are counted in — otherwise
            # "₱260.00 / Unit" is just noise.
            "sold_unit": (
                sold.unit_name if sold and sold.unit_name and factor != 1 else None
            ),
            "ref_link": ref_link,
        })
    rows.reverse()  # newest first for display

    return templates.TemplateResponse(
        "products/stock_card.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "product": product, "rows": rows, "opening": opening / view_factor,
            "current_total": current_total / view_factor,
            "total_in": total_in / view_factor, "total_out": total_out / view_factor,
            "count": len(movements),
            "back": safe_back_url(back, "/products"),
            "view_unit_id": unit, "view_unit_name": view_unit_name,
        },
    )


@router.get("/products/{product_id:int}/history")
def product_history(product_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Cost and selling-price over time, for the Dashboard's per-item tracking
    chart. Built from two real sources, not a dedicated price-log table:

      cost  — PurchaseLine.new_cost at each confirmed delivery (the actual
              moment a product's cost changes), plus any manual cost_price
              edit found in the Activity Log (a cost can also be corrected by
              hand, not only by receiving stock).
      price — the Activity Log's before/after on selling_price, since a price
              only ever changes through an admin edit — there's no separate
              event source for it the way purchases are for cost.

    Both series end with today's live value as the last point, so the line
    always reaches "now" even if the last recorded change was a while ago.
    The Activity Log only goes back to when it was built, so a product's
    history here effectively starts from then for anything not tied to a
    purchase.
    """
    if not user:
        return {"found": False}
    if not is_staff(user):
        return {"found": False}
    product = db.get(models.Product, product_id)
    if not product or not product.is_active:
        return {"found": False}

    today = date.today()
    now = datetime.now()
    # points[key] = {date, ts, value, label, old}. Keyed by a unique event id
    # (not by date) so multiple changes on the same day each keep their own
    # point instead of the later one overwriting the earlier one. `ts` (full
    # timestamp) drives ordering and the chart's x-position so same-day points
    # don't collapse onto each other; `date` is just the display label.
    # `old` (when known — a purchase line's old_cost, or an audit diff's
    # before-value) is what carried this series before this specific point,
    # so a single recorded change can still be seeded with an earlier
    # baseline instead of showing as a lone dot.
    cost_points = {}
    price_points = {}

    # ---- cost: confirmed purchases -----------------------------------------
    purchase_rows = (
        db.query(models.PurchaseLine, models.Purchase)
        .join(models.Purchase, models.PurchaseLine.purchase_id == models.Purchase.id)
        .filter(
            models.PurchaseLine.product_id == product_id,
            models.Purchase.txn_type == "receive",
            models.Purchase.status.in_(("confirmed", "paid")),
            models.Purchase.confirmed_at.isnot(None),
        )
        .order_by(models.Purchase.confirmed_at)
        .all()
    )
    for line, purchase in purchase_rows:
        cost_points[f"purchase-{line.id}"] = {
            "date": purchase.confirmed_at.date().isoformat(), "ts": purchase.confirmed_at.isoformat(),
            "value": float(line.new_cost or 0), "label": f"Received via {purchase.ref_no}",
            "old": float(line.old_cost or 0),
        }

    # ---- cost + price: manual edits, from the Activity Log ----------------
    audit_rows = (
        db.query(models.AuditLog)
        .filter(models.AuditLog.entity_type == "product", models.AuditLog.entity_id == product_id,
                models.AuditLog.action.in_(("create", "update")))
        .order_by(models.AuditLog.created_at)
        .all()
    )
    for row in audit_rows:
        if not row.changes:
            continue
        try:
            changes = json.loads(row.changes)
        except ValueError:
            continue
        d = row.created_at.date().isoformat()
        ts = row.created_at.isoformat()
        if "cost_price" in changes:
            try:
                old, new = changes["cost_price"]
                cost_points[f"audit-{row.id}-cost"] = {"date": d, "ts": ts, "value": float(new), "label": "Manual cost edit",
                                                         "old": float(old) if old not in (None, "") else None}
            except (TypeError, ValueError):
                pass
        if "selling_price" in changes:
            try:
                old, new = changes["selling_price"]
                price_points[f"audit-{row.id}-price"] = {"date": d, "ts": ts, "value": float(new), "label": "Manual price edit",
                                                           "old": float(old) if old not in (None, "") else None}
            except (TypeError, ValueError):
                pass

    def _finalize(points: dict, current_value: float, current_label: str) -> list:
        series = sorted(points.values(), key=lambda p: p["ts"])
        # A single change is still real movement — seed a baseline at the
        # product's creation date using that change's "before" value, so it
        # draws as a slope ("was X since added, changed to Y on this date")
        # instead of one disconnected dot. Only when that's actually earlier
        # than anything already in the series.
        if series and series[0].get("old") is not None and product.created_at:
            if product.created_at < datetime.fromisoformat(series[0]["ts"]):
                series.insert(0, {
                    "date": product.created_at.date().isoformat(), "ts": product.created_at.isoformat(),
                    "value": series[0]["old"], "label": "Since product was added",
                })
        today_iso = today.isoformat()
        if not series or series[-1]["date"] != today_iso:
            series.append({"date": today_iso, "ts": now.isoformat(), "value": current_value, "label": current_label})
        return [{"date": p["date"], "ts": p["ts"], "value": p["value"], "label": p["label"]} for p in series]

    cost_series = _finalize(cost_points, float(product.cost_price or 0), "Current cost")
    price_series = _finalize(price_points, float(product.selling_price or 0), "Current price")

    return {
        "found": True,
        "product": {"id": product.id, "name": product.name, "cost_price": float(product.cost_price or 0),
                     "selling_price": float(product.selling_price or 0)},
        "cost_series": cost_series,
        "price_series": price_series,
    }


# --------------------------------------------------------------------------- #
# Bulk import (Excel / CSV)
# --------------------------------------------------------------------------- #
def _parse_bool(value: str) -> bool:
    return str(value or "").strip().lower() in {"1", "y", "yes", "true", "vat", "x", "✓", "oui"}


def _parse_upload(filename: str, contents: bytes, header_map=None, fields=None,
                   required_field="name", required_label="Product Name",
                   keep_extra_columns=False):
    """Return (rows, error). rows is a list of dicts keyed by `fields`
    (defaults to the products FIELDS list; the Units & Conversions import
    passes HEADER_MAP_UNITS/FIELDS_UNITS instead).

    keep_extra_columns=True additionally puts every column the header_map
    doesn't recognise into record["_extra"] as {header label: value}. The
    stock-count sheet needs this: its per-unit columns are named after each
    product's own units (Sack, Elf Load, ...), so they can't be known ahead
    of time the way a fixed header map requires."""
    header_map = header_map if header_map is not None else HEADER_MAP
    fields = fields if fields is not None else FIELDS
    name = (filename or "").lower()
    try:
        if name.endswith(".csv"):
            text = contents.decode("utf-8-sig", errors="replace")
            table = list(csv.reader(io.StringIO(text)))
        elif name.endswith(".xlsx") or name.endswith(".xlsm"):
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            ws = wb.active
            table = [list(row) for row in ws.iter_rows(values_only=True)]
        else:
            return None, "Unsupported file type. Please upload a .xlsx or .csv file."
    except Exception as exc:  # noqa: BLE001
        return None, f"Could not read the file: {exc}"

    if not table:
        return None, "The file appears to be empty."

    header = table[0]
    idx = {}
    extra_idx = {}
    for i, cell in enumerate(header):
        label = str(cell or "").strip()
        key = label.lower()
        if key in header_map:
            idx[header_map[key]] = i
        elif keep_extra_columns and label:
            extra_idx[label] = i
    if required_field not in idx:
        return None, (
            f"Missing a '{required_label}' column. Download the template to see the "
            "expected format."
        )

    def cell(raw, field):
        i = idx.get(field)
        if i is None or i >= len(raw):
            return ""
        val = raw[i]
        return "" if val is None else str(val).strip()

    rows = []
    for raw in table[1:]:
        if not raw:
            continue
        record = {f: cell(raw, f) for f in fields}
        extra = {}
        if keep_extra_columns:
            for label, i in extra_idx.items():
                val = raw[i] if i < len(raw) else None
                extra[label] = "" if val is None else str(val).strip()
        if not any(record[f] for f in fields) and not any(extra.values()):  # skip blank rows
            continue
        if keep_extra_columns:
            record["_extra"] = extra
        # Shelf is the one column where "not in the file at all" has to stay
        # distinguishable from "in the file but blank": a re-import of an
        # updated price/stock list has no reason to mention shelves, and must
        # not silently wipe every shelf assignment the shop already made.
        # None = column absent (leave the product's shelf alone);
        # ""   = column present but blank for this row (clear it, on purpose).
        if "shelf" in fields and "shelf" not in idx:
            record["shelf"] = None
        rows.append(record)
    return rows, None


def _classify_import_rows(db: Session, numbered_rows):
    """Read-only pass over (line_no, record) pairs: figure out whether each
    row would create a new product, update an existing one (matched by
    name), or hit a barcode already used by a different product — without
    writing anything to the database. This is what the duplicate-review
    screen is built from, and it's re-run at confirm time against the
    then-current data in case anything changed in between."""
    classified = []
    for line_no, record in numbered_rows:
        name = (record["name"] or "").strip()
        if not name:
            classified.append({"line_no": line_no, "name": name, "record": record, "action": "blank"})
            continue
        existing = (
            db.query(models.Product)
            .filter(func.lower(models.Product.name) == name.lower())
            .filter(models.Product.is_active.is_(True))
            .first()
        )
        barcode = (record["barcode"] or "").strip()
        conflict = (
            db.query(models.Product)
            .filter(models.Product.barcode == barcode)
            .filter(models.Product.id != (existing.id if existing else 0))
            .first()
        ) if barcode else None
        if conflict:
            classified.append({
                "line_no": line_no, "name": name, "record": record, "action": "conflict",
                "message": f"Barcode “{barcode}” is already assigned to “{conflict.name}”.",
            })
        elif existing:
            classified.append({"line_no": line_no, "name": name, "record": record, "action": "update", "existing": existing})
        else:
            classified.append({"line_no": line_no, "name": name, "record": record, "action": "create"})
    return classified


def _apply_import_row(db: Session, file_label: str, record: dict, existing):
    """Create or update one product from a parsed import row. Returns
    ('created' | 'updated', None) on success, or (None, error_message) on
    failure. Caller wraps this in its own SAVEPOINT."""
    name = record["name"].strip()
    product = existing or models.Product()
    barcode = (record["barcode"] or "").strip()
    if barcode:
        conflict = (
            db.query(models.Product)
            .filter(models.Product.barcode == barcode)
            .filter(models.Product.id != (product.id or 0))
            .first()
        )
        if conflict:
            return None, f"Barcode “{barcode}” is already assigned to “{conflict.name}”."
    old_total = Decimal(str(product.total_qty or 0)) if existing else Decimal("0")
    product.name = name
    product.barcode = barcode or None
    product.category = _get_or_create_category(db, record["category"])
    product.subcategory = _get_or_create_subcategory(db, record["subcategory"], product.category)
    product.unit_type = _get_or_create_unit_type(db, record["unit_type"])
    # None means the uploaded file had no Shelf column at all, so it isn't
    # saying anything about shelves — keep whatever this product already has.
    if record.get("shelf") is not None:
        product.shelf = _get_or_create_shelf(db, record["shelf"])
    product.cost_price = _to_decimal(record["cost"])
    product.selling_price = _to_decimal(record["selling"])
    product.beginning_stock = _to_decimal(record["beginning"])
    product.stock_qty = _to_decimal(record["stocks"])
    product.is_vat = _parse_bool(record["vat"])
    if existing:
        # A re-import can silently move an existing product's stock with no
        # trace otherwise — record it the same way a manual edit would, so
        # it still shows up on the Stock Card and in Inventory Adjustments/P&L.
        delta = Decimal(str(product.total_qty or 0)) - old_total
        if delta != 0:
            unit_cost = Decimal(str(product.cost_price or 0))
            db.add(models.StockMovement(
                product_id=product.id, qty_base=delta, reason="adjustment",
                ref="bulk import", unit_cost=unit_cost, value=delta * unit_cost,
                note=f"Bulk import: {file_label}",
            ))
        return "updated", None
    db.add(product)
    return "created", None


def _run_import(db: Session, user, request, filename: str, classified, skip_line_nos=frozenset()):
    """Apply a classified row set. 'blank' rows are counted as skipped,
    'conflict' rows always become errors, and any row whose line_no is in
    skip_line_nos (an unchecked duplicate on the review screen) is skipped
    without being touched."""
    created = updated = skipped = 0
    errors = []
    for item in classified:
        line_no, name, action = item["line_no"], item["name"], item["action"]
        if action == "blank":
            skipped += 1
            continue
        if action == "conflict":
            errors.append({"row": line_no, "name": name, "message": item["message"]})
            continue
        if line_no in skip_line_nos:
            skipped += 1
            continue
        savepoint = db.begin_nested()
        try:
            status, err = _apply_import_row(db, filename, item["record"], item.get("existing"))
            if err:
                errors.append({"row": line_no, "name": name, "message": err})
                savepoint.rollback()
                continue
            if status == "created":
                created += 1
            else:
                updated += 1
            savepoint.commit()
        except Exception as exc:  # noqa: BLE001
            savepoint.rollback()
            errors.append({"row": line_no, "name": name, "message": str(exc)})
    if created or updated:
        audit.record(
            db, user=user, request=request, action="update", entity_type="product",
            entity_label=filename,
            summary=f"Bulk import from “{filename}”: {created} created, {updated} updated, {skipped} skipped",
        )
    db.commit()
    return {
        "created": created, "updated": updated, "skipped": skipped,
        "errors": errors, "total": len(classified), "filename": filename,
    }


@router.get("/products/import", response_class=HTMLResponse)
def import_form(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    return templates.TemplateResponse(
        "products/import.html",
        {"request": request, "app_name": request.app.title, "user": user, "result": None},
    )


@router.post("/products/import", response_class=HTMLResponse)
def import_upload(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)

    contents = file.file.read()
    rows, error = _parse_upload(file.filename, contents)

    if error:
        return templates.TemplateResponse(
            "products/import.html",
            {"request": request, "app_name": request.app.title, "user": user, "result": {"error": error}},
        )

    classified = _classify_import_rows(db, list(enumerate(rows, start=2)))  # row 2 = first data row
    has_duplicates = any(r["action"] == "update" for r in classified)

    if not has_duplicates:
        # Nothing ambiguous — same one-step import as before.
        result = _run_import(db, user, request, file.filename, classified)
        return templates.TemplateResponse(
            "products/import.html",
            {"request": request, "app_name": request.app.title, "user": user, "result": result},
        )

    # At least one row matches an existing product by name — stop and let the
    # admin choose per row instead of silently overwriting. Nothing has been
    # written to the database yet; the parsed rows ride along as hidden JSON
    # so Confirm doesn't need the file re-uploaded.
    payload = [{"line_no": r["line_no"], "record": r["record"]} for r in classified if r["action"] in ("create", "update")]
    return templates.TemplateResponse(
        "products/import_preview.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "filename": file.filename,
            "rows_json": json.dumps(payload),
            "new_rows": [r for r in classified if r["action"] == "create"],
            "update_rows": [r for r in classified if r["action"] == "update"],
            "conflict_rows": [r for r in classified if r["action"] == "conflict"],
            "blank_count": sum(1 for r in classified if r["action"] == "blank"),
        },
    )


@router.post("/products/import/confirm", response_class=HTMLResponse)
async def import_confirm(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)

    form = await request.form()

    # Which rows to apply is keyed by dynamic apply_<line_no> field names, so
    # this can't be fixed Form(...) params — the actual row-by-row import
    # (the blocking part, and the slowest route in the app for a big file)
    # runs off the event loop via run_in_threadpool instead.
    def _do():
        filename = form.get("filename") or "import"
        try:
            payload = json.loads(form.get("rows_json") or "[]")
        except ValueError:
            payload = []

        numbered_rows = [(int(item["line_no"]), item["record"]) for item in payload]
        # Re-classify against the *current* data rather than trusting the preview
        # snapshot — something could have changed (or another admin could have
        # imported) in the time between rendering the review screen and this submit.
        classified = _classify_import_rows(db, numbered_rows)

        skip_line_nos = {
            item["line_no"] for item in classified
            if item["action"] == "update" and form.get(f"apply_{item['line_no']}") is None
        }
        result = _run_import(db, user, request, filename, classified, skip_line_nos=skip_line_nos)
        return templates.TemplateResponse(
            "products/import.html",
            {"request": request, "app_name": request.app.title, "user": user, "result": result},
        )

    return await run_in_threadpool(_do)


@router.get("/products/import/template")
def download_template(user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(TEMPLATE_HEADERS)

    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # Example rows (delete these before importing your real data).
    # The second row deliberately leaves "Cost of Sales" and "Barcode" blank:
    # cost is optional and gets filled in automatically when you receive stock
    # in Purchasing, and barcode is optional too — scan it in with a barcode
    # scanner (it types like a keyboard) or leave it blank if the item has none.
    ws.append(["4800000000017", "Portland Cement 40kg", "Cement", "Bag Cement", "Bag", "Shelf 1", 220, 260, 10, 5])
    ws.append([None, "Common Wire Nail #4", "Fasteners", "Nails", "Kg", "Shelf 2", None, 95, 25.5, 0])

    widths = [18, 26, 16, 16, 12, 14, 14, 14, 24, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="product_import_template.xlsx"'},
    )


# --------------------------------------------------------------------------- #
# Bulk import (Excel / CSV) — Units & Conversions
# --------------------------------------------------------------------------- #
def _classify_unit_import_rows(db: Session, numbered_rows):
    """Group parsed rows by Product Name (case-insensitive) and look each one
    up now, so the confirm screen can show which products weren't found
    without touching the database."""
    groups = {}
    order = []
    for line_no, record in numbered_rows:
        pname = (record["product"] or "").strip()
        uname = (record["unit_name"] or "").strip()
        if not pname or not uname:
            continue
        key = pname.lower()
        if key not in groups:
            product = (
                db.query(models.Product)
                .filter(func.lower(models.Product.name) == key)
                .filter(models.Product.is_active.is_(True))
                .first()
            )
            groups[key] = {"name": pname, "product": product, "rows": [], "line_nos": []}
            order.append(key)
        fac = _to_decimal(record["factor"], "1")
        if fac <= 0:
            fac = Decimal("1")
        groups[key]["rows"].append({
            "name": uname,
            "relative_factor": fac,
            "price": _to_decimal(record["price"]),
            "relative_to": (record["relative_to"] or "").strip() or None,
            "markup_pct": _to_decimal(record["markup_pct"]),
            "margin_pct": _to_decimal(record["margin_pct"]),
        })
        groups[key]["line_nos"].append(line_no)
    return [groups[k] for k in order]


def _run_unit_import(db: Session, user, request, filename: str, groups):
    """Apply each product's full ladder in its own SAVEPOINT — a bad row for
    one product shouldn't block the rest of the file."""
    updated = 0
    errors = []
    for g in groups:
        if not g["product"]:
            errors.append({
                "row": g["line_nos"][0], "name": g["name"],
                "message": f"No active product named “{g['name']}” — add it first via the Products import or the Add Product form.",
            })
            continue
        savepoint = db.begin_nested()
        try:
            _apply_unit_rows(g["product"], db, g["rows"])
            savepoint.commit()
            updated += 1
        except Exception as exc:  # noqa: BLE001
            savepoint.rollback()
            errors.append({"row": g["line_nos"][0], "name": g["name"], "message": str(exc)})
    if updated:
        audit.record(
            db, user=user, request=request, action="update", entity_type="product",
            entity_label=filename,
            summary=f"Bulk unit-conversion import from “{filename}”: {updated} product(s) updated, {len(errors)} skipped",
        )
    db.commit()
    return {
        "created": 0, "updated": updated, "skipped": 0,
        "errors": errors, "total": len(groups), "filename": filename,
    }


@router.get("/products/import/units", response_class=HTMLResponse)
def import_units_form(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)
    return templates.TemplateResponse(
        "products/import_units.html",
        {"request": request, "app_name": request.app.title, "user": user, "result": None},
    )


@router.post("/products/import/units", response_class=HTMLResponse)
def import_units_upload(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)

    contents = file.file.read()
    rows, error = _parse_upload(
        file.filename, contents,
        header_map=HEADER_MAP_UNITS, fields=FIELDS_UNITS,
        required_field="product", required_label="Product Name",
    )
    if error:
        return templates.TemplateResponse(
            "products/import_units.html",
            {"request": request, "app_name": request.app.title, "user": user, "result": {"error": error}},
        )

    groups = _classify_unit_import_rows(db, list(enumerate(rows, start=2)))
    result = _run_unit_import(db, user, request, file.filename, groups)
    return templates.TemplateResponse(
        "products/import_units.html",
        {"request": request, "app_name": request.app.title, "user": user, "result": result},
    )


@router.get("/products/import/units/template")
def download_units_template(user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/products", status_code=302)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Units"
    ws.append(TEMPLATE_HEADERS_UNITS)

    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # Example: a product whose base unit is "Kg" also sells by the Sack (25kg)
    # and by the Elf Load, which is defined relative to Sack rather than base.
    ws.append(["GRAVEL", "Sack", 25, "", 500, None, None])
    ws.append(["GRAVEL", "Elf Load", 6, "Sack", 30000, None, None])

    widths = [22, 16, 14, 22, 12, 10, 10]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="unit_conversion_import_template.xlsx"'},
    )
