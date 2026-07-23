"""Sales history (fully-paid) and receivables (credit) with settlement."""
import io
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from . import models
from .database import get_db
from .deps import get_current_user, is_admin
from .templating import templates

router = APIRouter()

SETTLE_METHODS = [("cash", "Cash"), ("gcash", "GCash"), ("bank_transfer", "Bank Transfer"), ("cheque", "Cheque")]
PAGE_SIZE = 20
TYPE_LABELS = {"sale": "Sale", "refund": "Refund", "exchange": "Exchange"}


def _parse_date(s: str):
    try:
        return date.fromisoformat(s) if s else None
    except ValueError:
        return None


def _local_date(col):
    return func.date(func.timezone("Asia/Manila", col))


def _dec(value, default="0") -> Decimal:
    try:
        return Decimal(str(value).strip().replace(",", "") or default)
    except (InvalidOperation, AttributeError, ValueError):
        return Decimal(default)


def _settled_map(db: Session):
    rows = (
        db.query(models.ReceivableSettlement.sale_id, func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
        .group_by(models.ReceivableSettlement.sale_id)
        .all()
    )
    return {sid: Decimal(amt) for sid, amt in rows}


def _outstanding(sale, settled_map) -> Decimal:
    return (sale.receivable_amount or Decimal("0")) - settled_map.get(sale.id, Decimal("0"))


def _back_url(user, sale=None) -> str:
    """Where "back"/"cancel" should go — admins return to the receivables
    list; cashiers (who can't see that list) go to the customer's credit
    statement instead, since that's how they got here."""
    if is_admin(user):
        return "/sales/receivables"
    if sale and sale.customer_id:
        return f"/credits/{sale.customer_id}"
    return "/credits"


def _filtered_sales_query(db: Session, q: str, type_filter: str, date_from, date_to):
    """Fully-paid sales, filtered by search text / type / date. Filtering (not
    just fetch-then-check) happens at the SQL level so pagination counts and
    the Excel export both see the same accurate result set."""
    settled_sub = (
        db.query(
            models.ReceivableSettlement.sale_id.label("sid"),
            func.coalesce(func.sum(models.ReceivableSettlement.amount), 0).label("paid"),
        )
        .group_by(models.ReceivableSettlement.sale_id)
        .subquery()
    )
    query = (
        db.query(models.Sale)
        .outerjoin(settled_sub, settled_sub.c.sid == models.Sale.id)
        .filter(models.Sale.receivable_amount <= func.coalesce(settled_sub.c.paid, 0))
    )
    if q:
        like = f"%{q}%"
        query = query.filter(or_(models.Sale.invoice_no.ilike(like), models.Sale.customer_name.ilike(like)))
    if type_filter in TYPE_LABELS:
        query = query.filter(models.Sale.txn_type == type_filter)
    if date_from:
        query = query.filter(_local_date(models.Sale.created_at) >= date_from)
    if date_to:
        query = query.filter(_local_date(models.Sale.created_at) <= date_to)
    return query


@router.get("/sales", response_class=HTMLResponse)
def sales_history(
    request: Request,
    q: str = "",
    type_filter: str = "",
    date_from: str = "",
    date_to: str = "",
    page: int = 1,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_admin(user):
        return RedirectResponse("/pos", status_code=302)
    q = (q or "").strip()
    page = max(page, 1)
    df, dt = _parse_date(date_from), _parse_date(date_to)

    query = _filtered_sales_query(db, q, type_filter, df, dt)

    total_count, total_sales = query.with_entities(
        func.count(models.Sale.id), func.coalesce(func.sum(models.Sale.total), 0)
    ).one()
    pages = max((total_count + PAGE_SIZE - 1) // PAGE_SIZE, 1)
    page = min(page, pages)

    sales_page = (
        query.order_by(models.Sale.id.desc())
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )
    totals = {"count": total_count, "sales": Decimal(str(total_sales or 0))}
    return templates.TemplateResponse(
        "sales/list.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "sales": sales_page, "totals": totals, "tab": "all", "q": q,
         "type_filter": type_filter, "date_from": date_from, "date_to": date_to,
         "types": TYPE_LABELS, "page": page, "pages": pages},
    )


@router.get("/sales/export")
def export_sales(
    q: str = "",
    type_filter: str = "",
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_admin(user):
        return RedirectResponse("/pos", status_code=302)
    q = (q or "").strip()
    df, dt = _parse_date(date_from), _parse_date(date_to)

    query = _filtered_sales_query(db, q, type_filter, df, dt)
    sales = query.order_by(models.Sale.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    headers = ["Invoice #", "Type", "Date", "Customer", "Payment", "Total"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for s in sales:
        ws.append([
            s.invoice_no,
            TYPE_LABELS.get(s.txn_type, s.txn_type),
            s.created_at.strftime("%Y-%m-%d %I:%M %p") if s.created_at else "",
            s.customer_name or "Walk-in",
            s.payment_method or "",
            float(s.total or 0),
        ])

    widths = [16, 12, 20, 24, 20, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname_bits = ["sales"]
    if type_filter in TYPE_LABELS:
        fname_bits.append(type_filter)
    if date_from:
        fname_bits.append(date_from)
    if date_to and date_to != date_from:
        fname_bits.append(date_to)
    filename = "_".join(fname_bits) + ".xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/sales/receivables", response_class=HTMLResponse)
def receivables(
    request: Request,
    q: str = "",
    page: int = 1,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_admin(user):
        return RedirectResponse("/pos", status_code=302)
    q = (q or "").strip()
    page = max(page, 1)

    # Outstanding computed at the SQL level so LIMIT/OFFSET pagination and the
    # total-credit sum both reflect what's actually still owed, not the
    # original receivable_amount.
    settled_sub = (
        db.query(
            models.ReceivableSettlement.sale_id.label("sid"),
            func.coalesce(func.sum(models.ReceivableSettlement.amount), 0).label("paid"),
        )
        .group_by(models.ReceivableSettlement.sale_id)
        .subquery()
    )
    outstanding_expr = models.Sale.receivable_amount - func.coalesce(settled_sub.c.paid, 0)
    query = (
        db.query(models.Sale, outstanding_expr.label("outstanding"))
        .outerjoin(settled_sub, settled_sub.c.sid == models.Sale.id)
        .filter(models.Sale.receivable_amount > 0)
        .filter(outstanding_expr > 0)
    )
    if q:
        like = f"%{q}%"
        query = query.filter(or_(models.Sale.invoice_no.ilike(like), models.Sale.customer_name.ilike(like)))

    total_count, total_credit = query.with_entities(
        func.count(models.Sale.id), func.coalesce(func.sum(outstanding_expr), 0)
    ).one()
    pages = max((total_count + PAGE_SIZE - 1) // PAGE_SIZE, 1)
    page = min(page, pages)

    page_rows = (
        query.order_by(models.Sale.id.desc())
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )
    rows = [(s, Decimal(str(out))) for s, out in page_rows]
    # Balances waiting on a COD delivery are still owed, so they stay in this
    # list and in the total — but they're tagged, because they're awaiting a
    # handover, not a customer who is late paying their credit.
    from .deliveries import cod_pending_sale_ids
    cod_ids = cod_pending_sale_ids(db)
    return templates.TemplateResponse(
        "sales/receivables.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "rows": rows, "total_credit": Decimal(str(total_credit or 0)), "tab": "receivables", "q": q,
         "cod_ids": cod_ids,
         "page": page, "pages": pages, "total": total_count},
    )


@router.get("/sales/receivables/{sale_id:int}/pay", response_class=HTMLResponse)
def settle_form(sale_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    sale = db.get(models.Sale, sale_id)
    if not sale:
        return RedirectResponse(_back_url(user), status_code=302)
    settled = _settled_map(db)
    outstanding = _outstanding(sale, settled)
    return templates.TemplateResponse(
        "sales/settle.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "sale": sale, "outstanding": outstanding, "methods": SETTLE_METHODS, "error": None,
            "back_url": _back_url(user, sale),
        },
    )


@router.post("/sales/receivables/{sale_id:int}/pay")
async def settle_pay(sale_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    sale = db.get(models.Sale, sale_id)
    if not sale:
        return RedirectResponse(_back_url(user), status_code=302)

    settled = _settled_map(db)
    outstanding = _outstanding(sale, settled)
    form = await request.form()
    method = (form.get("method") or "cash").strip().lower()
    amount = _dec(form.get("amount"))

    def rerender(error):
        return templates.TemplateResponse(
            "sales/settle.html",
            {"request": request, "app_name": request.app.title, "user": user,
             "sale": sale, "outstanding": outstanding, "methods": SETTLE_METHODS, "error": error,
             "back_url": _back_url(user, sale)},
        )

    if amount <= 0:
        return rerender("Enter an amount greater than zero.")
    if amount > outstanding:
        amount = outstanding  # never collect more than owed

    if method == "cheque":
        # A post-dated cheque doesn't settle anything yet — it just goes into
        # the PDC register until the bank actually honors it.
        raw_date = (form.get("cheque_date") or "").strip()
        try:
            cheque_date = date.fromisoformat(raw_date)
        except ValueError:
            return rerender("Enter a valid cheque date (the date printed on the cheque).")
        pdc = models.PostDatedCheque(
            direction="received", amount=amount,
            bank=(form.get("bank") or "").strip() or None,
            cheque_no=(form.get("cheque_no") or "").strip() or None,
            cheque_date=cheque_date,
            sale_id=sale.id, customer_id=sale.customer_id,
            created_by=user.id,
        )
        db.add(pdc)
        db.flush()
        db.commit()
        return RedirectResponse(f"/pdc/{pdc.id}", status_code=status.HTTP_302_FOUND)

    db.add(models.ReceivableSettlement(
        sale_id=sale.id, method=method, amount=amount,
        cashier_id=user.id,
    ))
    db.commit()
    return RedirectResponse(_back_url(user, sale), status_code=status.HTTP_302_FOUND)
