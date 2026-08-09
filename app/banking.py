"""Cash & Banking: bank accounts and their deposit/withdrawal ledger.

A balance is never stored — it's opening_balance plus the sum of that
account's transactions, computed on the fly (same idea as how a sale's
outstanding credit is derived, not cached). Moving money between two
accounts is just a withdrawal on one and a deposit on the other; there's
no separate "transfer" record type.
"""
import csv
import io
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import case, func
from sqlalchemy.orm import Session

from . import accounting, audit, models
from .database import get_db
from .deps import get_current_user, is_staff
from .templating import templates

router = APIRouter()

PAGE_SIZE = 15
TXN_LABELS = {"deposit": "Deposit", "withdrawal": "Withdrawal"}
ACCOUNT_KINDS = [("bank", "Bank Account"), ("ewallet", "E-Wallet"), ("petty_cash", "Petty Cash")]
ZERO = Decimal("0")


def _dec(value, default="0") -> Decimal:
    try:
        return Decimal(str(value).strip().replace(",", "") or default)
    except (InvalidOperation, AttributeError, ValueError):
        return Decimal(default)


def _parse_date(s: str):
    try:
        return date.fromisoformat(s) if s else None
    except ValueError:
        return None


def _balances_for(db: Session, account_ids=None):
    """{account_id: (deposits_total, withdrawals_total)} for non-voided txns."""
    q = db.query(
        models.BankTransaction.account_id,
        func.coalesce(func.sum(case((models.BankTransaction.txn_type == "deposit", models.BankTransaction.amount), else_=0)), 0),
        func.coalesce(func.sum(case((models.BankTransaction.txn_type == "withdrawal", models.BankTransaction.amount), else_=0)), 0),
    ).filter(models.BankTransaction.is_voided.is_(False))
    if account_ids is not None:
        q = q.filter(models.BankTransaction.account_id.in_(account_ids))
    rows = q.group_by(models.BankTransaction.account_id).all()
    return {r[0]: (Decimal(str(r[1] or 0)), Decimal(str(r[2] or 0))) for r in rows}


def _account_balance(account, balances: dict) -> Decimal:
    dep, wd = balances.get(account.id, (ZERO, ZERO))
    return Decimal(str(account.opening_balance or 0)) + dep - wd


@router.get("/banking", response_class=HTMLResponse)
def list_accounts(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    accounts = db.query(models.BankAccount).filter(models.BankAccount.is_active.is_(True)).order_by(models.BankAccount.name).all()
    balances = _balances_for(db)
    rows = [{"account": a, "balance": _account_balance(a, balances)} for a in accounts]
    total_balance = sum((r["balance"] for r in rows), ZERO)

    return templates.TemplateResponse(
        "banking/accounts.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "rows": rows, "total_balance": total_balance,
        },
    )


def _render_account_form(request, db, user, account=None, error=None):
    ledger_accounts = db.query(models.Account).filter(
        models.Account.is_active.is_(True), models.Account.account_type == "asset",
    ).order_by(models.Account.code).all()
    return templates.TemplateResponse(
        "banking/account_form.html",
        {"request": request, "app_name": request.app.title, "user": user, "account": account, "error": error,
         "ledger_accounts": ledger_accounts, "account_kinds": ACCOUNT_KINDS},
    )


@router.get("/banking/accounts/new", response_class=HTMLResponse)
def new_account(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    return _render_account_form(request, db, user)


@router.post("/banking/accounts")
async def create_account(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    form = await request.form()
    name = (form.get("name") or "").strip()
    if not name:
        return _render_account_form(request, db, user, error="Account name is required.")
    if db.query(models.BankAccount).filter(func.lower(models.BankAccount.name) == name.lower()).first():
        return _render_account_form(request, db, user, error=f"An account named '{name}' already exists.")
    gl_account_id = (form.get("gl_account_id") or "").strip()
    account_kind = (form.get("account_kind") or "bank").strip()
    if account_kind not in dict(ACCOUNT_KINDS):
        account_kind = "bank"
    account = models.BankAccount(
        name=name,
        bank_name=(form.get("bank_name") or "").strip() or None,
        account_no=(form.get("account_no") or "").strip() or None,
        opening_balance=_dec(form.get("opening_balance")),
        gl_account_id=int(gl_account_id) if gl_account_id else None,
        account_kind=account_kind,
    )
    db.add(account)
    db.flush()
    audit.record(
        db, user=user, request=request, action="create", entity_type="bank_account",
        entity_id=account.id, entity_label=account.name,
        summary=f"Added bank account “{account.name}” (opening {account.opening_balance})",
    )
    db.commit()
    return RedirectResponse("/banking", status_code=status.HTTP_302_FOUND)


@router.get("/banking/accounts/{account_id:int}/edit", response_class=HTMLResponse)
def edit_account(account_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)
    return _render_account_form(request, db, user, account=account)


@router.post("/banking/accounts/{account_id:int}")
async def update_account(account_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)
    form = await request.form()
    name = (form.get("name") or "").strip()
    if not name:
        return _render_account_form(request, db, user, account=account, error="Account name is required.")
    clash = db.query(models.BankAccount).filter(func.lower(models.BankAccount.name) == name.lower(), models.BankAccount.id != account.id).first()
    if clash:
        return _render_account_form(request, db, user, account=account, error=f"An account named '{name}' already exists.")
    before = audit.snapshot(account, ["name", "bank_name", "account_no", "opening_balance", "is_active", "account_kind"])
    account.name = name
    account.bank_name = (form.get("bank_name") or "").strip() or None
    account.account_no = (form.get("account_no") or "").strip() or None
    account.opening_balance = _dec(form.get("opening_balance"))
    account.is_active = (form.get("status") or "active") == "active"
    gl_account_id = (form.get("gl_account_id") or "").strip()
    account.gl_account_id = int(gl_account_id) if gl_account_id else None
    account_kind = (form.get("account_kind") or "bank").strip()
    account.account_kind = account_kind if account_kind in dict(ACCOUNT_KINDS) else "bank"
    db.flush()
    after = audit.snapshot(account, ["name", "bank_name", "account_no", "opening_balance", "is_active", "account_kind"])
    changes = audit.diff(before, after)
    if changes:
        audit.record(
            db, user=user, request=request, action="update", entity_type="bank_account",
            entity_id=account.id, entity_label=account.name,
            summary=f"Edited bank account “{account.name}”", changes=changes,
        )
    db.commit()
    return RedirectResponse("/banking", status_code=status.HTTP_302_FOUND)


@router.get("/banking/accounts/{account_id:int}", response_class=HTMLResponse)
def view_account(
    account_id: int,
    request: Request,
    txn_type: str = "",
    date_from: str = "",
    date_to: str = "",
    page: int = 1,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)

    page = max(page, 1)
    df, dt = _parse_date(date_from), _parse_date(date_to)

    query = db.query(models.BankTransaction).filter(
        models.BankTransaction.account_id == account_id,
        models.BankTransaction.is_voided.is_(False),
    )
    if txn_type in TXN_LABELS:
        query = query.filter(models.BankTransaction.txn_type == txn_type)
    if df:
        query = query.filter(models.BankTransaction.txn_date >= df)
    if dt:
        query = query.filter(models.BankTransaction.txn_date <= dt)

    total = query.count()
    pages = max((total + PAGE_SIZE - 1) // PAGE_SIZE, 1)
    page = min(page, pages)
    txns = (
        query.order_by(models.BankTransaction.txn_date.desc(), models.BankTransaction.id.desc())
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )

    balances = _balances_for(db, [account_id])
    balance = _account_balance(account, balances)
    dep_total, wd_total = balances.get(account_id, (ZERO, ZERO))

    return templates.TemplateResponse(
        "banking/account_view.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "account": account, "balance": balance, "dep_total": dep_total, "wd_total": wd_total,
            "txns": txns, "txn_type": txn_type, "date_from": date_from, "date_to": date_to,
            "labels": TXN_LABELS, "page": page, "pages": pages, "total": total,
        },
    )


@router.get("/banking/accounts/{account_id:int}/transactions/new", response_class=HTMLResponse)
def new_transaction(
    account_id: int,
    request: Request,
    txn_type: str = "deposit",
    error: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)
    if txn_type not in TXN_LABELS:
        txn_type = "deposit"
    contra_accounts = db.query(models.Account).filter(models.Account.is_active.is_(True)).order_by(models.Account.code).all()
    return templates.TemplateResponse(
        "banking/transaction_form.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "account": account, "txn_type": txn_type, "today": date.today().isoformat(), "error": error,
            "contra_accounts": contra_accounts,
        },
    )


@router.post("/banking/accounts/{account_id:int}/transactions")
async def create_transaction(account_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)
    form = await request.form()
    txn_type = (form.get("txn_type") or "").strip().lower()
    if txn_type not in TXN_LABELS:
        txn_type = "deposit"
    amount = _dec(form.get("amount"))
    if amount <= 0:
        return RedirectResponse(
            f"/banking/accounts/{account_id}/transactions/new?txn_type={txn_type}&error=Enter+an+amount+greater+than+zero.",
            status_code=302,
        )
    contra_account_id = (form.get("contra_account_id") or "").strip()
    txn = models.BankTransaction(
        account_id=account.id,
        txn_type=txn_type,
        amount=amount,
        txn_date=_parse_date((form.get("txn_date") or "").strip()) or date.today(),
        description=(form.get("description") or "").strip() or None,
        reference_no=(form.get("reference_no") or "").strip() or None,
        contra_account_id=int(contra_account_id) if contra_account_id else None,
        created_by=user.id,
    )
    db.add(txn)
    db.flush()
    try:
        accounting.post_bank_transaction(db, txn, entered_by_id=user.id)
    except accounting.PostingError:
        pass
    audit.record(
        db, user=user, request=request, action="create", entity_type="bank_transaction",
        entity_id=txn.id, entity_label=account.name,
        summary=f"{TXN_LABELS[txn_type]} of {amount} on “{account.name}”",
    )
    db.commit()
    return RedirectResponse(f"/banking/accounts/{account_id}", status_code=status.HTTP_302_FOUND)


@router.post("/banking/transactions/{txn_id:int}/void")
def void_transaction(txn_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    txn = db.get(models.BankTransaction, txn_id)
    if not txn:
        return RedirectResponse("/banking", status_code=302)
    account_id = txn.account_id
    txn.is_voided = True
    accounting.reverse_bank_transaction_posting(db, txn, reason=f"Voided {TXN_LABELS.get(txn.txn_type, txn.txn_type)}", entered_by_id=user.id)
    audit.record(
        db, user=user, request=request, action="void", entity_type="bank_transaction",
        entity_id=txn.id, entity_label=(txn.account.name if txn.account else None),
        summary=f"Voided {TXN_LABELS.get(txn.txn_type, txn.txn_type)} of {txn.amount}",
    )
    db.commit()
    return RedirectResponse(f"/banking/accounts/{account_id}", status_code=status.HTTP_302_FOUND)


@router.get("/banking/accounts/{account_id:int}/reconcile", response_class=HTMLResponse)
def reconcile_account(
    account_id: int, request: Request, statement_balance: str = "",
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)

    balances = _balances_for(db, [account_id])
    book_balance = _account_balance(account, balances)

    unreconciled = (
        db.query(models.BankTransaction)
        .filter(models.BankTransaction.account_id == account_id, models.BankTransaction.is_voided.is_(False),
                models.BankTransaction.reconciled_at.is_(None))
        .order_by(models.BankTransaction.txn_date, models.BankTransaction.id)
        .all()
    )
    # Split for display — "outstanding checks" (withdrawals) vs "outstanding
    # deposits", the two categories a real bank reconciliation names them by.
    outstanding_checks = [t for t in unreconciled if t.txn_type == "withdrawal"]
    outstanding_deposits = [t for t in unreconciled if t.txn_type == "deposit"]
    reconciled = (
        db.query(models.BankTransaction)
        .filter(models.BankTransaction.account_id == account_id, models.BankTransaction.is_voided.is_(False),
                models.BankTransaction.reconciled_at.isnot(None))
        .order_by(models.BankTransaction.txn_date.desc(), models.BankTransaction.id.desc())
        .limit(20)
        .all()
    )

    stmt_balance = _dec(statement_balance) if statement_balance else None
    diff = (book_balance - stmt_balance) if stmt_balance is not None else None

    return templates.TemplateResponse(
        "banking/reconcile.html",
        {"request": request, "app_name": request.app.title, "user": user,
         "account": account, "book_balance": book_balance,
         "unreconciled": unreconciled, "outstanding_checks": outstanding_checks,
         "outstanding_deposits": outstanding_deposits, "reconciled": reconciled,
         "statement_balance": statement_balance, "stmt_balance": stmt_balance, "diff": diff,
         "labels": TXN_LABELS},
    )


@router.post("/banking/accounts/{account_id:int}/reconcile")
async def reconcile_submit(account_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)

    form = await request.form()
    txn_ids = [int(v) for v in form.getlist("txn_id")]
    statement_balance = form.get("statement_balance", "")
    if txn_ids:
        now = datetime.now(timezone.utc)
        (
            db.query(models.BankTransaction)
            .filter(models.BankTransaction.account_id == account_id, models.BankTransaction.id.in_(txn_ids))
            .update({"reconciled_at": now, "reconciled_by_id": user.id}, synchronize_session=False)
        )
        audit.record(
            db, user=user, request=request, action="reconcile", entity_type="bank_account",
            entity_id=account.id, entity_label=account.name,
            summary=f"Reconciled {len(txn_ids)} transaction(s) on “{account.name}”",
        )
        db.commit()
    return RedirectResponse(
        f"/banking/accounts/{account_id}/reconcile?statement_balance={statement_balance}", status_code=status.HTTP_302_FOUND
    )


@router.get("/banking/accounts/{account_id:int}/reconcile/export")
def export_reconciliation(account_id: int, statement_balance: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)

    balances = _balances_for(db, [account_id])
    book_balance = _account_balance(account, balances)
    stmt_balance = _dec(statement_balance) if statement_balance else None
    unreconciled = (
        db.query(models.BankTransaction)
        .filter(models.BankTransaction.account_id == account_id, models.BankTransaction.is_voided.is_(False),
                models.BankTransaction.reconciled_at.is_(None))
        .order_by(models.BankTransaction.txn_date, models.BankTransaction.id)
        .all()
    )
    outstanding_checks = [t for t in unreconciled if t.txn_type == "withdrawal"]
    outstanding_deposits = [t for t in unreconciled if t.txn_type == "deposit"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    header_fill = PatternFill("solid", fgColor="1F6FEB")

    ws.append(["Bank Reconciliation", account.name])
    ws.append(["Book Balance", float(book_balance)])
    if stmt_balance is not None:
        ws.append(["Statement Balance", float(stmt_balance)])
        ws.append(["Difference", float(book_balance - stmt_balance)])
    ws.append([])
    ws.append(["Outstanding Checks (withdrawals)"])
    ws.append(["Date", "Description", "Reference", "Amount"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for t in outstanding_checks:
        ws.append([t.txn_date.isoformat(), t.description or "", t.reference_no or "", float(t.amount)])
    ws.append([])
    ws.append(["Outstanding Deposits"])
    ws.append(["Date", "Description", "Reference", "Amount"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for t in outstanding_deposits:
        ws.append([t.txn_date.isoformat(), t.description or "", t.reference_no or "", float(t.amount)])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"reconciliation_{account.name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Header synonyms tried in order, case-insensitive, so a CSV export from any
# bank's portal has a decent chance of auto-detecting correctly without
# forcing the user through a manual column-mapping step every time.
_DATE_HEADERS = ("date", "transaction date", "posting date", "value date")
_DESC_HEADERS = ("description", "particulars", "details", "narration")
_DEBIT_HEADERS = ("debit", "withdrawal", "dr")
_CREDIT_HEADERS = ("credit", "deposit", "cr")
_AMOUNT_HEADERS = ("amount",)
_REF_HEADERS = ("reference", "reference no", "reference no.", "ref", "ref no", "check no", "cheque no")


def _find_col(fieldnames, candidates):
    lowered = {f.strip().lower(): f for f in fieldnames}
    for c in candidates:
        if c in lowered:
            return lowered[c]
    return None


@router.post("/banking/accounts/{account_id:int}/reconcile/import", response_class=HTMLResponse)
async def import_statement(
    account_id: int, request: Request, statement_file: UploadFile,
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    account = db.get(models.BankAccount, account_id)
    if not account:
        return RedirectResponse("/banking", status_code=302)

    raw = (await statement_file.read()).decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(raw))
    fieldnames = reader.fieldnames or []
    date_col = _find_col(fieldnames, _DATE_HEADERS)
    desc_col = _find_col(fieldnames, _DESC_HEADERS)
    debit_col = _find_col(fieldnames, _DEBIT_HEADERS)
    credit_col = _find_col(fieldnames, _CREDIT_HEADERS)
    amount_col = _find_col(fieldnames, _AMOUNT_HEADERS)
    ref_col = _find_col(fieldnames, _REF_HEADERS)

    if not date_col or not (amount_col or debit_col or credit_col):
        return templates.TemplateResponse(
            "banking/reconcile_import_review.html",
            {"request": request, "app_name": request.app.title, "user": user, "account": account,
             "error": f"Couldn't find a Date column and an Amount/Debit/Credit column in this file. "
                      f"Columns found: {', '.join(fieldnames) or '(none — is this a CSV?)'}",
             "matched": [], "unmatched_rows": [], "detected": {}},
        )

    statement_rows = []
    for row in reader:
        raw_date = (row.get(date_col) or "").strip()
        txn_date = None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
            try:
                txn_date = datetime.strptime(raw_date, fmt).date()
                break
            except ValueError:
                continue
        if not txn_date:
            continue
        if amount_col:
            amount = _dec((row.get(amount_col) or "0").replace(",", ""))
        else:
            debit = _dec((row.get(debit_col) or "0").replace(",", "")) if debit_col else Decimal("0")
            credit = _dec((row.get(credit_col) or "0").replace(",", "")) if credit_col else Decimal("0")
            amount = credit - debit  # positive = deposit, negative = withdrawal
        if amount == 0:
            continue
        statement_rows.append({
            "date": txn_date, "amount": abs(amount),
            "txn_type": "deposit" if amount > 0 else "withdrawal",
            "description": (row.get(desc_col) or "").strip() if desc_col else "",
            "reference": (row.get(ref_col) or "").strip() if ref_col else "",
        })

    unreconciled = (
        db.query(models.BankTransaction)
        .filter(models.BankTransaction.account_id == account_id, models.BankTransaction.is_voided.is_(False),
                models.BankTransaction.reconciled_at.is_(None))
        .all()
    )
    # Match by exact amount + txn_type, date within 3 days (banks often post
    # a day or two off from when the transaction was recorded here).
    used_ids = set()
    matched = []
    unmatched_rows = []
    for srow in statement_rows:
        candidate = next(
            (t for t in unreconciled
             if t.id not in used_ids and t.txn_type == srow["txn_type"] and t.amount == srow["amount"]
             and abs((t.txn_date - srow["date"]).days) <= 3),
            None,
        )
        if candidate:
            used_ids.add(candidate.id)
            matched.append({"statement": srow, "txn": candidate})
        else:
            unmatched_rows.append(srow)

    return templates.TemplateResponse(
        "banking/reconcile_import_review.html",
        {"request": request, "app_name": request.app.title, "user": user, "account": account,
         "error": None, "matched": matched, "unmatched_rows": unmatched_rows,
         "detected": {"date": date_col, "description": desc_col, "debit": debit_col,
                      "credit": credit_col, "amount": amount_col, "reference": ref_col}},
    )


@router.post("/banking/transactions/{txn_id:int}/unreconcile")
def unreconcile_transaction(txn_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    txn = db.get(models.BankTransaction, txn_id)
    if not txn:
        return RedirectResponse("/banking", status_code=302)
    account_id = txn.account_id
    txn.reconciled_at = None
    txn.reconciled_by_id = None
    db.commit()
    return RedirectResponse(f"/banking/accounts/{account_id}/reconcile", status_code=status.HTTP_302_FOUND)
