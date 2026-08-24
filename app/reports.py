"""Reports: cross-module summaries that don't live neatly in one module —
Profit & Loss (ties Sales, COGS and Expenses together) and Inventory
Valuation. The hub also points at the exportable lists other modules
already have (Sales, Expenses, Purchases, Cheques).
"""
import io
import json
from datetime import date, datetime, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import models, pricing, settings_store
from .database import get_db
from .deps import get_current_user, is_staff
from .products import low_stock_expr
from .search_utils import multi_word_ilike
from .templating import templates

router = APIRouter()

MANILA = ZoneInfo("Asia/Manila")
ZERO = Decimal("0")


def _today() -> date:
    return datetime.now(MANILA).date()


def _local_date(col):
    return func.date(func.timezone("Asia/Manila", col))


def _parse_date(s: str):
    try:
        return date.fromisoformat(s) if s else None
    except ValueError:
        return None


def _resolve_period(days: int, date_from: str, date_to: str):
    """Same custom-range-overrides-preset logic as the Dashboard, duplicated
    here rather than imported since each module owns its own small helpers."""
    today = _today()
    df, dt = _parse_date(date_from), _parse_date(date_to)
    custom = bool(df and dt)
    if custom:
        if dt > today:
            dt = today
        if df > dt:
            df, dt = dt, df
        if (dt - df).days > 365:
            df = dt - timedelta(days=365)
        return df, dt, custom
    if days not in (7, 30, 90):
        days = 30
    return today - timedelta(days=days - 1), today, custom


@router.get("/reports", response_class=HTMLResponse)
def reports_hub(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    return templates.TemplateResponse(
        "reports/hub.html",
        {"request": request, "app_name": request.app.title, "user": user},
    )


def _pl_data(db: Session, period_start: date, period_end: date):
    """Same formulas the Dashboard uses, so the numbers agree with what the
    owner already sees there: Revenue = net sales (sale + refund + exchange
    totals); Gross Profit = revenue from 'sale' lines minus their frozen cost."""
    revenue = (
        db.query(func.coalesce(func.sum(models.Sale.total), 0))
        .filter(
            _local_date(models.Sale.created_at).between(period_start, period_end),
            models.Sale.is_voided.is_(False),
        )
        .scalar()
    )
    revenue = Decimal(str(revenue or 0))

    cogs_expr = models.SaleLine.qty * models.SaleLine.unit_factor * models.SaleLine.unit_cost
    gross_profit = (
        db.query(func.coalesce(func.sum(models.SaleLine.line_total - cogs_expr), 0))
        .join(models.Sale, models.SaleLine.sale_id == models.Sale.id)
        .filter(
            models.Sale.txn_type == "sale",
            models.Sale.is_voided.is_(False),
            _local_date(models.Sale.created_at).between(period_start, period_end),
        )
        .scalar()
    )
    gross_profit = Decimal(str(gross_profit or 0))

    expense_rows = (
        db.query(models.ExpenseCategory.name, func.coalesce(func.sum(models.Expense.amount), 0))
        .select_from(models.Expense)
        .outerjoin(models.ExpenseCategory, models.Expense.category_id == models.ExpenseCategory.id)
        .filter(models.Expense.is_voided.is_(False), models.Expense.expense_date.between(period_start, period_end))
        .group_by(models.ExpenseCategory.name)
        .all()
    )
    expenses_by_category = sorted(
        [{"name": name or "Uncategorized", "amount": Decimal(str(amt or 0))} for name, amt in expense_rows],
        key=lambda r: r["amount"], reverse=True,
    )
    total_expenses = sum((r["amount"] for r in expenses_by_category), ZERO)

    inventory_adjustment_total = (
        db.query(func.coalesce(func.sum(models.StockMovement.value), 0))
        .filter(models.StockMovement.reason.in_(("adjustment", "stock_count")),
                _local_date(models.StockMovement.created_at).between(period_start, period_end))
        .scalar()
    )
    inventory_adjustment_total = Decimal(str(inventory_adjustment_total or 0))

    return {
        "revenue": revenue,
        "gross_profit": gross_profit,
        "expenses_by_category": expenses_by_category,
        "total_expenses": total_expenses,
        "inventory_adjustment_total": inventory_adjustment_total,
        "net_profit": gross_profit - total_expenses + inventory_adjustment_total,
    }


@router.get("/reports/profit-loss", response_class=HTMLResponse)
def profit_loss(
    request: Request,
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    data = _pl_data(db, period_start, period_end)

    return templates.TemplateResponse(
        "reports/profit_loss.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "days": days, "date_from": date_from, "date_to": date_to,
            "period_start": period_start, "period_end": period_end, "custom": custom,
            **data,
        },
    )


@router.get("/reports/profit-loss/export")
def export_profit_loss(
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, _ = _resolve_period(days, date_from, date_to)
    data = _pl_data(db, period_start, period_end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"
    header_fill = PatternFill("solid", fgColor="1F6FEB")

    def header_row(cells):
        ws.append(cells)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    ws.append([f"Profit & Loss — {period_start.isoformat()} to {period_end.isoformat()}"])
    ws.append([])
    header_row(["Line", "Amount"])
    ws.append(["Revenue (net of refunds/exchanges)", float(data["revenue"])])
    ws.append(["Gross Profit (from goods sold)", float(data["gross_profit"])])
    ws.append([])
    header_row(["Expenses by category", "Amount"])
    for row in data["expenses_by_category"]:
        ws.append([row["name"], float(row["amount"])])
    ws.append(["Total Expenses", float(data["total_expenses"])])
    ws.append([])
    ws.append(["Inventory Shrinkage / Gain (adjustments & stock counts)", float(data["inventory_adjustment_total"])])
    ws.append([])
    ws.append(["Net Profit (Gross Profit − Expenses + Inventory Adjustments)", float(data["net_profit"])])

    for cell in ws["A"]:
        if cell.value in ("Total Expenses", "Net Profit (Gross Profit − Expenses + Inventory Adjustments)"):
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"profit_loss_{period_start.isoformat()}_{period_end.isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _qty_expr():
    return models.Product.beginning_stock + models.Product.stock_qty


def _valuation_rows(db: Session):
    products = (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True))
        .order_by(models.Product.name)
        .all()
    )
    rows = []
    for p in products:
        qty = Decimal(str(p.total_qty or 0))
        cost_val = qty * Decimal(str(p.cost_price or 0))
        retail_val = qty * Decimal(str(p.selling_price or 0))
        rows.append({
            "product": p,
            "category": p.category.name if p.category else "Uncategorized",
            "qty": qty,
            "cost_value": cost_val,
            "retail_value": retail_val,
        })
    return rows


INV_COST_PAGE_SIZE = 50


def _cost_change_lookup(db: Session):
    """Latest confirmed-delivery cost change and latest manual cost_price
    edit, per product_id, across the whole catalog (not just one page) —
    same two sources the Dashboard's per-item history chart uses."""
    latest_purchase = {}
    purchase_rows = (
        db.query(models.PurchaseLine, models.Purchase)
        .join(models.Purchase, models.PurchaseLine.purchase_id == models.Purchase.id)
        .filter(
            models.Purchase.txn_type == "receive",
            models.Purchase.status.in_(("confirmed", "paid")),
            models.Purchase.confirmed_at.isnot(None),
        )
        .order_by(models.Purchase.confirmed_at)
        .all()
    )
    for line, purchase in purchase_rows:
        latest_purchase[line.product_id] = {
            "ts": purchase.confirmed_at,
            "old": Decimal(str(line.old_cost or 0)),
            "new": Decimal(str(line.new_cost or 0)),
            "source": f"Delivery {purchase.ref_no}",
        }

    latest_audit = {}
    audit_rows = (
        db.query(models.AuditLog)
        .filter(models.AuditLog.entity_type == "product", models.AuditLog.action.in_(("create", "update")))
        .order_by(models.AuditLog.created_at)
        .all()
    )
    for row in audit_rows:
        if not row.changes:
            continue
        try:
            changes = json.loads(row.changes)
        except ValueError:
            continue
        if "cost_price" not in changes:
            continue
        try:
            old, new = changes["cost_price"]
        except (TypeError, ValueError):
            continue
        latest_audit[row.entity_id] = {
            "ts": row.created_at,
            "old": Decimal(str(old)) if old not in (None, "") else Decimal("0"),
            "new": Decimal(str(new)) if new not in (None, "") else Decimal("0"),
            "source": "Manual edit",
        }
    return latest_purchase, latest_audit


def _last_change_for(product_id, latest_purchase, latest_audit):
    candidates = [c for c in (latest_purchase.get(product_id), latest_audit.get(product_id)) if c]
    return max(candidates, key=lambda c: c["ts"]) if candidates else None


def _cost_rows(db: Session, q: str = "", page: int = 1, page_size: int = INV_COST_PAGE_SIZE):
    """Per-product current cost plus its most recent cost change, one page
    at a time. Returns (rows, total, pages, page, changed_count) — changed_count
    is the count across the WHOLE filtered set, not just the current page."""
    query = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if q:
        query = query.filter(multi_word_ilike(models.Product.name, q))

    total = query.count()
    pages = max((total + page_size - 1) // page_size, 1)
    page = min(max(page, 1), pages)

    products = (
        query.order_by(models.Product.name)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    latest_purchase, latest_audit = _cost_change_lookup(db)

    rows = []
    for p in products:
        last_change = _last_change_for(p.id, latest_purchase, latest_audit)
        rows.append({
            "product": p,
            "category": p.category.name if p.category else "Uncategorized",
            "unit_name": p.unit_type.name if p.unit_type else "unit",
            "last_change": last_change,
        })

    filtered_ids = [pid for (pid,) in query.with_entities(models.Product.id).all()]
    changed_count = sum(
        1 for pid in filtered_ids
        if (lc := _last_change_for(pid, latest_purchase, latest_audit)) and lc["new"] != lc["old"]
    )

    return rows, total, pages, page, changed_count


def _sales_by_product(db: Session, period_start: date, period_end: date):
    """Per-product sales in the window: units sold, revenue and gross profit.
    Uses 'sale' lines only (same basis as the Dashboard's top-sellers), so a
    product's movement here reads as gross demand, not net-of-returns."""
    cogs_expr = models.SaleLine.qty * models.SaleLine.unit_factor * models.SaleLine.unit_cost
    rows = (
        db.query(
            models.SaleLine.product_name,
            func.coalesce(func.sum(models.SaleLine.qty), 0).label("qty"),
            func.coalesce(func.sum(models.SaleLine.line_total), 0).label("revenue"),
            func.coalesce(func.sum(models.SaleLine.line_total - cogs_expr), 0).label("profit"),
        )
        .join(models.Sale, models.SaleLine.sale_id == models.Sale.id)
        .filter(
            models.Sale.txn_type == "sale",
            models.Sale.is_voided.is_(False),
            _local_date(models.Sale.created_at).between(period_start, period_end),
        )
        .group_by(models.SaleLine.product_name)
        .all()
    )
    out = [
        {
            "name": r.product_name,
            "qty": Decimal(str(r.qty or 0)),
            "revenue": Decimal(str(r.revenue or 0)),
            "profit": Decimal(str(r.profit or 0)),
        }
        for r in rows
    ]
    out.sort(key=lambda r: r["revenue"], reverse=True)
    return out


@router.get("/reports/sales-by-product", response_class=HTMLResponse)
def sales_by_product(
    request: Request,
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    rows = _sales_by_product(db, period_start, period_end)
    totals = {
        "qty": sum((r["qty"] for r in rows), ZERO),
        "revenue": sum((r["revenue"] for r in rows), ZERO),
        "profit": sum((r["profit"] for r in rows), ZERO),
    }
    return templates.TemplateResponse(
        "reports/sales_by_product.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "days": days, "date_from": date_from, "date_to": date_to,
            "period_start": period_start, "period_end": period_end, "custom": custom,
            "rows": rows, "totals": totals,
        },
    )


@router.get("/reports/sales-by-product/export")
def export_sales_by_product(
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, _ = _resolve_period(days, date_from, date_to)
    rows = _sales_by_product(db, period_start, period_end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales by Product"
    headers = ["Product", "Units Sold", "Revenue", "Gross Profit"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in rows:
        ws.append([r["name"], float(r["qty"]), float(r["revenue"]), float(r["profit"])])
    widths = [32, 14, 16, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"sales_by_product_{period_start.isoformat()}_{period_end.isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _price_override_rows(db: Session, period_start: date, period_end: date):
    """Sale lines charged at a different price than what the product is
    CURRENTLY listed at for that same price tier (fixed/markup/margin) —
    either a cashier typed a lower number to give a customer a break, or a
    higher one (a special/rush order, a small item priced up to round the
    total, etc). Neither shows up anywhere else: the receipt just shows
    whatever price was typed, no flag either way.

    Compared against today's set price, not a historical snapshot (none is
    kept per line) — so a genuine price change since the sale can make an
    old, perfectly correct sale look flagged here. Reading from the most
    recent sale downward keeps that mostly a non-issue in practice."""
    rows = (
        db.query(models.SaleLine, models.Sale, models.Product)
        .join(models.Sale, models.SaleLine.sale_id == models.Sale.id)
        .join(models.Product, models.SaleLine.product_id == models.Product.id)
        .filter(
            models.Sale.txn_type == "sale",
            models.Sale.is_voided.is_(False),
            models.SaleLine.qty > 0,
            _local_date(models.Sale.created_at).between(period_start, period_end),
        )
        .order_by(models.Sale.created_at.desc())
        .all()
    )
    out = []
    for line, sale, product in rows:
        tier = line.price_tier or "fixed"
        if tier == "markup":
            benchmark = Decimal(str(product.markup_price or 0))
        elif tier == "margin":
            benchmark = Decimal(str(product.margin_price or 0))
        else:
            benchmark = Decimal(str(product.selling_price or 0))
        if benchmark <= 0:
            continue  # nothing set for this tier to compare against
        unit_price = Decimal(str(line.unit_price or 0))
        diff = unit_price - benchmark  # >0 = charged more than list, <0 = charged less
        if abs(diff) <= Decimal("0.01"):  # a one-centavo gap is rounding, not an override
            continue
        qty = Decimal(str(line.qty or 0))
        out.append({
            "sale": sale, "product": product, "tier": tier,
            "benchmark": benchmark, "unit_price": unit_price, "qty": qty,
            "direction": "over" if diff > 0 else "under",
            "diff_per_unit": abs(diff), "diff_total": abs(diff) * qty,
        })
    out.sort(key=lambda r: r["diff_total"], reverse=True)
    return out


@router.get("/reports/price-overrides", response_class=HTMLResponse)
def price_overrides(
    request: Request,
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    direction: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    all_rows = _price_override_rows(db, period_start, period_end)
    direction = direction if direction in ("under", "over") else ""
    rows = [r for r in all_rows if not direction or r["direction"] == direction]
    total_under = sum((r["diff_total"] for r in all_rows if r["direction"] == "under"), ZERO)
    total_over = sum((r["diff_total"] for r in all_rows if r["direction"] == "over"), ZERO)
    return templates.TemplateResponse(
        "reports/price_overrides.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "days": days, "date_from": date_from, "date_to": date_to,
            "period_start": period_start, "period_end": period_end, "custom": custom,
            "direction": direction, "rows": rows,
            "under_count": sum(1 for r in all_rows if r["direction"] == "under"),
            "over_count": sum(1 for r in all_rows if r["direction"] == "over"),
            "total_under": total_under, "total_over": total_over,
        },
    )


@router.get("/reports/price-overrides/export")
def export_price_overrides(
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    direction: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, _ = _resolve_period(days, date_from, date_to)
    direction = direction if direction in ("under", "over") else ""
    rows = _price_override_rows(db, period_start, period_end)
    if direction:
        rows = [r for r in rows if r["direction"] == direction]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Overrides"
    headers = ["Invoice #", "Date", "Cashier", "Product", "Direction", "Qty", "List Price", "Sold At", "Diff/Unit", "Total Diff"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in rows:
        sale = r["sale"]
        ws.append([
            sale.invoice_no,
            sale.created_at.strftime("%Y-%m-%d %H:%M") if sale.created_at else "",
            (sale.cashier.full_name or sale.cashier.username) if sale.cashier else "",
            r["product"].name, "Charged more" if r["direction"] == "over" else "Charged less",
            float(r["qty"]), float(r["benchmark"]), float(r["unit_price"]),
            float(r["diff_per_unit"]), float(r["diff_total"]),
        ])
    widths = [14, 16, 16, 32, 14, 10, 12, 12, 12, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"price_overrides_{period_start.isoformat()}_{period_end.isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _sales_by_unit(db: Session, period_start: date, period_end: date):
    """Which selling unit actually moves for each product — e.g. is GRAVEL
    going out as whole FORWARD loads or mostly as Elf 1/4?

    Grouped by product + unit name. `base_qty` restates every unit in the
    product's base unit so volumes are comparable across rows (2 FORWARD and
    12 Elf are not the same amount of gravel); `share` is that row's cut of
    its own product's revenue, which is what makes one unit "the best seller".
    Sale lines only, same gross-demand basis as _sales_by_product.
    """
    rows = (
        db.query(
            models.SaleLine.product_name,
            models.SaleLine.unit_name,
            func.coalesce(func.sum(models.SaleLine.qty), 0).label("qty"),
            func.coalesce(func.sum(models.SaleLine.qty * models.SaleLine.unit_factor), 0).label("base_qty"),
            func.coalesce(func.sum(models.SaleLine.line_total), 0).label("revenue"),
            func.count(models.SaleLine.id).label("times"),
        )
        .join(models.Sale, models.SaleLine.sale_id == models.Sale.id)
        .filter(
            models.Sale.txn_type == "sale",
            models.Sale.is_voided.is_(False),
            models.SaleLine.qty > 0,
            _local_date(models.Sale.created_at).between(period_start, period_end),
        )
        .group_by(models.SaleLine.product_name, models.SaleLine.unit_name)
        .all()
    )

    by_product = {}
    for r in rows:
        by_product.setdefault(r.product_name, []).append({
            "unit": r.unit_name or "—",
            "qty": Decimal(str(r.qty or 0)),
            "base_qty": Decimal(str(r.base_qty or 0)),
            "revenue": Decimal(str(r.revenue or 0)),
            "times": int(r.times or 0),
        })

    groups = []
    for name, units in by_product.items():
        total_rev = sum((u["revenue"] for u in units), ZERO)
        for u in units:
            u["share"] = float(u["revenue"] / total_rev * 100) if total_rev > 0 else 0.0
        units.sort(key=lambda u: u["revenue"], reverse=True)
        for i, u in enumerate(units):
            u["is_top"] = i == 0 and u["revenue"] > 0
        groups.append({
            "name": name,
            "units": units,
            "revenue": total_rev,
            "base_qty": sum((u["base_qty"] for u in units), ZERO),
            "times": sum(u["times"] for u in units),
        })
    groups.sort(key=lambda g: g["revenue"], reverse=True)
    return groups


@router.get("/reports/sales-by-unit", response_class=HTMLResponse)
def sales_by_unit(
    request: Request,
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    groups = _sales_by_unit(db, period_start, period_end)
    today = _today()
    return templates.TemplateResponse(
        "reports/sales_by_unit.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "days": days, "date_from": date_from, "date_to": date_to,
            "period_start": period_start, "period_end": period_end, "custom": custom,
            "groups": groups,
            "today": today,
            "month_start": today.replace(day=1),
            "this_month": custom and date_from == today.replace(day=1).isoformat() and date_to == today.isoformat(),
        },
    )


@router.get("/reports/sales-by-unit/export")
def export_sales_by_unit(
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, _ = _resolve_period(days, date_from, date_to)
    groups = _sales_by_unit(db, period_start, period_end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales by Unit"
    ws.append(["Product", "Sold As", "Qty Sold", "In Base Units", "Times Sold", "Revenue", "% of Product", "Best Seller"])
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for g in groups:
        for u in g["units"]:
            ws.append([
                g["name"], u["unit"], float(u["qty"]), float(u["base_qty"]),
                u["times"], float(u["revenue"]), round(u["share"], 1),
                "Yes" if u["is_top"] else "",
            ])
    widths = [28, 16, 12, 16, 12, 16, 14, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"sales_by_unit_{period_start.isoformat()}_{period_end.isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _low_margin_rows(db: Session, threshold: float):
    """Same rule as the 'thin margin' Notification: profitable products
    (selling above cost) whose true margin still falls under the shop's
    target. Products at/below cost are a harder problem — they're already
    surfaced separately — so they're excluded here."""
    products = (
        db.query(models.Product)
        .filter(
            models.Product.is_active.is_(True),
            models.Product.cost_price > 0,
            models.Product.selling_price > models.Product.cost_price,
        )
        .order_by(models.Product.name)
        .all()
    )
    rows = []
    for p in products:
        price = Decimal(str(p.selling_price or 0))
        cost = Decimal(str(p.cost_price or 0))
        margin_pct = float((price - cost) / price * 100) if price > 0 else 0.0
        if margin_pct < threshold:
            rows.append({
                "product": p,
                "cost": cost,
                "price": price,
                "margin_pct": margin_pct,
                "gap_pct": threshold - margin_pct,
            })
    rows.sort(key=lambda r: r["margin_pct"])
    return rows


@router.get("/reports/low-margin", response_class=HTMLResponse)
def low_margin_report(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    threshold = settings_store.min_margin_pct()
    rows = _low_margin_rows(db, threshold) if threshold is not None else []

    return templates.TemplateResponse(
        "reports/low_margin.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "rows": rows, "threshold": threshold,
        },
    )


@router.get("/reports/low-margin/export")
def export_low_margin(db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    threshold = settings_store.min_margin_pct()
    rows = _low_margin_rows(db, threshold) if threshold is not None else []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Low Margin"
    ws.append([f"Products below {threshold:g}% margin target" if threshold is not None else "No margin target set"])
    ws.append([])
    headers = ["Product", "Cost", "Selling Price", "Margin %", "Gap to Target %"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in rows:
        ws.append([r["product"].name, float(r["cost"]), float(r["price"]), round(r["margin_pct"], 1), round(r["gap_pct"], 1)])
    widths = [30, 14, 14, 12, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="low_margin_{_today().isoformat()}.xlsx"'},
    )


@router.get("/reports/inventory-valuation", response_class=HTMLResponse)
def inventory_valuation(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    rows = _valuation_rows(db)
    total_cost = sum((r["cost_value"] for r in rows), ZERO)
    total_retail = sum((r["retail_value"] for r in rows), ZERO)

    by_cat = {}
    for r in rows:
        c = by_cat.setdefault(r["category"], {"category": r["category"], "cost_value": ZERO, "retail_value": ZERO, "count": 0})
        c["cost_value"] += r["cost_value"]
        c["retail_value"] += r["retail_value"]
        c["count"] += 1
    by_category = sorted(by_cat.values(), key=lambda r: r["cost_value"], reverse=True)

    return templates.TemplateResponse(
        "reports/inventory_valuation.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "rows": rows, "by_category": by_category,
            "total_cost": total_cost, "total_retail": total_retail,
            "today": _today(),
        },
    )


@router.get("/reports/inventory-valuation/export")
def export_inventory_valuation(db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    rows = _valuation_rows(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Valuation"
    headers = ["Product", "Category", "Qty on Hand", "Cost Price", "Cost Value", "Selling Price", "Retail Value"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for r in rows:
        p = r["product"]
        ws.append([
            p.name, r["category"], float(r["qty"]),
            float(p.cost_price or 0), float(r["cost_value"]),
            float(p.selling_price or 0), float(r["retail_value"]),
        ])

    widths = [28, 18, 14, 14, 14, 14, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventory_valuation_{_today().isoformat()}.xlsx"'},
    )


@router.get("/reports/inventory-costs", response_class=HTMLResponse)
def inventory_costs(
    request: Request, q: str = "", page: int = 1,
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    rows, total, pages, page, changed_count = _cost_rows(db, q, page)

    return templates.TemplateResponse(
        "reports/inventory_costs.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "rows": rows, "changed_count": changed_count, "q": q,
            "page": page, "pages": pages, "total": total,
        },
    )


INV_PRICING_PAGE_SIZE = 50


def _pricing_rows_for(p: models.Product) -> dict:
    """One product's base pricing plus a row per unit in its conversion
    ladder — each unit carries its own cost (base cost x factor_to_base),
    Fixed price, markup %, margin % and a "true margin" sanity check (the
    Selling Price tab's own math, see pricing.true_margin), since a unit's
    price can be hand-typed and drift from what its stored markup/margin %
    would actually produce."""
    cost = Decimal(str(p.cost_price or 0))
    selling = Decimal(str(p.selling_price or 0))
    units = []
    for u in p.units:
        factor = Decimal(str(u.factor_to_base or 0))
        unit_cost = cost * factor
        unit_price = Decimal(str(u.price or 0))
        units.append({
            "name": u.name,
            "factor": factor,
            "cost": unit_cost,
            "price": unit_price,
            "markup_pct": Decimal(str(u.markup_pct or 0)),
            "margin_pct": Decimal(str(u.margin_pct or 0)),
            "true_margin": pricing.true_margin(unit_price, unit_cost),
        })
    return {
        "product": p,
        "category": p.category.name if p.category else "Uncategorized",
        "base_unit_name": p.unit_type.name if p.unit_type else "unit",
        "cost": cost,
        "price": selling,
        "markup_pct": Decimal(str(p.markup_pct or 0)),
        "margin_pct": Decimal(str(p.margin_pct or 0)),
        "true_margin": pricing.true_margin(selling, cost),
        "units": units,
    }


def _pricing_query(db: Session, q: str = "", category_id: int = 0, shelf_id: int = 0):
    query = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if q:
        query = query.filter(multi_word_ilike(models.Product.name, q))
    if category_id:
        query = query.filter(models.Product.category_id == category_id)
    if shelf_id:
        query = query.filter(models.Product.shelf_id == shelf_id)
    return query


def _pricing_rows(
    db: Session, q: str = "", category_id: int = 0, shelf_id: int = 0,
    page: int = 1, page_size: int = INV_PRICING_PAGE_SIZE,
):
    query = _pricing_query(db, q, category_id, shelf_id)

    total = query.count()
    pages = max((total + page_size - 1) // page_size, 1)
    page = min(max(page, 1), pages)

    products = (
        query.order_by(models.Product.name)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    rows = [_pricing_rows_for(p) for p in products]
    return rows, total, pages, page


@router.get("/reports/inventory-pricing", response_class=HTMLResponse)
def inventory_pricing(
    request: Request, q: str = "", category_id: int = 0, shelf_id: int = 0, page: int = 1,
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    rows, total, pages, page = _pricing_rows(db, q, category_id, shelf_id, page)

    # Quick-filter pills: only categories/shelves actually in use, with a live
    # count each against the current search (not the pill's own filter), so
    # switching pills reflects what's really there — same idea as Inventory's.
    base_for_counts = db.query(models.Product).filter(models.Product.is_active.is_(True))
    if q:
        base_for_counts = base_for_counts.filter(multi_word_ilike(models.Product.name, q))
    cat_counts = dict(
        base_for_counts.filter(models.Product.category_id.isnot(None))
        .with_entities(models.Product.category_id, func.count(models.Product.id))
        .group_by(models.Product.category_id)
        .all()
    )
    categories = (
        db.query(models.Category).filter(models.Category.id.in_(cat_counts.keys())).order_by(models.Category.name).all()
    ) if cat_counts else []
    shelf_counts = dict(
        base_for_counts.filter(models.Product.shelf_id.isnot(None))
        .with_entities(models.Product.shelf_id, func.count(models.Product.id))
        .group_by(models.Product.shelf_id)
        .all()
    )
    shelves = (
        db.query(models.Shelf).filter(models.Shelf.id.in_(shelf_counts.keys())).order_by(models.Shelf.name).all()
    ) if shelf_counts else []

    return templates.TemplateResponse(
        "reports/inventory_pricing.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "rows": rows, "q": q, "page": page, "pages": pages, "total": total,
            "category_id": category_id, "categories": categories, "cat_counts": cat_counts,
            "shelf_id": shelf_id, "shelves": shelves, "shelf_counts": shelf_counts,
        },
    )


@router.get("/reports/inventory-pricing/export")
def export_inventory_pricing(
    q: str = "", category_id: int = 0, shelf_id: int = 0,
    db: Session = Depends(get_db), user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    products = _pricing_query(db, q, category_id, shelf_id).order_by(models.Product.name).all()
    rows = [_pricing_rows_for(p) for p in products]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Pricing"
    headers = ["Product", "Category", "Unit", "Cost", "Selling Price", "Markup %", "Margin %", "True Margin %"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for r in rows:
        p = r["product"]
        ws.append([
            p.name, r["category"], r["base_unit_name"] + " (base)",
            float(r["cost"]), float(r["price"]), float(r["markup_pct"]), float(r["margin_pct"]), float(r["true_margin"]),
        ])
        for u in r["units"]:
            ws.append([
                p.name, r["category"], u["name"],
                float(u["cost"]), float(u["price"]), float(u["markup_pct"]), float(u["margin_pct"]), float(u["true_margin"]),
            ])

    widths = [28, 18, 16, 12, 14, 12, 12, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventory_pricing_{_today().isoformat()}.xlsx"'},
    )


def _inventory_adjustment_rows(db: Session, period_start: date, period_end: date):
    """Every manual stock edit or completed stock count in the window,
    valued at the product's cost at the time — negative value = shrinkage
    (loss), positive = a find (gain)."""
    rows = (
        db.query(models.StockMovement, models.Product)
        .join(models.Product, models.StockMovement.product_id == models.Product.id)
        .filter(
            models.StockMovement.reason.in_(("adjustment", "stock_count")),
            _local_date(models.StockMovement.created_at).between(period_start, period_end),
        )
        .order_by(models.StockMovement.created_at.desc())
        .all()
    )
    out = []
    for mv, product in rows:
        out.append({
            "created_at": mv.created_at,
            "product_name": product.name,
            "reason": "Stock count" if mv.reason == "stock_count" else "Manual edit",
            "note": mv.note or "—",
            "ref": mv.ref or "—",
            "qty_base": Decimal(str(mv.qty_base or 0)),
            "unit_cost": Decimal(str(mv.unit_cost or 0)),
            "value": Decimal(str(mv.value or 0)),
        })
    return out


@router.get("/reports/inventory-adjustments", response_class=HTMLResponse)
def inventory_adjustments(
    request: Request,
    days: int = 30,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    period_start, period_end, custom = _resolve_period(days, date_from, date_to)
    rows = _inventory_adjustment_rows(db, period_start, period_end)
    loss_total = sum((r["value"] for r in rows if r["value"] < 0), ZERO)
    gain_total = sum((r["value"] for r in rows if r["value"] > 0), ZERO)
    net_total = loss_total + gain_total

    today = _today()
    month_start = today.replace(day=1)
    this_month = custom and period_start == month_start and period_end == today

    return templates.TemplateResponse(
        "reports/inventory_adjustments.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "days": days, "date_from": date_from, "date_to": date_to,
            "period_start": period_start, "period_end": period_end, "custom": custom,
            "month_start": month_start, "today": today, "this_month": this_month,
            "rows": rows, "loss_total": loss_total, "gain_total": gain_total, "net_total": net_total,
        },
    )


@router.get("/reports/month-end-rollover", response_class=HTMLResponse)
def month_end_rollover_history(
    request: Request,
    q: str = "",
    period: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """History of month-end rollover: for each product, exactly how much
    Stocks Qty got folded into Actual Beginning, and when — so "what number
    got added to this item's beginning stock in September" has a real
    answer instead of just one summary line in the Activity Log."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    q = (q or "").strip()
    periods = [
        p[0] for p in db.query(models.MonthEndRolloverLine.period)
        .distinct().order_by(models.MonthEndRolloverLine.period.desc()).all()
    ]

    query = db.query(models.MonthEndRolloverLine)
    if q:
        query = query.filter(models.MonthEndRolloverLine.product_name.ilike(f"%{q}%"))
    if period:
        query = query.filter(models.MonthEndRolloverLine.period == period)
    rows = query.order_by(
        models.MonthEndRolloverLine.period.desc(),
        models.MonthEndRolloverLine.product_name,
    ).all()

    return templates.TemplateResponse(
        "reports/month_end_rollover.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "q": q, "period": period, "periods": periods, "rows": rows,
        },
    )


def _week_bounds(today: date) -> tuple[date, date]:
    """Monday..Sunday of the week containing `today`."""
    monday = today - timedelta(days=today.weekday())
    return monday, monday + timedelta(days=6)


def _weekly_summary_data(db: Session, period_start: date, period_end: date):
    """Everything for one page: the same P&L numbers as the Dashboard, plus
    purchasing, credit collections and a low-stock snapshot — the batch of
    numbers an owner who only checks in once a week actually wants together."""
    pl = _pl_data(db, period_start, period_end)

    sales_count = (
        db.query(func.count(models.Sale.id))
        .filter(
            models.Sale.txn_type == "sale",
            models.Sale.is_voided.is_(False),
            _local_date(models.Sale.created_at).between(period_start, period_end),
        )
        .scalar()
    ) or 0
    refunds_total = (
        db.query(func.coalesce(func.sum(-models.Sale.total), 0))
        .filter(models.Sale.txn_type == "refund", _local_date(models.Sale.created_at).between(period_start, period_end))
        .scalar()
    )
    refunds_total = Decimal(str(refunds_total or 0))

    purchases_total = (
        db.query(func.coalesce(func.sum(models.Purchase.total), 0))
        .filter(
            models.Purchase.txn_type == "receive", models.Purchase.status != "cancelled",
            _local_date(models.Purchase.confirmed_at).between(period_start, period_end),
        )
        .scalar()
    )
    purchases_total = Decimal(str(purchases_total or 0))
    purchases_count = (
        db.query(func.count(models.Purchase.id))
        .filter(
            models.Purchase.txn_type == "receive", models.Purchase.status != "cancelled",
            _local_date(models.Purchase.confirmed_at).between(period_start, period_end),
        )
        .scalar()
    ) or 0

    credits_collected = (
        db.query(func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
        .filter(_local_date(models.ReceivableSettlement.created_at).between(period_start, period_end))
        .scalar()
    )
    credits_collected = Decimal(str(credits_collected or 0))

    expenses_paid = (
        db.query(func.coalesce(func.sum(models.Expense.amount), 0))
        .filter(models.Expense.is_voided.is_(False), models.Expense.expense_date.between(period_start, period_end))
        .scalar()
    )
    expenses_paid = Decimal(str(expenses_paid or 0))

    top_products = _sales_by_product(db, period_start, period_end)[:8]

    low_stock = (
        db.query(models.Product)
        .filter(models.Product.is_active.is_(True), low_stock_expr(settings_store.default_low_stock_pct()))
        .order_by(models.Product.name)
        .all()
    )

    return {
        **pl,
        "sales_count": sales_count,
        "refunds_total": refunds_total,
        "purchases_total": purchases_total,
        "purchases_count": purchases_count,
        "credits_collected": credits_collected,
        "expenses_paid": expenses_paid,
        "top_products": top_products,
        "low_stock": low_stock,
    }


@router.get("/reports/weekly-summary", response_class=HTMLResponse)
def weekly_summary(
    request: Request,
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)

    today = _today()
    df, dt = _parse_date(date_from), _parse_date(date_to)
    if df and dt:
        if dt > today:
            dt = today
        if df > dt:
            df, dt = dt, df
        period_start, period_end = df, dt
    else:
        period_start, period_end = _week_bounds(today)

    data = _weekly_summary_data(db, period_start, period_end)

    return templates.TemplateResponse(
        "reports/weekly_summary.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "period_start": period_start, "period_end": period_end,
            "date_from": date_from, "date_to": date_to,
            "today": today, **data,
        },
    )
