"""Customer accounts (name, TIN, address) and receivable helper."""
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Request, status
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from . import audit, models, settings_store
from .database import get_db
from .deps import get_current_user, is_staff, safe_back_url
from .templating import templates

router = APIRouter()

PAGE_SIZE = 15


def get_or_create_customer(db: Session, name: str):
    name = (name or "").strip()
    if not name:
        return None
    existing = (
        db.query(models.Customer)
        .filter(func.lower(models.Customer.name) == name.lower())
        .first()
    )
    if existing:
        return existing
    cust = models.Customer(name=name)
    db.add(cust)
    db.flush()
    return cust


@router.post("/customers/quick")
async def quick_customer(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Create (or update) a customer straight from POS, without leaving the
    sale — same idea as /suppliers/quick on the Purchases form. Unlike that
    one, an existing match by name gets its details UPDATED here rather than
    left alone: the whole point of this popup is "I already have this
    walk-in typed in, let me also attach their TIN/address/terms now,"
    which only makes sense if a second save on the same name actually
    changes something."""
    if not user:
        return JSONResponse({"ok": False, "error": "unauthorized"}, status_code=401)
    if not is_staff(user):
        return JSONResponse({"ok": False, "error": "forbidden"}, status_code=403)

    data = await request.json()
    name = (data.get("name") or "").strip()
    if not name:
        return JSONResponse({"ok": False, "error": "Customer name is required."}, status_code=400)

    tin = (data.get("tin") or "").strip() or None
    address = (data.get("address") or "").strip() or None
    raw_days = (data.get("credit_days") or "").strip() if isinstance(data.get("credit_days"), str) else data.get("credit_days")
    try:
        credit_days = int(raw_days) if raw_days not in (None, "") else None
    except (TypeError, ValueError):
        credit_days = None

    customer = db.query(models.Customer).filter(func.lower(models.Customer.name) == name.lower()).first()
    if customer:
        if tin:
            customer.tin = tin
        if address:
            customer.address = address
        if credit_days is not None:
            customer.credit_days = credit_days
    else:
        customer = models.Customer(
            name=name, tin=tin, address=address,
            credit_days=credit_days if credit_days is not None else 15,
        )
        db.add(customer)
    db.commit()
    db.refresh(customer)
    return {"ok": True, "customer": {"id": customer.id, "name": customer.name}}


@router.get("/customers/search")
def search_customers(q: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    """JSON autocomplete for the POS customer field."""
    if not user:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    q = (q or "").strip()
    query = db.query(models.Customer).filter(models.Customer.is_active.is_(True))
    if q:
        query = query.filter(models.Customer.name.ilike(f"%{q}%"))
    customers = query.order_by(models.Customer.name).limit(20).all()
    return {
        "customers": [
            {"id": c.id, "name": c.name, "tin": c.tin or "", "address": c.address or ""}
            for c in customers
        ]
    }


@router.get("/customers", response_class=HTMLResponse)
def list_customers(
    request: Request,
    q: str = "",
    page: int = 1,
    cust_msg: str = "",
    cust_error: str = "",
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=302)
    q = (q or "").strip()
    page = max(page, 1)
    query = db.query(models.Customer).filter(models.Customer.is_active.is_(True))
    if q:
        query = query.filter(models.Customer.name.ilike(f"%{q}%"))
    total = query.count()
    pages = max((total + PAGE_SIZE - 1) // PAGE_SIZE, 1)
    page = min(page, pages)
    customers = (
        query.order_by(models.Customer.name)
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )

    # Site-wide receivables snapshot — same basis as the Credits page. Money
    # figures stay admin-only, same as Credits itself.
    total_receivable, customers_owing = Decimal("0"), 0
    if is_staff(user):
        settled_sub = (
            db.query(
                models.ReceivableSettlement.sale_id.label("sid"),
                func.coalesce(func.sum(models.ReceivableSettlement.amount), 0).label("paid"),
            )
            .group_by(models.ReceivableSettlement.sale_id)
            .subquery()
        )
        outstanding_expr = models.Sale.receivable_amount - func.coalesce(settled_sub.c.paid, 0)
        total_receivable, customers_owing = (
            db.query(func.coalesce(func.sum(outstanding_expr), 0), func.count(func.distinct(models.Sale.customer_id)))
            .outerjoin(settled_sub, settled_sub.c.sid == models.Sale.id)
            .filter(models.Sale.receivable_amount > 0, outstanding_expr > 0)
            .one()
        )
        total_receivable = Decimal(str(total_receivable or 0))

    return templates.TemplateResponse(
        "customers/list.html",
        {
            "request": request,
            "app_name": request.app.title,
            "user": user,
            "customers": customers,
            "q": q,
            "page": page,
            "pages": pages,
            "total": total,
            "total_receivable": total_receivable,
            "customers_owing": customers_owing or 0,
            "cust_msg": cust_msg,
            "cust_error": cust_error,
            "is_staff_user": is_staff(user),
        },
    )


@router.get("/customers/new", response_class=HTMLResponse)
def new_customer(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse(
        "customers/form.html",
        {"request": request, "app_name": request.app.title, "user": user, "customer": None, "error": None,
         "back": "/customers"},
    )


@router.get("/customers/{customer_id:int}/edit", response_class=HTMLResponse)
def edit_customer(customer_id: int, request: Request, back: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=302)
    return templates.TemplateResponse(
        "customers/form.html",
        {"request": request, "app_name": request.app.title, "user": user, "customer": customer, "error": None,
         # The filtered list they came from, so a search survives an edit.
         "back": safe_back_url(back, "/customers")},
    )


@router.get("/customers/{customer_id:int}/history", response_class=HTMLResponse)
def customer_history(customer_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=302)

    sales = (
        db.query(models.Sale)
        .filter(models.Sale.customer_id == customer_id, models.Sale.is_voided.is_(False))
        .order_by(models.Sale.id.desc())
        .all()
    )

    # settlements collected per sale, to show each sale's paid/credit status
    sale_ids = [s.id for s in sales]
    settled = {}
    if sale_ids:
        rows = (
            db.query(models.ReceivableSettlement.sale_id, func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
            .filter(models.ReceivableSettlement.sale_id.in_(sale_ids))
            .group_by(models.ReceivableSettlement.sale_id)
            .all()
        )
        settled = {sid: Decimal(amt) for sid, amt in rows}

    rows = []
    total_spent = Decimal("0")
    total_out = Decimal("0")
    for s in sales:
        outstanding = (s.receivable_amount or Decimal("0")) - settled.get(s.id, Decimal("0"))
        rows.append({"sale": s, "outstanding": outstanding})
        if s.txn_type == "sale":
            total_spent += (s.total or Decimal("0"))
        if outstanding > 0:
            total_out += outstanding

    return templates.TemplateResponse(
        "customers/history.html",
        {
            "request": request, "app_name": request.app.title, "user": user,
            "customer": customer, "rows": rows, "count": len(rows),
            "total_spent": total_spent, "total_out": total_out,
        },
    )


@router.get("/customers/{customer_id:int}/history/export.xlsx")
def export_customer_history_excel(customer_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Excel export of one customer's full transaction history — every
    invoice/refund/exchange row exactly as shown on the history page."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=302)

    sales = (
        db.query(models.Sale)
        .filter(models.Sale.customer_id == customer_id, models.Sale.is_voided.is_(False))
        .order_by(models.Sale.id.desc())
        .all()
    )
    sale_ids = [s.id for s in sales]
    settled = {}
    if sale_ids:
        rows = (
            db.query(models.ReceivableSettlement.sale_id, func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
            .filter(models.ReceivableSettlement.sale_id.in_(sale_ids))
            .group_by(models.ReceivableSettlement.sale_id)
            .all()
        )
        settled = {sid: Decimal(amt) for sid, amt in rows}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase History"
    headers = ["Invoice #", "Type", "Date", "Payment", "Total", "Status"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F6FEB")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for s in sales:
        outstanding = (s.receivable_amount or Decimal("0")) - settled.get(s.id, Decimal("0"))
        status_label = f"Credit {outstanding:,.2f}" if outstanding > 0 else "Paid"
        ws.append([
            s.invoice_no,
            s.txn_type.capitalize() if s.txn_type else "",
            s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
            s.payment_method or "",
            float(s.total or 0),
            status_label,
        ])

    widths = [14, 12, 18, 16, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c for c in customer.name if c.isalnum() or c in (" ", "_", "-")).strip() or "customer"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_history.xlsx"'},
    )


@router.get("/customers/{customer_id:int}/history/export.pdf")
def export_customer_history_pdf(customer_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """PDF version of the full Purchase History export — same rows as the
    Excel export, letterhead-styled to match the other simple PDFs."""
    if not user:
        return RedirectResponse("/login", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=302)

    sales = (
        db.query(models.Sale)
        .filter(models.Sale.customer_id == customer_id, models.Sale.is_voided.is_(False))
        .order_by(models.Sale.id.desc())
        .all()
    )
    sale_ids = [s.id for s in sales]
    settled = {}
    if sale_ids:
        srows = (
            db.query(models.ReceivableSettlement.sale_id, func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
            .filter(models.ReceivableSettlement.sale_id.in_(sale_ids))
            .group_by(models.ReceivableSettlement.sale_id)
            .all()
        )
        settled = {sid: Decimal(amt) for sid, amt in srows}

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from .pdf_utils import letterhead

    biz = settings_store.get_all(db)
    party_lines = [customer.name]
    if customer.tin:
        party_lines.append(f"TIN {customer.tin}")
    if customer.address:
        party_lines.append(customer.address)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=18 * mm, rightMargin=18 * mm)
    styles = getSampleStyleSheet()
    elements = letterhead(biz, "Purchase History", [f"{len(sales)} transaction(s)"], "Customer", party_lines)

    table_data = [["Invoice #", "Type", "Date", "Payment", "Total", "Status"]]
    for s in sales:
        outstanding = (s.receivable_amount or Decimal("0")) - settled.get(s.id, Decimal("0"))
        status_label = f"Credit {outstanding:,.2f}" if outstanding > 0 else "Paid"
        table_data.append([
            s.invoice_no,
            s.txn_type.capitalize() if s.txn_type else "",
            s.created_at.strftime("%b %d, %Y") if s.created_at else "-",
            s.payment_method or "-",
            f"{s.total or 0:,.2f}",
            status_label,
        ])

    table = Table(table_data, colWidths=[65, 65, 80, 100, 70, 90], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F6FEB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    safe_name = "".join(c for c in customer.name if c.isalnum() or c in (" ", "_", "-")).strip() or "customer"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_history.pdf"'},
    )


@router.post("/customers")
async def create_customer(request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    form = await request.form()
    back = safe_back_url(form.get("back") or "", "/customers")
    name = (form.get("name") or "").strip()
    if not name:
        return templates.TemplateResponse(
            "customers/form.html",
            {"request": request, "app_name": request.app.title, "user": user, "customer": None,
             "error": "Customer name is required.", "back": back},
        )
    cust = models.Customer(
        name=name,
        tin=(form.get("tin") or "").strip() or None,
        address=(form.get("address") or "").strip() or None,
        credit_days=int(form.get("credit_days") or 15),
    )
    db.add(cust)
    db.commit()
    return RedirectResponse(back, status_code=status.HTTP_302_FOUND)


@router.post("/customers/{customer_id:int}")
async def update_customer(customer_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=302)
    form = await request.form()
    back = safe_back_url(form.get("back") or "", "/customers")
    name = (form.get("name") or "").strip()
    if not name:
        return templates.TemplateResponse(
            "customers/form.html",
            {"request": request, "app_name": request.app.title, "user": user, "customer": customer,
             "error": "Customer name is required.", "back": back},
        )
    customer.name = name
    customer.tin = (form.get("tin") or "").strip() or None
    customer.address = (form.get("address") or "").strip() or None
    try:
        customer.credit_days = int(form.get("credit_days") or 15)
    except (TypeError, ValueError):
        customer.credit_days = 15
    db.commit()
    return RedirectResponse(back, status_code=status.HTTP_302_FOUND)


def _customer_outstanding(db: Session, customer_id: int) -> Decimal:
    """What this customer still owes across all their credit sales."""
    owed = (
        db.query(func.coalesce(func.sum(models.Sale.receivable_amount), 0))
        .filter(models.Sale.customer_id == customer_id)
        .scalar()
    )
    paid = (
        db.query(func.coalesce(func.sum(models.ReceivableSettlement.amount), 0))
        .join(models.Sale, models.ReceivableSettlement.sale_id == models.Sale.id)
        .filter(models.Sale.customer_id == customer_id)
        .scalar()
    )
    return Decimal(str(owed or 0)) - Decimal(str(paid or 0))


def _customer_has_history(db: Session, customer_id: int) -> bool:
    """True if anything real points at this customer. Those foreign keys are
    not ON DELETE CASCADE on purpose — a sale from three months ago must not
    vanish because someone tidied the customer list — so a hard DELETE here
    would either fail outright or quietly break that history."""
    checks = [
        (models.Sale, models.Sale.customer_id),
        (models.Quotation, models.Quotation.customer_id),
        (models.PostDatedCheque, models.PostDatedCheque.customer_id),
    ]
    for model, col in checks:
        if db.query(model.id).filter(col == customer_id).first():
            return True
    return False


@router.post("/customers/{customer_id:int}/delete")
def delete_customer(customer_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Remove a customer from the list.

    Hard-deletes only one that has never been transacted with (a typo, or a
    name created by accident). Anything with real history is ARCHIVED instead
    — hidden from the list and pickers while its sales stay attached and
    attributable, which is the same trade-off products already make.

    Refused outright while they still owe money: hiding an account with an
    open balance is how a debt quietly stops being chased.
    """
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if not customer:
        return RedirectResponse("/customers", status_code=302)

    back = safe_back_url(request.query_params.get("back") or "", "/customers")
    sep = "&" if "?" in back else "?"

    if _customer_outstanding(db, customer_id) > 0:
        return RedirectResponse(f"{back}{sep}cust_error=owing", status_code=status.HTTP_302_FOUND)

    name = customer.name
    if _customer_has_history(db, customer_id):
        customer.is_active = False
        audit.record(
            db, user=user, request=request, action="archive", entity_type="customer",
            entity_id=customer.id, entity_label=name,
            summary=f"Archived customer “{name}” — has sales history, so kept for the record",
        )
        db.commit()
        return RedirectResponse(f"{back}{sep}cust_msg=archived", status_code=status.HTTP_302_FOUND)

    audit.record(
        db, user=user, request=request, action="delete", entity_type="customer",
        entity_id=customer.id, entity_label=name,
        summary=f"Deleted customer “{name}” (never had any activity)",
    )
    db.delete(customer)
    db.commit()
    return RedirectResponse(f"{back}{sep}cust_msg=deleted", status_code=status.HTTP_302_FOUND)


@router.get("/customers/archived", response_class=HTMLResponse)
def archived_customers(request: Request, q: str = "", db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    q = (q or "").strip()
    query = db.query(models.Customer).filter(models.Customer.is_active.is_(False))
    if q:
        query = query.filter(models.Customer.name.ilike(f"%{q}%"))
    customers = query.order_by(models.Customer.name).all()
    return templates.TemplateResponse(
        "customers/archived.html",
        {"request": request, "app_name": request.app.title, "user": user, "customers": customers, "q": q},
    )


@router.post("/customers/{customer_id:int}/restore")
def restore_customer(customer_id: int, request: Request, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=302)
    if not is_staff(user):
        return RedirectResponse("/pos", status_code=302)
    customer = db.get(models.Customer, customer_id)
    if customer:
        customer.is_active = True
        audit.record(
            db, user=user, request=request, action="restore", entity_type="customer",
            entity_id=customer.id, entity_label=customer.name,
            summary=f"Restored customer “{customer.name}” from archive",
        )
        db.commit()
    return RedirectResponse("/customers/archived", status_code=status.HTTP_302_FOUND)
